from email import message
from msilib.schema import RadioButton
from turtle import textinput
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Color, numbers
from openpyxl.worksheet.datavalidation import DataValidation
import os
from struct import pack
import tkinter as tk
from PIL import Image, ImageTk
from tkinter.filedialog import askopenfile
from tkinter import CENTER, filedialog
from tkinter import HORIZONTAL, messagebox
from tkinter import ttk
from tkinter import IntVar
import tkinter.ttk
import time
import string
from tkinter.filedialog import askopenfile

from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re

import pandas as pd
from pandas import read_html
import html5lib

url = 'DIDCheck_DPT.html'



        # print(result_of_string)

# ----------------------------------------------------------------
# excel
try:
    # os.system('TASKKILL /F /IM EXCEL.exe')
    os.remove("Report_review.xlsx")
    wb_report_review = Workbook()
    ws_report_review = wb_report_review.active

    ws_report_review.title = "Report_review"
    wb_report_review.save("Report_review.xlsx")
    os.close("Report_review.xlsx")
except:
    try:
        wb_report_review = Workbook()
        ws_report_review = wb_report_review.active
        ws_report_review.title = "Report_review"
        wb_report_review.save("Report_review.xlsx")
        # os.system('TASKKILL /F /IM EXCEL.exe')
    except OSError:
        print('Failed creating the file')
    else:
        print('File created')


# ------------------------------------------------------------------------------------------------------------------------
def create_DPT_can_load_file():
    try:
        with open("DPT_can_load.xlsx", "r") as file:
            # Print the success message
            print("File is already haved")
        # fd = os.open("RFvalue.xlsx", os.O_RDWR)
        # os.close(fd)
    except OSError:

        wb_convert_DPT = load_workbook('DIA_Questionnaire_LAS_CA_C281D_20220626.xlsx')

        ws_convert_DPT = wb_convert_DPT.active

        ws_convert_DPT = wb_convert_DPT['7-1_DIDList']

        sheet_convert_DPT = wb_convert_DPT.worksheets[0]

        wb_convert_DPT.save("DPT_can_load.xlsx")
        wb_convert_DPT.close()

        # messagebox.showinfo("COMPLETE", "File RFvalue.xlsx has been created in the same folder tool successfully, Please fill all value")
        # print('tao thanh cong')

# def excute_excell_file():
#     wb_read_DPT = load_workbook('DPT_can_load.xlsx')
#     wb_add_dpt_value_to_report_review = load_workbook('Report_review.xlsx')
    
#     ws_read_DPT = wb_read_DPT.active
#     ws_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review.active

#     ws_read_DPT = wb_read_DPT['7-1_DIDList']
#     ws_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review['Report_review']

#     sheet_read_DPT = wb_read_DPT.worksheets[12]
#     sheet_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review.worksheets[0]
    
#     row_count_read_DPT = sheet_read_DPT.max_row    
#     row_count_add_dpt_value_to_report_review = sheet_add_dpt_value_to_report_review.max_row
    
#     k = 4
#     o = 4
#     j = 5
#     count_length_byte = 0
#     count_Value_DPT = 0
#     length_byte = []
#     Value_DPT = []
#     load_second = 0
#     # print(row_read_DPT)
    
#     while k < row_count_read_DPT:
#         for row in range(o, j):
#             for col in range(1, 2):
#                 char = get_column_letter(col)
#                 row_list_DID = ws_read_DPT[char + str(row)].value
#                 # print(row_list_DID)
                
#         for row in range(o, j):
#             for col in range(3, 4):
#                 char = get_column_letter(col)
#                 row_list_length_byte = ws_read_DPT[char + str(row)].value
#                 # print(row_list_length_byte)

#         for row in range(o, j):
#             for col in range(21, 22):
#                 char = get_column_letter(col)
#                 row_list_value = ws_read_DPT[char + str(row)].value
#                 # print(row_list_value)
        
#         if str(row_list_length_byte) != 'None':
#             length_byte = length_byte + [row_list_length_byte]
#             count_length_byte +=  1
        
        
        
#         if str(row_list_value) != 'None':
#             Value_DPT = Value_DPT + [row_list_value]
#             count_Value_DPT +=  1
        
#         o += 1
#         j += 1
#         k += 1
    
    

    

#     for i in range(count_length_byte):
#         length_byte = length_byte + [length_byte[i]]

#     # count_length_byte = int(count_length_byte) * 2

#     # for i in range(count_length_byte):
#     #     print(length_byte[i], end = ' ')
#     #     # print(Value_DPT[i])

#     for i in range(count_Value_DPT):
#         Value_DPT = Value_DPT + [Value_DPT[i]]

#     # count_Value_DPT = int(count_Value_DPT) * 2

#     # for i in range(count_Value_DPT):
#     #     print(Value_DPT[i], end = ' ')
#         # print(Value_DPT[i])
    
#     # print(length_byte[1])





def review_report(ws_report_review, wb_report_review):

    wb_report_review = load_workbook('Report_review.xlsx')
    wb_add_dpt_value_to_report_review = load_workbook('Report_review.xlsx')
    wb_read_DPT = load_workbook('DPT_can_load.xlsx')
    wb_report_review_add_result = load_workbook('Report_review.xlsx')

    ws_report_review = wb_report_review.active
    ws_read_DPT = wb_read_DPT.active
    ws_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review.active
    ws_report_review_add_result = wb_report_review.active

    ws_report_review = wb_report_review['Report_review']
    ws_read_DPT = wb_read_DPT['7-1_DIDList']
    ws_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review['Report_review']
    ws_report_review_add_result = wb_report_review_add_result['Report_review']

    sheet_report_review = wb_report_review.worksheets[0]
    sheet_read_DPT = wb_read_DPT.worksheets[12]
    sheet_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review.worksheets[0]
    sheet_report_review_add_result = wb_report_review_add_result.worksheets[0]


    row_count_report_review = sheet_report_review.max_row
    row_count_read_DPT = sheet_read_DPT.max_row    
    row_count_add_dpt_value_to_report_review = sheet_add_dpt_value_to_report_review.max_row


    noneFill = PatternFill(start_color='00FFFFFF',
                            end_color='00FFFFFF',
                            fill_type='solid'
                            )
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    font_text = Font(name="Calibri", size=14, color='00FFFFFF', bold=True)
    font_text3 = Font(name="Calibri", size=11, color='00FFFFFF', bold=True)
    font_text2 = Font(name="Calibri", size=11, color='000000', bold=False, italic = True)
    alignment = Alignment(horizontal='center', vertical='center')

    
    sheet_report_review.column_dimensions['D'].width = 40
    sheet_report_review.column_dimensions['E'].width = 50
    sheet_report_review.column_dimensions['F'].width = 50
    sheet_report_review.column_dimensions['G'].width = 50
    sheet_report_review.column_dimensions['I'].width = 50
    sheet_report_review.column_dimensions['L'].width = 20

    sheet_report_review.column_dimensions['A'].number_format = numbers.FORMAT_TEXT
    sheet_report_review.column_dimensions['B'].number_format = numbers.FORMAT_TEXT
    sheet_report_review.column_dimensions['C'].number_format = numbers.FORMAT_TEXT
    sheet_report_review.column_dimensions['D'].number_format = numbers.FORMAT_TEXT
    sheet_report_review.column_dimensions['E'].number_format = numbers.FORMAT_TEXT
    sheet_report_review.column_dimensions['F'].number_format = numbers.FORMAT_TEXT
    sheet_report_review.column_dimensions['G'].number_format = numbers.FORMAT_TEXT
    sheet_report_review.column_dimensions['H'].number_format = numbers.FORMAT_TEXT
    sheet_report_review.column_dimensions['I'].number_format = numbers.FORMAT_TEXT
    sheet_report_review.column_dimensions['J'].number_format = numbers.FORMAT_TEXT
    sheet_report_review.column_dimensions['K'].number_format = numbers.FORMAT_TEXT
    sheet_report_review.column_dimensions['L'].number_format = numbers.FORMAT_TEXT
    sheet_report_review.column_dimensions['A'].alignment = alignment
    sheet_report_review.column_dimensions['B'].alignment = alignment
    sheet_report_review.column_dimensions['C'].alignment = alignment
    sheet_report_review.column_dimensions['D'].alignment = alignment
    sheet_report_review.column_dimensions['E'].alignment = alignment
    sheet_report_review.column_dimensions['F'].alignment = alignment
    sheet_report_review.column_dimensions['G'].alignment = alignment
    sheet_report_review.column_dimensions['h'].alignment = alignment
    sheet_report_review.column_dimensions['I'].alignment = alignment
    sheet_report_review.column_dimensions['J'].alignment = alignment
    sheet_report_review.column_dimensions['K'].alignment = alignment
    sheet_report_review.column_dimensions['L'].alignment = alignment

    # for col in range(1, 12):
    #     cell_header = ws_report_review.cell(1, col)
    # # used hex code for red color
    # cell_header.fill = PatternFill(
    #     start_color='000066CC', end_color='000066CC', fill_type="solid")
    # cell_header.border = border
    # cell_header.font = font_text
    # cell_header.alignment = alignment

    k = 5
    o = 4
    j = 5
    count_length_byte = 0
    count_Value_DPT = 0
    count_DID_In_DPT = 0
    length_byte = []
    Value_DPT = []
    DID_In_DPT = []
    load_second = 0
    # print(row_read_DPT)
    
    while k < row_count_read_DPT:
        for row in range(o, j):
            for col in range(1, 2):
                char = get_column_letter(col)
                row_list_DID_In_DPT = ws_read_DPT[char + str(row)].value
                # print(row_list_DID)
                
        for row in range(o, j):
            for col in range(3, 4):
                char = get_column_letter(col)
                row_list_length_byte = ws_read_DPT[char + str(row)].value
                # print(row_list_length_byte)

        for row in range(o, j):
            for col in range(21, 22):
                char = get_column_letter(col)
                row_list_value = ws_read_DPT[char + str(row)].value
                # print(row_list_value)
        
        if str(row_list_DID_In_DPT) != 'None':
            DID_In_DPT = DID_In_DPT + [row_list_DID_In_DPT]
            count_DID_In_DPT +=  1

        if str(row_list_length_byte) != 'None':
            length_byte = length_byte + [row_list_length_byte]
            count_length_byte +=  1
        
        
        if str(row_list_value) != 'None':
            Value_DPT = Value_DPT + [row_list_value]
            count_Value_DPT +=  1
        
        o += 1
        j += 1
        k += 1
    
    

    

    # for i in range(count_length_byte):
    #     length_byte = length_byte + [length_byte[i]]

    # # count_length_byte = int(count_length_byte) * 2

    # # for i in range(count_length_byte):
    # #     print(length_byte[i], end = ' ')
    # #     # print(Value_DPT[i])

    # for i in range(count_Value_DPT):
    #     Value_DPT = Value_DPT + [Value_DPT[i]]

    # count_Value_DPT = int(count_Value_DPT) * 2

    # for i in range(count_Value_DPT):
    #     print(Value_DPT[i], end = ' ')
        # print(Value_DPT[i])


    
    id = 1
    r = 2
    with open(url, 'r') as f:

        ws_report_review.append(['ID' , 'DID','Length Byte RECEIVED','ASCII RECEIVED Value', 'HEX RECEIVED Value','HEX EXPECTED Value', 'ASCII EXPECTED Value','Length Byte EXPECTED', 'DPT Value', 'Length Byte DPT','Result'])
        
        for col in range(1, 12):
            cell_header = ws_report_review.cell(1, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='000066CC', end_color='000066CC', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
            cell_header.alignment = alignment

        result_of_string = f.read()
        
        result_of_string = result_of_string.partition("EXPECTED: 62")[2]

        while len(result_of_string) > 0:
            id += 1
            length_of_EXPECTED = result_of_string.find('<')
            # print(length_of_EXPECTED)
            

            result_of_EXPECTED = result_of_string
            result_of_EXPECTED = result_of_EXPECTED.split('<', 1)[0]

            value_of_EXPECTED = result_of_EXPECTED[4:length_of_EXPECTED]
            DID_of_EXPECTED = result_of_EXPECTED[0:4]

            # print(value_of_EXPECTED)

            try:
                bytes_object_of_EXPECTED = bytes.fromhex(value_of_EXPECTED)
                
            except:
                missing_lenght_byte_of_EXPECTED = value_of_EXPECTED.partition(".{")[2]
                missing_lenght_byte_of_EXPECTED = missing_lenght_byte_of_EXPECTED[:-1]
                # print(missing_lenght_byte_of_EXPECTED)

                # missing_lenght_byte_of_EXPECTED = int(missing_lenght_byte_of_EXPECTED) // 2


                length_of_EXPECTED = int(length_of_EXPECTED) + int(missing_lenght_byte_of_EXPECTED) - 4
                # print(length_of_EXPECTED)

                value_of_EXPECTED = value_of_EXPECTED.split('.', 1)[0]
                # print(value_of_EXPECTED)

                bytes_object_of_EXPECTED = bytes.fromhex(value_of_EXPECTED)
                # print(bytes_object_of_EXPECTED)

                ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")
                # print(ascii_string_of_EXPECTED)
                # Convert to bytes object
            try:
                ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")

            # Convert to ASCII representation
            except:
                # ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")
                # ascii_string_of_EXPECTED = ILLEGAL_CHARACTERS_RE.sub(r'', ascii_string_of_EXPECTED)
                print('khong')
            if str(value_of_EXPECTED) != '':
                length_of_EXPECTED = (length_of_EXPECTED - 4) // 2
                # print('value khong trong '+ str(length_of_EXPECTED))
            else:
                length_of_EXPECTED = (length_of_EXPECTED // 2) - 2
                # print('value trong')

            # print('ASCII Value EXPECTED: ' + ascii_string_of_EXPECTED)
            # print('HEX Value EXPECTED: ' + value_of_EXPECTED)

            result_of_string = result_of_string.partition("RECEIVED: 62")[2] 
            # print(result_of_string)

            ### This is the part that measures the length before '<'
            length_of_RECEIVED = result_of_string.find('<')
            # print(length)
            

            result_of_RECEIVED = result_of_string
            result_of_RECEIVED = result_of_RECEIVED.split('<', 1)[0]

            value_of_RECEIVED = result_of_RECEIVED[4:length_of_RECEIVED]
            DID_of_RECEIVED = result_of_RECEIVED[0:4]

            bytes_object_of_RECEIVED = bytes.fromhex(value_of_RECEIVED)
                # Convert to bytes object

            length_of_RECEIVED = (length_of_RECEIVED - 4) // 2

            try:
                ascii_string_of_RECEIVED = bytes_object_of_RECEIVED.decode("ASCII")
                # print(ascii_string_of_RECEIVED)
                ascii_string_of_RECEIVED = ILLEGAL_CHARACTERS_RE.sub(r'', ascii_string_of_RECEIVED)
                # print(ascii_string_of_RECEIVED)
                
            # Convert to ASCII representation
            except:
                ascii_string_of_RECEIVED = ILLEGAL_CHARACTERS_RE.sub(r'', ascii_string_of_RECEIVED)
                # print(ascii_string_of_RECEIVED)


            result_of_string = result_of_string.partition("EXPECTED: 62")[2]

            try:
                ws_report_review.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(length_of_RECEIVED),str(ascii_string_of_RECEIVED), str(value_of_RECEIVED),str(value_of_EXPECTED), str(ascii_string_of_EXPECTED),str(length_of_EXPECTED),'' , '', ''])
            except:
                ws_report_review.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(length_of_RECEIVED),'', str(value_of_RECEIVED),str(value_of_EXPECTED), '',str(length_of_EXPECTED),'' , '', ''])
            
            r += 1

    data_validation_data = '"PASS, FAIL"'
    
    for row in range(2, row_count_read_DPT):
        data_validation = DataValidation(type='list', formula1 = data_validation_data)
        ws_report_review.add_data_validation(data_validation)
        data_validation.add(ws_report_review['K'+str(row)])
    
    wb_report_review.save("Report_review.xlsx")
    wb_report_review.close()

    l = 2
    h = 0
    
    wb_report_review_add_dptvalue = load_workbook('Report_review.xlsx')
    ws_report_review_add_dptvalue = wb_report_review_add_dptvalue.active
    ws_report_review_add_dptvalue = wb_report_review_add_dptvalue['Report_review']

    sheet_report_review_add_dptvalue = wb_report_review_add_dptvalue.worksheets[0]
    row_count_report_review = sheet_report_review_add_dptvalue.max_row
    
    # row_count_report_review = int(row_count_report_review) // 2
    # print(row_count_report_review)



    while l < row_count_report_review:
        ws_report_review_add_dptvalue['I'+ str(l)] = str(Value_DPT[h])
        ws_report_review_add_dptvalue['J'+ str(l)] = str(length_byte[h])

        l += 1
        # r += 1

        ws_report_review_add_dptvalue['I'+ str(l)] = str(Value_DPT[h])
        ws_report_review_add_dptvalue['J'+ str(l)] = str(length_byte[h])
        
        # r += 1
        l += 1
        h += 1

    wb_report_review_add_dptvalue.save("Report_review.xlsx")
    wb_report_review_add_dptvalue.close()


    wb_report_review_add_result = load_workbook('Report_review.xlsx')
    ws_report_review_add_result = wb_report_review_add_result.active
    ws_report_review_add_result = wb_report_review_add_result['Report_review']

    sheet_report_review_add_result = wb_report_review_add_result.worksheets[0]
    row_count_report_review_add_result = sheet_report_review_add_result.max_row


    z = 2
    m = 2
    n = 3
    v = 2
    while z < row_count_report_review + 1:
        for row in range(m, n):
            for col in range(3, 4):
                char = get_column_letter(col)
                length_of_RECEIVED = ws_report_review_add_result[char + str(row)].value
                # print(length_of_RECEIVED)
                
        for row in range(m, n):
            for col in range(4, 5):
                char = get_column_letter(col)
                ascii_string_of_RECEIVED = ws_report_review_add_result[char + str(row)].value
                # print(ascii_string_of_RECEIVED)

        for row in range(m, n):
            for col in range(5, 6):
                char = get_column_letter(col)
                value_of_RECEIVED = ws_report_review_add_result[char + str(row)].value
                # print(value_of_RECEIVED)
        
        for row in range(m, n):
            for col in range(6, 7):
                char = get_column_letter(col)
                value_of_EXPECTED = ws_report_review_add_result[char + str(row)].value
                # print(value_of_EXPECTED)
        
        for row in range(m, n):
            for col in range(7, 8):
                char = get_column_letter(col)
                ascii_string_of_EXPECTED = ws_report_review_add_result[char + str(row)].value
                # print(ascii_string_of_EXPECTED)

        for row in range(m, n):
            for col in range(8, 9):
                char = get_column_letter(col)
                length_of_EXPECTED = ws_report_review_add_result[char + str(row)].value
                # print(length_of_EXPECTED)
            
        for row in range(m, n):
            for col in range(10, 11):
                char = get_column_letter(col)
                length_of_DID_DPT = ws_report_review_add_result[char + str(row)].value
                # print(length_of_DID_DPT)

        
        


        if str(ascii_string_of_EXPECTED) == str(ascii_string_of_RECEIVED) or str(value_of_EXPECTED) == str(value_of_RECEIVED) and str(length_of_EXPECTED) == str(length_of_RECEIVED) and str(length_of_EXPECTED) == str(length_of_DID_DPT):
            result = 'PASS'
        else:
            result = 'FAIL'
        
        ws_report_review_add_result['K'+ str(z)] = str(result)

        if result == 'FAIL':
            for col in range(1, 12):
                cell_header = ws_report_review_add_result.cell(v, col)
            # used hex code for red color
                cell_header.fill = PatternFill(
                    start_color='00FF0000', end_color='00FF0000', fill_type="solid")
                cell_header.border = border
                cell_header.font = font_text3
                cell_header.alignment = alignment
        if result == 'PASS':
            for col in range(11, 12):
                cell_header = ws_report_review_add_result.cell(v, col)
            # used hex code for red color
                cell_header.fill = PatternFill(
                    start_color='0099CC00', end_color='0099CC00', fill_type="solid")
                cell_header.border = border
                cell_header.font = font_text3
                cell_header.alignment = alignment

        n += 1
        m += 1
        v += 1
        z += 1

    wb_report_review_add_result.save("Report_review.xlsx")
    wb_report_review_add_result.close()

create_DPT_can_load_file()
review_report(ws_report_review, wb_report_review)
# excute_excell_file()
