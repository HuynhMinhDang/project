from email import message
from msilib.schema import RadioButton
from turtle import textinput
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Color, numbers
from openpyxl.worksheet.datavalidation import DataValidation
import os
from struct import pack
import tkinter as tk
from PIL import Image, ImageTk
from tkinter.filedialog import askopenfile
from tkinter import CENTER, filedialog
from tkinter import HORIZONTAL, messagebox
from tkinter import ttk
from tkinter import IntVar
import tkinter.ttk
import time
import string
from tkinter.filedialog import askopenfile

from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re

import pandas as pd
from pandas import read_html
import html5lib





        # print(result_of_string)

# ----------------------------------------------------------------
# excel
try:
    # os.system('TASKKILL /F /IM EXCEL.exe')
    os.remove("Report_review.xlsx")
    wb_report_review = Workbook()
    ws_report_review = wb_report_review.active

    ws_report_review.title = "Report_review"
    wb_report_review.save("Report_review.xlsx")
    os.close("Report_review.xlsx")
except:
    try:
        wb_report_review = Workbook()
        ws_report_review = wb_report_review.active
        ws_report_review.title = "Report_review"
        wb_report_review.save("Report_review.xlsx")
        # os.system('TASKKILL /F /IM EXCEL.exe')
    except OSError:
        print('Failed creating the file')
    else:
        print('File created')

# ------------------------------------------------------------------------------------------------------------------------
def create_DPT_file():
    try:
        with open("DPT.xlsx", "r") as file:
            # Print the success message
            print("File is already haved")
        # fd = os.open("RFvalue.xlsx", os.O_RDWR)
        # os.close(fd)
    except OSError:
        wb_read_DPT = load_workbook('DPT.xlsx')

        ws_read_DPT = wb_read_DPT.active

        ws_read_DPT = wb_read_DPT['DPT']

        sheet_read_DPT = wb_read_DPT.worksheets[0]

        noneFill = PatternFill(start_color='00FFFFFF',
                                end_color='00FFFFFF',
                                fill_type='solid'
                                )
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
        font_text = Font(name="Calibri", size=14, color='00FFFFFF', bold=True)
        font_text2 = Font(name="Calibri", size=11, color='000000', bold=False, italic = True)
        alignment = Alignment(horizontal='center', vertical='center')

        sheet_read_DPT.column_dimensions['B'].width = 50
        sheet_read_DPT.column_dimensions['C'].width = 20
        sheet_read_DPT.column_dimensions['D'].width = 30
        sheet_read_DPT.column_dimensions['E'].width = 30
        sheet_read_DPT.column_dimensions['F'].width = 40
        sheet_read_DPT.column_dimensions['G'].width = 40
        sheet_read_DPT.column_dimensions['H'].width = 35
        sheet_read_DPT.column_dimensions['A'].number_format = numbers.FORMAT_TEXT
        sheet_read_DPT.column_dimensions['B'].number_format = numbers.FORMAT_TEXT
        sheet_read_DPT.column_dimensions['C'].number_format = numbers.FORMAT_TEXT
        sheet_read_DPT.column_dimensions['D'].number_format = numbers.FORMAT_TEXT
        sheet_read_DPT.column_dimensions['E'].number_format = numbers.FORMAT_TEXT
        sheet_read_DPT.column_dimensions['F'].number_format = numbers.FORMAT_TEXT
        sheet_read_DPT.column_dimensions['G'].number_format = numbers.FORMAT_TEXT
        sheet_read_DPT.column_dimensions['A'].alignment = alignment
        sheet_read_DPT.column_dimensions['B'].alignment = alignment
        sheet_read_DPT.column_dimensions['C'].alignment = alignment
        sheet_read_DPT.column_dimensions['D'].alignment = alignment
        sheet_read_DPT.column_dimensions['E'].alignment = alignment
        sheet_read_DPT.column_dimensions['F'].alignment = alignment
        sheet_read_DPT.column_dimensions['G'].alignment = alignment

        ws_read_DPT.title = "DPT"
        ws_read_DPT.append(['DID', 'Length_Byte','ASCII Value', 'HEX_Value', 'Type'])

        # data_validation_data = '"APPL, FBL, RBEOL"'
        # for row in range(2, 30):
        #     data_validation = DataValidation(type='list', formula1 = data_validation_data)
        #     ws5.add_data_validation(data_validation)
        #     data_validation.add(ws5['E'+str(row)])

        wb_read_DPT.save("DPT.xlsx")
        wb_read_DPT.close()
        # os.system('TASKKILL /F /IM EXCEL.exe')
        messagebox.showinfo("COMPLETE", "File RFvalue.xlsx has been created in the same folder tool successfully, Please fill all value")
        print('tao thanh cong')


url = 'DIDCheck_DPT.html'

def review_report(ws_report_review, wb_report_review):
    wb_report_review = load_workbook('Report_review.xlsx')

    ws_report_review = wb_report_review.active
    ws_report_review = wb_report_review['Report_review']
    sheet_report_review = wb_report_review.worksheets[0]
    row_count_report_review = sheet_report_review.max_row
    # if direct == '':
    #     wb_read_DPT = load_workbook('DPT.xlsx')
    # else:
    #     wb_read_DPT = load_workbook(str(direct))
    # ws_read_DPT = wb_read_DPT.active
    # ws_read_DPT = wb_read_DPT['DPT']
    # sheet_read_DPT = wb_read_DPT.worksheets[0]
    # row_count_read_DPT = sheet_read_DPT.max_row
    id = 1
    with open(url, 'r') as f:

        ws_report_review.append(['ID_'+str(id) , 'DID','Length Byte RECEIVED','ASCII RECEIVED Value', 'HEX RECEIVED Value','HEX EXPECTED Value', 'ASCII EXPECTED Value','Length Byte EXPECTED', '', '', 'DPT'])
            
        result_of_string = f.read()
        i = 0
        result_of_string = result_of_string.partition("EXPECTED: 62")[2]

        while len(result_of_string) > 0:
            id += 1
            length_of_EXPECTED = result_of_string.find('<')
            # print(length_of_EXPECTED)
            

            result_of_EXPECTED = result_of_string
            result_of_EXPECTED = result_of_EXPECTED.split('<', 1)[0]

            value_of_EXPECTED = result_of_EXPECTED[4:length_of_EXPECTED]
            DID_of_EXPECTED = result_of_EXPECTED[0:4]

            # print(value_of_EXPECTED)

            try:
                bytes_object_of_EXPECTED = bytes.fromhex(value_of_EXPECTED)
                
            except:
                missing_lenght_byte_of_EXPECTED = value_of_EXPECTED.partition(".{")[2]
                missing_lenght_byte_of_EXPECTED = missing_lenght_byte_of_EXPECTED[:-1]
                # print(missing_lenght_byte_of_EXPECTED)

                # missing_lenght_byte_of_EXPECTED = int(missing_lenght_byte_of_EXPECTED) // 2


                length_of_EXPECTED = int(length_of_EXPECTED) + int(missing_lenght_byte_of_EXPECTED) - 4
                # print(length_of_EXPECTED)

                value_of_EXPECTED = value_of_EXPECTED.split('.', 1)[0]
                print(value_of_EXPECTED)

                bytes_object_of_EXPECTED = bytes.fromhex(value_of_EXPECTED)
                # print(bytes_object_of_EXPECTED)

                ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")
                # print(ascii_string_of_EXPECTED)
                # Convert to bytes object
            try:
                ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")

            # Convert to ASCII representation
            except:
                # ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")
                # ascii_string_of_EXPECTED = ILLEGAL_CHARACTERS_RE.sub(r'', ascii_string_of_EXPECTED)
                print('khong')
            if str(value_of_EXPECTED) != '':
                length_of_EXPECTED = (length_of_EXPECTED - 4) // 2
                # print('value khong trong '+ str(length_of_EXPECTED))
            else:
                length_of_EXPECTED = (length_of_EXPECTED // 2) - 2
                # print('value trong')

            # length_of_EXPECTED = (length_of_EXPECTED - 4) // 2

            
            

            # print('ASCII Value EXPECTED: ' + ascii_string_of_EXPECTED)
            # print(result_of_RECEIVED)
            # print('HEX Value EXPECTED: ' + value_of_EXPECTED)

            result_of_string = result_of_string.partition("RECEIVED: 62")[2] 
            # print(result_of_string)

            ### This is the part that measures the length before '<'
            length_of_RECEIVED = result_of_string.find('<')
            # print(length)
            

            result_of_RECEIVED = result_of_string
            result_of_RECEIVED = result_of_RECEIVED.split('<', 1)[0]

            value_of_RECEIVED = result_of_RECEIVED[4:length_of_RECEIVED]
            DID_of_RECEIVED = result_of_RECEIVED[0:4]

            bytes_object_of_RECEIVED = bytes.fromhex(value_of_RECEIVED)
                # Convert to bytes object

            length_of_RECEIVED = (length_of_RECEIVED - 4) // 2

            try:
                ascii_string_of_RECEIVED = bytes_object_of_RECEIVED.decode("ASCII")
                # print(ascii_string_of_RECEIVED)
                ascii_string_of_RECEIVED = ILLEGAL_CHARACTERS_RE.sub(r'', ascii_string_of_RECEIVED)
                # print(ascii_string_of_RECEIVED)
                
            # Convert to ASCII representation
            except:
                ascii_string_of_RECEIVED = ILLEGAL_CHARACTERS_RE.sub(r'', ascii_string_of_RECEIVED)
                print(ascii_string_of_RECEIVED)

                # print('')
                # ws_report_review.append(['ID_'+str(id) , str(DID_of_EXPECTED),'', str(value_of_RECEIVED),str(value_of_EXPECTED), '', '', '', 'DPT'])

            # print('ASCII Value RECEIVED: ' + ascii_string_of_RECEIVED)
            # print(result_of_RECEIVED)
            # print('HEX Value RECEIVED: ' + value_of_RECEIVED)
            # print(result_of_RECEIVED)

            # wb_report_review.save("Report_review.xlsx")
            # try:
            result_of_string = result_of_string.partition("EXPECTED: 62")[2]
            # except:
                # result_of_string = result_of_string.partition("RECEIVED: 62")[2]
                
            # ws_report_review.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(length_of_RECEIVED),str(ascii_string_of_RECEIVED), str(value_of_RECEIVED),str(value_of_EXPECTED), str(ascii_string_of_EXPECTED),str(length_of_EXPECTED), '', '', 'DPT'])
            
            if str(ascii_string_of_EXPECTED) == str(ascii_string_of_RECEIVED) and str(value_of_EXPECTED) == str(value_of_RECEIVED) or str(length_of_EXPECTED) == str(length_of_RECEIVED):
                result = 'PASS'
            else:
                result = 'FAIL'
            
            try:
                ws_report_review.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(length_of_RECEIVED),str(ascii_string_of_RECEIVED), str(value_of_RECEIVED),str(value_of_EXPECTED), str(ascii_string_of_EXPECTED),str(length_of_EXPECTED), '', '', 'DPT',str(result)])
            except:
                ws_report_review.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(length_of_RECEIVED),'', str(value_of_RECEIVED),str(value_of_EXPECTED), '',str(length_of_EXPECTED),'', '', 'DPT',str(result)])
            
            # print(result_of_string)
            # print(len(f_contens))


                # print('PASS')
            
            
            
            # ws_report_review.append(['ID_'+str(id) + str(DID_of_EXPECTED),str(ascii_string_of_RECEIVED), str(value_of_RECEIVED),str(value_of_EXPECTED), str(ascii_string_of_EXPECTED), '', '', 'DPT'])
        

            
            i = len(result_of_string)
    
    with open('Report_review.txt', 'w') as f:
        f.write(result_of_string)

    wb_report_review.save("Report_review.xlsx")
    wb_report_review.close()
    

review_report(ws_report_review, wb_report_review)



