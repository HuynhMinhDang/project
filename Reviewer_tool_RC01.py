from email import message
from msilib.schema import RadioButton
from turtle import textinput
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Color, numbers
from openpyxl.worksheet.datavalidation import DataValidation
import os
from struct import pack
import tkinter as tk
from PIL import Image, ImageTk
from tkinter.filedialog import askopenfile
from tkinter import CENTER, filedialog
from tkinter import HORIZONTAL, messagebox
from tkinter import ttk
from tkinter import IntVar
import tkinter.ttk
import time
import string
from tkinter.filedialog import askopenfile

import pandas as pd
from pandas import read_html
import html5lib

url = 'DIDCheck_DPT.html'



        # print(result_of_string)

# ----------------------------------------------------------------
# excel
try:
    # os.system('TASKKILL /F /IM EXCEL.exe')
    os.remove("Report_review.xlsx")
    wb_report_review = Workbook()
    ws_report_review = wb_report_review.active

    ws_report_review.title = "Report_review"
    wb_report_review.save("Report_review.xlsx")
    os.close("Report_review.xlsx")
except:
    try:
        wb_report_review = Workbook()
        ws_report_review = wb_report_review.active
        ws_report_review.title = "Report_review"
        wb_report_review.save("Report_review.xlsx")
        # os.system('TASKKILL /F /IM EXCEL.exe')
    except OSError:
        print('Failed creating the file')
    else:
        print('File created')

# ------------------------------------------------------------------------------------------------------------------------
def create_DPT_file():
    try:
        with open("DPT.xlsx", "r") as file:
            # Print the success message
            print("File is already haved")
        # fd = os.open("RFvalue.xlsx", os.O_RDWR)
        # os.close(fd)
    except OSError:
        wb_read_DPT = load_workbook('DPT.xlsx')

        ws_read_DPT = wb_read_DPT.active

        ws_read_DPT = wb_read_DPT['DPT']

        sheet_read_DPT = wb_read_DPT.worksheets[0]

        noneFill = PatternFill(start_color='00FFFFFF',
                                end_color='00FFFFFF',
                                fill_type='solid'
                                )
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
        font_text = Font(name="Calibri", size=14, color='00FFFFFF', bold=True)
        font_text2 = Font(name="Calibri", size=11, color='000000', bold=False, italic = True)
        alignment = Alignment(horizontal='center', vertical='center')

        sheet_read_DPT.column_dimensions['B'].width = 50
        sheet_read_DPT.column_dimensions['C'].width = 20
        sheet_read_DPT.column_dimensions['D'].width = 30
        sheet_read_DPT.column_dimensions['E'].width = 30
        sheet_read_DPT.column_dimensions['F'].width = 40
        sheet_read_DPT.column_dimensions['G'].width = 40
        sheet_read_DPT.column_dimensions['H'].width = 35
        sheet_read_DPT.column_dimensions['A'].number_format = numbers.FORMAT_TEXT
        sheet_read_DPT.column_dimensions['B'].number_format = numbers.FORMAT_TEXT
        sheet_read_DPT.column_dimensions['C'].number_format = numbers.FORMAT_TEXT
        sheet_read_DPT.column_dimensions['D'].number_format = numbers.FORMAT_TEXT
        sheet_read_DPT.column_dimensions['E'].number_format = numbers.FORMAT_TEXT
        sheet_read_DPT.column_dimensions['F'].number_format = numbers.FORMAT_TEXT
        sheet_read_DPT.column_dimensions['G'].number_format = numbers.FORMAT_TEXT
        sheet_read_DPT.column_dimensions['A'].alignment = alignment
        sheet_read_DPT.column_dimensions['B'].alignment = alignment
        sheet_read_DPT.column_dimensions['C'].alignment = alignment
        sheet_read_DPT.column_dimensions['D'].alignment = alignment
        sheet_read_DPT.column_dimensions['E'].alignment = alignment
        sheet_read_DPT.column_dimensions['F'].alignment = alignment
        sheet_read_DPT.column_dimensions['G'].alignment = alignment

        ws_read_DPT.title = "DPT"
        ws_read_DPT.append(['DID', 'Length_Byte','ASCII Value', 'HEX_Value', 'Type'])

        # data_validation_data = '"APPL, FBL, RBEOL"'
        # for row in range(2, 30):
        #     data_validation = DataValidation(type='list', formula1 = data_validation_data)
        #     ws5.add_data_validation(data_validation)
        #     data_validation.add(ws5['E'+str(row)])

        wb_read_DPT.save("DPT.xlsx")
        wb_read_DPT.close()
        # os.system('TASKKILL /F /IM EXCEL.exe')
        messagebox.showinfo("COMPLETE", "File RFvalue.xlsx has been created in the same folder tool successfully, Please fill all value")
        print('tao thanh cong')



def review_report(ws_report_review, wb_report_review):
    wb_report_review = load_workbook('Report_review.xlsx')

    ws_report_review = wb_report_review.active
    ws_report_review = wb_report_review['Report_review']
    sheet_report_review = wb_report_review.worksheets[0]
    row_count_report_review = sheet_report_review.max_row
    # if direct == '':
    #     wb_read_DPT = load_workbook('DPT.xlsx')
    # else:
    #     wb_read_DPT = load_workbook(str(direct))
    # ws_read_DPT = wb_read_DPT.active
    # ws_read_DPT = wb_read_DPT['DPT']
    # sheet_read_DPT = wb_read_DPT.worksheets[0]
    # row_count_read_DPT = sheet_read_DPT.max_row
    id = 1
    with open(url, 'r') as f:

        ws_report_review.append(['ID_'+str(id) , 'DID','ASCII RECEIVED Value', 'HEX RECEIVED Value','HEX EXPECTED Value', 'ASCII EXPECTED Value', '', '', 'DPT'])
            
        f_contens = f.read()
        
        result_of_string = f_contens.partition("EXPECTED: 62")[2]

        length = result_of_string.find('<')

        result_of_EXPECTED = result_of_string
        result_of_EXPECTED = result_of_EXPECTED.split('<', 1)[0]

        value_of_EXPECTED = result_of_EXPECTED[4:length]
        DID_of_EXPECTED = result_of_EXPECTED[0:4]


        bytes_object_of_EXPECTED = bytes.fromhex(value_of_EXPECTED)
            # Convert to bytes object
        ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")
        # Convert to ASCII representation

        print('ASCII Value EXPECTED: ' + ascii_string_of_EXPECTED)
        # print(result_of_RECEIVED)
        print('HEX Value EXPECTED: ' + value_of_EXPECTED)

        result_of_string = result_of_string.partition("RECEIVED: 62")[2] 
        # print(result_of_string)

        ### This is the part that measures the length before '<'
        length = result_of_string.find('<')
        # print(length)

        result_of_RECEIVED = result_of_string
        result_of_RECEIVED = result_of_RECEIVED.split('<', 1)[0]

        value_of_RECEIVED = result_of_RECEIVED[4:length]
        DID_of_RECEIVED = result_of_RECEIVED[0:4]

        bytes_object_of_RECEIVED = bytes.fromhex(value_of_RECEIVED)
            # Convert to bytes object


        ascii_string_of_RECEIVED = bytes_object_of_RECEIVED.decode("ASCII")
        # Convert to ASCII representation

        print('ASCII Value RECEIVED: ' + ascii_string_of_RECEIVED)
        # print(result_of_RECEIVED)
        print('HEX Value RECEIVED: ' + value_of_RECEIVED)
        # print(result_of_RECEIVED)

        ws_report_review.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(ascii_string_of_RECEIVED), str(value_of_RECEIVED),str(value_of_EXPECTED), str(ascii_string_of_EXPECTED), '', '', 'DPT'])
        
        # wb_report_review.save("Report_review.xlsx")

        result_of_string = result_of_string.partition("RECEIVED: 62")[2]
        
        # print(result_of_string)
        # print(len(f_contens))
        i = 0

        if str(ascii_string_of_EXPECTED) == str(ascii_string_of_RECEIVED) and str(value_of_EXPECTED) == str(value_of_RECEIVED):
            print('PASS')
        
        
        
        ws_report_review.append(['ID_'+str(id) + str(DID_of_EXPECTED),str(ascii_string_of_RECEIVED), str(value_of_RECEIVED),str(value_of_EXPECTED), str(ascii_string_of_EXPECTED), '', '', 'DPT'])
        
        # wb_report_review.save("Report_review.xlsx")
        # wb_report_review.close()
        # os.system('TASKKILL /F /IM EXCEL.exe')
        # messagebox.showinfo("COMPLETE", "File RFvalue.xlsx has been created in the same folder tool successfully, Please fill all value")
        # print('tao thanh cong')

        while len(result_of_string) > 0:
            # ws_report_review.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(ascii_string_of_RECEIVED), str(value_of_RECEIVED),str(value_of_EXPECTED), str(ascii_string_of_EXPECTED), '', '', 'DPT'])
            
            ### This is the part that measures the length before '<'
            length = result_of_string.find('<')
            # print(length)

            result_of_RECEIVED = result_of_string
            result_of_RECEIVED = result_of_RECEIVED.split('<', 1)[0]
            
            value_of_RECEIVED = result_of_RECEIVED[4:length]
            DID_of_RECEIVED = result_of_RECEIVED[0:4]

            bytes_object_of_RECEIVED = bytes.fromhex(value_of_RECEIVED)
            # Convert to bytes object

            try:
                ascii_string_of_RECEIVED = bytes_object_of_RECEIVED.decode("ASCII")
            except:
                print("")
            # ascii_string_of_RECEIVED = bytes_object_of_RECEIVED.decode("ASCII")
            
            # Convert to ASCII representation

            print('ASCII Value: ' + ascii_string_of_RECEIVED)
            # print(result_of_RECEIVED)
            print('HEX Value: ' + value_of_RECEIVED)
            # print(DID)
            # ws_report_review.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(ascii_string_of_RECEIVED), str(value_of_RECEIVED),str(value_of_EXPECTED), str(ascii_string_of_EXPECTED), '', '', 'DPT'])
                
            
            result_of_string = result_of_string.partition("RECEIVED: 62")[2]
            
            i = len(result_of_string)

review_report(ws_report_review, wb_report_review)
wb_report_review.save("Report_review.xlsx")
wb_report_review.close()



# def DPT(ws_report_review, wb_read_DPT, id, number1, number2, number3, number4, direct):
#     if direct == '':
#         wb_read_DPT = load_workbook('DPT.xlsx')
#     else:
#         wb_read_DPT = load_workbook(str(direct))
#     ws_read_DPT = wb_read_DPT.active
#     ws_read_DPT = wb_read_DPT['DPT']
#     sheet_read_DPT = wb_read_DPT.worksheets[0]
#     row_count_read_DPT = sheet_read_DPT.max_row

#     i = 0
#     o = 2
#     j = 3

#     k = 1

#     i = 0

#     hexvalue_baseSW = ""
#     while k < row_count_read_DPT:
#         for row in range(o, j):
#             for col in range(1, 2):
#                 char = get_column_letter(col)
#                 row_list_DID_baseSW = ws_read_DPT[char + str(row)].value
#                 row_list_DID_baseSW_lowercase = str(row_list_DID_baseSW).lower()

#         for row in range(o, j):
#             for col in range(2, 3):
#                 char = get_column_letter(col)
#                 row_list_length_byte_baseSW = ws[char + str(row)].value

#         for row in range(o, j):
#             for col in range(3, 4):
#                 char = get_column_letter(col)
#                 row_list_ASCII_values_baseSW = ws[char + str(row)].value
        
#         for row in range(o, j):
#             for col in range(4, 5):
#                 char = get_column_letter(col)
#                 row_list_hex_values_baseSW = ws[char + str(row)].value


#         id += 1
#         # check lenghth byte
#         count_hexvalue_baseSW = 0
#         hexvalue_baseSW = ""
#         length_byte = 0
#         if str(row_list_ASCII_values_baseSW) == 'None'  and str(row_list_hex_values_baseSW) == 'None':
#             messagebox.showerror("ERROR", "Value in DPT.xlsx file is EMPTY!!!, Pleas fill the value")
#         else:
#             if str(row_list_hex_values_baseSW) == "None":
#                 # change ascii sang hex value
#                 for i in str(row_list_ASCII_values_baseSW):
#                     hexvalue_baseSW += hex(ord(i))[2:]

#             # danh cho co length byter thi su dung
#                 # print(hexvalue_baseSW)
#                 # count_hexvalue_baseSW = len(hexvalue_baseSW)
#                 # count_hexvalue_baseSW = int(count_hexvalue_baseSW) // 2
#                 # # print(count_hexvalue_baseSW)
#                 # # print(type(row_list_length_byte_baseSW))
#                 # if str(count_hexvalue_baseSW) < row_list_length_byte_baseSW:
#                 #     # print("Day la do dai byte",row_list_length_byte_baseSW)
#                 #     length_byte = (int(row_list_length_byte_baseSW) -
#                 #                 int(count_hexvalue_baseSW)) * 2
#                 #     # print("byte bi thieu", length_byte)
#                 #     hexvalue_baseSW = hexvalue_baseSW.lower()
#                 #     hexvalue_baseSW = str(hexvalue_baseSW + ".{" + str(length_byte) + "}")
# # --------------------------------------------------------------------------------------------------------------

#                 hexvalue_baseSW = hexvalue_baseSW.lower()

#             else:

#             # danh cho co length byter thi su dung
#                 count_hexvalue_baseSW = len(row_list_hex_values_baseSW)
#                 count_hexvalue_baseSW = int(count_hexvalue_baseSW) // 2
#                 if str(count_hexvalue_baseSW) < row_list_length_byte_baseSW:
#                     length_byte = (int(row_list_length_byte_baseSW) - int(count_hexvalue_baseSW)) * 2
#                     hexvalue_baseSW = row_list_hex_values_baseSW.lower()
#                     hexvalue_baseSW = str(hexvalue_baseSW + ".{" + str(length_byte) + "}")
# # ------------------------------------------------------------------------------------------------------------------

            
#             # print("hoan thanh",hexvalue_baseSW)
#         if str(row_list_name_baseSW) == "None":
#             row_list_name_baseSW = ""
            
#         # if str(row_list_name_baseSW) != "Supplier Software number":

        
#         ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW) + ' in RBEOL', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID ' +
#             str(row_list_DID_baseSW) + ' using physical addressing\n5) Reset ECU\n6) Wait 3s\n7) Send 1001\n8) Wait 3s', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -',
#             '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)\n5) RequestResponse(1101, 5101, Equal)\n6) wait(3000)\n7) RequestResponse(1001, 5001.*, Regexp)\n8) wait(3000)', 'Automated Testcase', 'implemented', baseSW, ''])
            
#         number4 += 1
#         o += 1
#         j += 1
#         k += 1

#     return id



# wb_read_DPT = load_workbook('DPT.xlsx')

# ws_read_DPT = wb_read_DPT.active

# ws_read_DPT = wb_read_DPT['DPT']

# sheet_read_DPT = wb_read_DPT.worksheets[0]
# row_count_read_DPT = sheet_read_DPT.max_row

# tasks = row_count_read_DPT



# with open('DIDCheck_DPT.txt', 'w') as f:
#     f.write(result_of_string)


