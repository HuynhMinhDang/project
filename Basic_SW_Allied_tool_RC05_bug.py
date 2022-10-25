################################################################################################################################################################################################################################################################################################################################################################################################################################################################################################################################################################################################################################
# Date : 14/10/2022                                                                
# Basic_SW_Allied_Tool RC05                                                                
# developer : Huynh Minh Dang                                                          
# version : V1.1.1                                                                     
# Description : bug 
######################################################################################################################################################################################################################################################################################################################################################################################################################################################################################################################################################################################################

from email import message
from msilib.schema import RadioButton
from turtle import textinput
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Color, numbers
from openpyxl.worksheet.datavalidation import DataValidation
import os
from struct import pack
import tkinter as tk
from PIL import Image, ImageTk
from tkinter.filedialog import askopenfile
from tkinter import CENTER, filedialog
from tkinter import HORIZONTAL, messagebox
from tkinter import ttk
from tkinter import IntVar
import tkinter.ttk
import time
import string
from tkinter.filedialog import askopenfile


try:
    # os.system('TASKKILL /F /IM EXCEL.exe')
    os.remove("TC_RF.xlsx")
    wb_TC_RF = Workbook()
    ws_TC_RF = wb_TC_RF.active
    ws_TC_RF.title = "TC_RF"
    wb_TC_RF.save("TC_RF.xlsx")
    os.close("TC_RF.xlsx")
    wb_TC_RF.close()
except:
    try:
        wb_TC_RF = Workbook()
        ws_TC_RF = wb_TC_RF.active
        ws_TC_RF.title = "TC_RF"
        wb_TC_RF.save("TC_RF.xlsx")
        wb_TC_RF.close()
        # os.system('TASKKILL /F /IM EXCEL.exe')
    except OSError:
        print('Failed creating the file TC_RF.xlsx')
    else:
        print('File TC_RF.xlsx created')

try:
    # os.system('TASKKILL /F /IM EXCEL.exe')
    os.remove("TC_FBL.xlsx")
    wb_TC_FBL = Workbook()
    ws_TC_FBL = wb_TC_FBL.active
    ws_TC_FBL.title = "TC_FBL"
    wb_TC_FBL.save("TC_FBL.xlsx")
    wb_TC_FBL.close()
    os.close("TC_FBL.xlsx")
except:
    try:
        wb_TC_FBL = Workbook()
        ws_TC_FBL = wb_TC_FBL.active
        ws_TC_FBL.title = "TC_FBL"
        wb_TC_FBL.save("TC_FBL.xlsx")
        wb_TC_FBL.close()
        # os.system('TASKKILL /F /IM EXCEL.exe')
    except OSError:
        print('Failed creating the file TC_FBL.xlsx')
    else:
        print('File TC_FBL.xlsx created ')


def create_value_file():
    try:
        with open("BSWAvalue.xlsx", "r") as file:
            # Print the success message
            print("File is already haved")
        # fd = os.open("BSWAvalue.xlsx", os.O_RDWR)
        # os.close(fd)
    except OSError:
        z = 0
        x = 0
        wb5 = Workbook()
        ws5 = wb5.active
        sheet = wb5.worksheets[0]
        noneFill = PatternFill(start_color='00FFFFFF',
                                end_color='00FFFFFF',
                                fill_type='solid'
                                )
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
        font_text = Font(name="Calibri", size=14, color='00FFFFFF', bold=True)
        font_text2 = Font(name="Calibri", size=11, color='000000', bold=False, italic = True)
        alignment = Alignment(horizontal='center', vertical='center')

        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['F'].width = 30
        sheet.column_dimensions['G'].width = 40
        sheet.column_dimensions['H'].width = 40
        sheet.column_dimensions['I'].width = 35
        sheet.column_dimensions['A'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['B'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['C'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['D'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['E'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['F'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['G'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['H'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['A'].alignment = alignment
        sheet.column_dimensions['B'].alignment = alignment
        sheet.column_dimensions['C'].alignment = alignment
        sheet.column_dimensions['D'].alignment = alignment
        sheet.column_dimensions['E'].alignment = alignment
        sheet.column_dimensions['F'].alignment = alignment
        sheet.column_dimensions['G'].alignment = alignment
        sheet.column_dimensions['H'].alignment = alignment


# tao ra sheet base sw

        ws5.title = "RFvalue_baseSW"
        ws5.append(['DID', 'Description', 'Length (Byte)','ASCII Value', 'HEX_Value', 'Type'])
        # ws5.append(['DID', 'Description','ASCII Value', 'HEX_Value', 'Type'])

        # ws5['F1'] = 'BaseSW Name'
        # cell_header = ws5.cell(1, 6)
        # cell_header.fill = PatternFill(
        #     start_color='000066CC', end_color='000066CC', fill_type="solid")
        # cell_header.border = border
        # cell_header.font = font_text
        # cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'Ticket BaseSW'])
        ws5['G2'] = 'Variant Name baseSW'
        cell_header = ws5.cell(2, 7)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'Variant BaseSW'])
        ws5['G3'] = 'Variant BaseSW'
        cell_header = ws5.cell(3, 7)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'DID check variant BaseSW'])
        ws5['G4'] = 'DID check variant BaseSW'
        cell_header = ws5.cell(4, 7)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['G5'] = 'Security Unlock Level'
        cell_header = ws5.cell(5, 7)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['I5'] = '0: No level; 1: level1; 2: level2; 3: level3'
        cell_header = ws5.cell(5, 10)
        cell_header.border = border
        cell_header.font = font_text2



        for col in range(1, 8):
            cell_header = ws5.cell(1, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='000066CC', end_color='000066CC', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
            cell_header.alignment = alignment

        data_validation_data = '"APPL, FBL, RBEOL"'
        for row in range(2, 30):
            data_validation = DataValidation(type='list', formula1 = data_validation_data)
            ws5.add_data_validation(data_validation)
            data_validation.add(ws5['F'+str(row)])



        # cell_header = ws5.cell(1, 5)
        # cell_header.fill = noneFill
        # cell_header.border = border

# ket thuc sheet baseSW

# tao ra sheet latest SW

        wb5.create_sheet("RFvalue_latestSW")
        ws5 = wb5['RFvalue_latestSW']
        sheet2 = wb5.worksheets[1]
        sheet2.column_dimensions['B'].width = 50
        sheet2.column_dimensions['C'].width = 20
        sheet2.column_dimensions['D'].width = 20
        sheet2.column_dimensions['E'].width = 30
        sheet2.column_dimensions['F'].width = 20
        sheet2.column_dimensions['G'].width = 30
        sheet2.column_dimensions['H'].width = 40
        sheet2.column_dimensions['I'].width = 40
        sheet2.column_dimensions['J'].width = 35
        sheet2.column_dimensions['A'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['B'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['C'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['D'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['E'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['F'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['G'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['H'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['I'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['A'].alignment = alignment
        sheet2.column_dimensions['B'].alignment = alignment
        sheet2.column_dimensions['C'].alignment = alignment
        sheet2.column_dimensions['D'].alignment = alignment
        sheet2.column_dimensions['E'].alignment = alignment
        sheet2.column_dimensions['F'].alignment = alignment
        sheet2.column_dimensions['G'].alignment = alignment
        sheet2.column_dimensions['H'].alignment = alignment
        sheet2.column_dimensions['I'].alignment = alignment
        # column_count2 = sheet2.max_column
        ws5.append(['DID', 'Description', 'Length (Byte)','ASCII Value', 'HEX_Value', 'Type','DummySW value(hex)'])
        # ws5.append(['DID', 'Description','ASCII Value', 'HEX_Value', 'Type','DummySW value(hex)'])

        ws5['H2'] = 'Variant Name LatestSW'
        cell_header = ws5.cell(2, 8)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'Variant LatestSW'])
        ws5['H3'] = 'Variant LatestSW'
        cell_header = ws5.cell(3, 8)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'DID check variant LatestSW'])
        ws5['H4'] = 'DID check variant LatestSW'
        cell_header = ws5.cell(4, 8)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['H5'] = 'Security Unlock Level'
        cell_header = ws5.cell(5, 8)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['J5'] = '0: No level; 1: level1; 2: level2; 3: level3'
        cell_header = ws5.cell(5, 11)
        cell_header.border = border
        cell_header.font = font_text2



        for col in range(1, 9):
            cell_header = ws5.cell(1, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='000066CC', end_color='000066CC', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
            cell_header.alignment = alignment
            
        data_validation_data = '"APPL, FBL, RBEOL"'
        
        for row in range(2, 30):
            data_validation = DataValidation(type='list', formula1 = data_validation_data)
            ws5.add_data_validation(data_validation)
            data_validation.add(ws5['F'+str(row)])


# ket thuc sheet latest SW


# tao ra sheet DID_DPT

        wb5.create_sheet("DID_DPT")
        ws5 = wb5['DID_DPT']
        sheet3 = wb5.worksheets[2]
        sheet3.column_dimensions['B'].width = 50
        sheet3.column_dimensions['C'].width = 20
        sheet3.column_dimensions['D'].width = 30
        sheet3.column_dimensions['E'].width = 20
        sheet3.column_dimensions['F'].width = 20
        sheet3.column_dimensions['G'].width = 25
        sheet3.column_dimensions['A'].number_format = numbers.FORMAT_TEXT
        sheet3.column_dimensions['B'].number_format = numbers.FORMAT_TEXT
        sheet3.column_dimensions['C'].number_format = numbers.FORMAT_TEXT
        sheet3.column_dimensions['D'].number_format = numbers.FORMAT_TEXT
        sheet3.column_dimensions['E'].number_format = numbers.FORMAT_TEXT
        sheet3.column_dimensions['F'].number_format = numbers.FORMAT_TEXT
        sheet3.column_dimensions['G'].number_format = numbers.FORMAT_TEXT
        sheet3.column_dimensions['A'].alignment = alignment
        sheet3.column_dimensions['B'].alignment = alignment
        sheet3.column_dimensions['C'].alignment = alignment
        sheet3.column_dimensions['D'].alignment = alignment
        sheet3.column_dimensions['E'].alignment = alignment
        sheet3.column_dimensions['F'].alignment = alignment
        sheet3.column_dimensions['G'].alignment = alignment
        # column_count2 = sheet2.max_column
        ws5.append(['DID', 'Description', 'Length (Byte)','ASCII Value', 'HEX_Value', 'Type', 'Programming Type'])
        # # ws5.append(['DID', 'Description','ASCII Value', 'HEX_Value', 'Type'])

        for col in range(1, 8):
            cell_header = ws5.cell(1, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='000066CC', end_color='000066CC', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
            cell_header.alignment = alignment
            
        data_validation_data_FBL = '"APL, FCN , APL and FCN, NONE"'
        
        for row in range(2, 30):
            data_validation_FBL = DataValidation(type='list', formula1 = data_validation_data_FBL)
            ws5.add_data_validation(data_validation_FBL)
            data_validation_FBL.add(ws5['F'+str(row)])
            data_validation_FBL.add(ws5['G'+str(row)])

        


# ket thuc sheet DID_DPT

# tao ra sheet general information

        wb5.create_sheet("General Information")
        ws5 = wb5['General Information']
        sheet4 = wb5.worksheets[3]
        sheet4.column_dimensions['A'].width = 50
        sheet4.column_dimensions['B'].width = 50
        sheet4.column_dimensions['C'].width = 50
        
        
        sheet4.column_dimensions['A'].number_format = numbers.FORMAT_TEXT
        sheet4.column_dimensions['B'].number_format = numbers.FORMAT_TEXT
        sheet4.column_dimensions['C'].number_format = numbers.FORMAT_TEXT
        
        sheet4.column_dimensions['A'].alignment = alignment
        sheet4.column_dimensions['B'].alignment = alignment
        sheet4.column_dimensions['C'].alignment = alignment

        ws5['A1'] = 'BaseSW Name'
        cell_header = ws5.cell(1, 1)
        cell_header.fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['A2'] = 'LatestSW Name'
        cell_header = ws5.cell(2, 1)
        cell_header.fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['A3'] = 'Ticket ID'
        cell_header = ws5.cell(3, 1)
        cell_header.fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['A4'] = 'Programming Counter DID'
        cell_header = ws5.cell(4, 1)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['A5'] = 'Programming Attempt Counter DID'
        cell_header = ws5.cell(5, 1)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['A6'] = 'Step counter of PC DID'
        cell_header = ws5.cell(6, 1)
        cell_header.fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['A7'] = 'Step counter of PAC DID'
        cell_header = ws5.cell(7, 1)
        cell_header.fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['A8'] = 'DID Check active session'
        cell_header = ws5.cell(8, 1)
        cell_header.fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['C6'] = '*PC-Programing_counter'
        cell_header = ws5.cell(6, 3)
        cell_header.font = font_text2
    
        ws5['C7'] = '*PAC-Programing_Attempt_Counter'
        cell_header = ws5.cell(7, 3)
        cell_header.font = font_text2

        
        for col in range(1, 2):
            cell_header = ws5.cell(1, col)
            # used hex code for red color
            cell_header.fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
            cell_header.alignment = alignment




# ket thuc sheet general information

        # cell_header = ws5.cell(1, 5)
        # cell_header.fill = noneFill
        # cell_header.border = border

        wb5.save("BSWAvalue.xlsx")
        wb5.close()
        # os.system('TASKKILL /F /IM EXCEL.exe')
        messagebox.showinfo("COMPLETE", "File BSWAvalue.xlsx has been created in the same folder tool successfully, Please fill all value")
        print('tao thanh cong')


check_DID = 0

def DID_baseSW(ws_RF_BaseSW, wb_RF_baseSW, id, number1, number2, number3, number4, direct):
    if direct == '':
        wb_RF_baseSW = load_workbook('BSWAvalue.xlsx')
        wb_General_Infomation = load_workbook('BSWAvalue.xlsx')
    else:
        wb_RF_baseSW = load_workbook(str(direct))
        wb_General_Infomation = load_workbook(str(direct))

    ws_RF_BaseSW = wb_RF_baseSW.active
    ws_General_Infomation = wb_General_Infomation.active

    ws_RF_BaseSW = wb_RF_baseSW['RFvalue_baseSW']
    ws_General_Infomation = wb_General_Infomation['General Information']

    sheet_RF_BaseSW = wb_RF_baseSW.worksheets[0]
    row_count_RF_BaseSW = sheet_RF_BaseSW.max_row
    for row in range(1, 2):
        for col in range(2, 3):
            char = get_column_letter(col)
            baseSW = ws_General_Infomation[char + str(row)].value


    i = 0
    o = 2
    j = 3

    k = 1
    number4 += 1

    c = 0

    i = 0
    count_string_number = 0
    hexvalue_baseSW = ""
    while k < row_count_RF_BaseSW:
        for row in range(o, j):
            for col in range(1, 2):
                char = get_column_letter(col)
                row_list_DID_baseSW = ws_RF_BaseSW[char + str(row)].value
                row_list_DID_baseSW_lowercase = str(row_list_DID_baseSW).lower()

        for row in range(o, j):
            for col in range(2, 3):
                char = get_column_letter(col)
                row_list_name_baseSW = ws_RF_BaseSW[char + str(row)].value
                if str(row_list_DID_baseSW) == "None":
                    messagebox.showerror("ERROR", "No DID or DID invalid, Please add DID")
                    break
                # print(ws_RF_BaseSW[char + str(row)].value)


        for row in range(o, j):
            for col in range(3, 4):
                char = get_column_letter(col)
                row_list_length_byte_baseSW = ws_RF_BaseSW[char + str(row)].value
                # print(row_list_length_byte_baseSW)

        for row in range(o, j):
            for col in range(4, 5):
                char = get_column_letter(col)
                row_list_ASCII_values_baseSW = ws_RF_BaseSW[char + str(row)].value
        
        for row in range(o, j):
            for col in range(5, 6):
                char = get_column_letter(col)
                row_list_hex_values_baseSW = ws_RF_BaseSW[char + str(row)].value

        for row in range(o, j):
            for col in range(6, 7):
                char = get_column_letter(col)
                row_list_type_baseSW = ws_RF_BaseSW[char + str(row)].value
                # print(row_list_type_baseSW)

        id += 1
        # check lenghth byte
        count_hexvalue_baseSW = 0
        hexvalue_baseSW = ""
        length_byte = 0
        if str(row_list_ASCII_values_baseSW) == 'None'  and str(row_list_hex_values_baseSW) == 'None':
            # hexvalue_baseSW = str(".{" + str(row_list_length_byte_baseSW) + "}")
            if str(row_list_length_byte_baseSW) != 'None':
                    length_byte = int(row_list_length_byte_baseSW) * 2
                    # print(length_byte)
                    hexvalue_baseSW = str(".{" + str(length_byte) + "}")
            else:
                    hexvalue_baseSW = str(".*")
                # print('dung')
                # print(hexvalue_baseSW)
            c = 1
        else:
            if str(row_list_hex_values_baseSW) == "None":
                # change ascii sang hex value
                for i in str(row_list_ASCII_values_baseSW):
                    hexvalue_baseSW += hex(ord(i))[2:]

            # danh cho co length byter thi su dung
                if str(row_list_length_byte_baseSW) != 'None':
                    # print(hexvalue_baseSW)
                    count_hexvalue_baseSW = len(hexvalue_baseSW)
                    count_hexvalue_baseSW = int(count_hexvalue_baseSW) // 2
                    # print('count: ' + str(count_hexvalue_baseSW))
                    # print(type(row_list_length_byte_baseSW))
                    if str(count_hexvalue_baseSW) < row_list_length_byte_baseSW:
                        # print("Day la do dai byte",row_list_length_byte_baseSW)
                        length_byte = (int(row_list_length_byte_baseSW) -
                                    int(count_hexvalue_baseSW)) * 2
                        # print("byte bi thieu", length_byte)
                        hexvalue_baseSW = hexvalue_baseSW.lower()
                        hexvalue_baseSW = str(hexvalue_baseSW + ".{" + str(length_byte) + "}")
                        c = 1
                    else:
                        c = 0
                if str(row_list_length_byte_baseSW) == 'None':
                    hexvalue_baseSW = hexvalue_baseSW.lower()
                    hexvalue_baseSW = str( hexvalue_baseSW + ".*")
                    c = 1
                
            
            else:

            # danh cho co length byter thi su dung
                if str(row_list_length_byte_baseSW) != 'None':
                    count_hexvalue_baseSW = len(row_list_hex_values_baseSW)
                    count_hexvalue_baseSW = int(count_hexvalue_baseSW) // 2
                    if str(count_hexvalue_baseSW) < row_list_length_byte_baseSW:
                        length_byte = (int(row_list_length_byte_baseSW) - int(count_hexvalue_baseSW)) * 2
                        hexvalue_baseSW = row_list_hex_values_baseSW.lower()
                        hexvalue_baseSW = str(hexvalue_baseSW + ".{" + str(length_byte) + "}")
                        c = 1
                else:
                    hexvalue_baseSW = row_list_hex_values_baseSW.lower()
                    hexvalue_baseSW = str(hexvalue_baseSW + ".*")
                    c = 1

                # hexvalue_baseSW = row_list_hex_values_baseSW.lower()
                # hexvalue_baseSW = str(hexvalue_baseSW + ".*")
                # c = 1
            
            # print("hoan thanh",hexvalue_baseSW)
        if str(row_list_name_baseSW) == "None":
            row_list_name_baseSW = ""
            
        # if str(row_list_name_baseSW) != "Supplier Software number":
        
        if c == 1:
    
            if str(row_list_type_baseSW) == "APPL":
                
                ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW)+' in Application', '1) Send service 0x22 to the camera for the DID ' +
                            str(row_list_DID_baseSW) + ' using physical addressing', '1) -', '1) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
            
            if str(row_list_type_baseSW) == "FBL":
                # ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW) + ' in Programming', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Wait 3s\n4) Check active session should be Extended\n4) Wait 3s\n5) Change to Programming session with Service 0x10 02\n6) Wait 3s\n7) Check active session should be Programming\n8) Wait 3s\n9) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + ' using physical addressing\n10) Wait 3s\n11) Change to Default session with Service 0x10 01\n12) Wait 3s\n13) Check active session should be Default\n14) Wait 3s\n15) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + 'using physical addressing', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -\n13) -\n14) -\n15) -', '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(3000)\n5) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}3, Regexp)\n6) wait(1000)\n7) RequestResponse(1002, 5002.*, Regexp)\n8) wait(3000)\n9) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}2, Regexp)\n10) wait(3000)\n11) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)\n12) RequestResponse(1001, 5001.*, Regexp)\n13) wait(3000)\n14) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}2, Regexp)\n15) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
                ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW) + ' in Programming', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Wait 1s\n4) Change to Programming session with Service 0x10 02\n5) Wait 5s\n6) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + ' using physical addressing\n7) Wait 3s\n8) Change to Default session with Service 0x10 01\n9) Wait 1s\n10) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + 'using physical addressing', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -', '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(1002, 5002.*, Regexp)\n5) wait(5000)\n6) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)\n7) wait(3000)\n8) RequestResponse(1001, 5001.*, Regexp)\n9) wait(1000)\n10) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

            if str(row_list_type_baseSW) == "RBEOL":
                ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW) + ' in RBEOL', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID ' +
                    str(row_list_DID_baseSW) + ' using physical addressing\n5) Reset ECU\n6) Wait 3s\n7) Send 1001\n8) Wait 3s', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -',
                    '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)\n5) RequestResponse(1101, 5101, Equal)\n6) wait(3000)\n7) RequestResponse(1001, 5001.*, Regexp)\n8) wait(3000)', 'Automated Testcase', 'implemented', baseSW, ''])
        if c == 0:
            
            if str(row_list_type_baseSW) == "APPL":
                
                ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW)+' in Application', '1) Send service 0x22 to the camera for the DID ' +
                            str(row_list_DID_baseSW) + ' using physical addressing', '1) -', '1) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Equal)', 'Automated Testcase', 'implemented', baseSW, ''])
            
            if str(row_list_type_baseSW) == "FBL":
                # ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW) + ' in Programming', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Wait 3s\n4) Check active session should be Extended\n4) Wait 3s\n5) Change to Programming session with Service 0x10 02\n6) Wait 3s\n7) Check active session should be Programming\n8) Wait 3s\n9) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + ' using physical addressing\n10) Wait 3s\n11) Change to Default session with Service 0x10 01\n12) Wait 3s\n13) Check active session should be Default\n14) Wait 3s\n15) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + 'using physical addressing', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -\n13) -\n14) -\n15) -', '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(3000)\n5) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}3, Regexp)\n6) wait(1000)\n7) RequestResponse(1002, 5002.*, Regexp)\n8) wait(3000)\n9) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}2, Regexp)\n10) wait(3000)\n11) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)\n12) RequestResponse(1001, 5001.*, Regexp)\n13) wait(3000)\n14) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}2, Regexp)\n15) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
                ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW) + ' in Programming', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Wait 1s\n4) Change to Programming session with Service 0x10 02\n5) Wait 5s\n6) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + ' using physical addressing\n7) Wait 3s\n8) Change to Default session with Service 0x10 01\n9) Wait 3s\n10) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + 'using physical addressing', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -', '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(1002, 5002.*, Regexp)\n5) wait(5000)\n6) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Equal)\n7) wait(3000)\n8) RequestResponse(1001, 5001.*, Regexp)\n9) wait(1000)\n10) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

            if str(row_list_type_baseSW) == "RBEOL":
                ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW) + ' in RBEOL', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID ' +
                    str(row_list_DID_baseSW) + ' using physical addressing\n5) Reset ECU\n6) Wait 3s\n7) Send 1001\n8) Wait 3s', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -',
                    '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Equal)\n5) RequestResponse(1101, 5101, Equal)\n6) wait(3000)\n7) RequestResponse(1001, 5001.*, Regexp)\n8) wait(3000)', 'Automated Testcase', 'implemented', baseSW, ''])
            
        number4 += 1
        o += 1
        j += 1
        k += 1

    return id

def DID_latestSW(ws_RF_LatestSW, wb_RF_LatestSW, id, number1, number2, number3, number4, direct,dummy):
    if direct == '':
        wb_RF_LatestSW = load_workbook('BSWAvalue.xlsx')
        wb_General_Infomation = load_workbook('BSWAvalue.xlsx')
    else:
        wb_RF_LatestSW = load_workbook(str(direct))
        wb_General_Infomation = load_workbook(str(direct))

    ws_RF_LatestSW = wb_RF_LatestSW .active
    ws_General_Infomation = wb_General_Infomation.active

    ws_RF_LatestSW = wb_RF_LatestSW['RFvalue_latestSW']
    ws_General_Infomation = wb_General_Infomation['General Information']

    sheet_RF_LatestSW = wb_RF_LatestSW.worksheets[1]
    row_count_RF_LatestSW = sheet_RF_LatestSW.max_row

    for row in range(2, 3):
        for col in range(2, 3):
            char = get_column_letter(col)
            latestSW = ws_General_Infomation[char + str(row)].value

    i = 0
    o = 2
    j = 3
    k = 1
    c = 0
    status_dummy = 0
    number4 += 1

    i = 0

    while k < row_count_RF_LatestSW:
        for row in range(o, j):
            for col in range(1, 2):
                char = get_column_letter(col)
                row_list_DID_latestSW = ws_RF_LatestSW[char + str(row)].value
                row_list_DID_latestSW_lowercase = str(
                    row_list_DID_latestSW).lower()
                if str(row_list_DID_latestSW) == "None":
                    return id
                # print(ws_RF_LatestSW[char + str(row)].value)
        for row in range(o, j):
            for col in range(2, 3):
                char = get_column_letter(col)
                row_list_name_latestSW = ws_RF_LatestSW[char + str(row)].value
                if str(row_list_DID_latestSW) == "None":
                    return id
                # print(ws_RF_LatestSW[char + str(row)].value)
        for row in range(o, j):
            for col in range(3, 4):
                char = get_column_letter(col)
                row_list_length_byte_latestSW = ws_RF_LatestSW[char + str(row)].value
                if str(row_list_DID_latestSW) == "None":
                    return id
                # print(ws[char + str(row)].value)
                # print(row_list_length_byte_latestSW)
        for row in range(o, j):
            for col in range(4, 5):
                char = get_column_letter(col)
                row_list_ASCII_values_latestSW = ws_RF_LatestSW[char + str(row)].value
                if str(row_list_DID_latestSW) == "None":
                    return id
                # print(ws_RF_LatestSW[char + str(row)].value)
        
        for row in range(o, j):
            for col in range(5, 6):
                char = get_column_letter(col)
                row_list_hex_values_latestSW = ws_RF_LatestSW[char + str(row)].value

        for row in range(o, j):
            for col in range(6, 7):
                char = get_column_letter(col)
                row_list_type_latestSW = ws_RF_LatestSW[char + str(row)].value
                # print("latest" + str(row_list_type_latestSW))

        for row in range(o, j):
            for col in range(7, 8):
                char = get_column_letter(col)
                row_list_hex_values_DummySW = ws_RF_LatestSW[char + str(row)].value
                # print("latest" + str(row_list_hex_values_DummySW))
        
        if str(row_list_DID_latestSW) != "None" :
            id += 1
            # check lenghth byte
            hexvalue_latestSW = ""
            
            # length_byte = ""
            if dummy == "dummy":
                if str(row_list_name_latestSW) == "Supplier Software number":
                    if str(row_list_length_byte_latestSW) != 'None':
                        length_byte = int(row_list_length_byte_latestSW) * 2
                        # print(length_byte)
                        hexvalue_latestSW = str(".{" + str(length_byte) + "}")
                    else:
                        hexvalue_latestSW = str(".*")
                    # print('dung')
                    #print(hexvalue_latestSW)
                    dummy = ""
                    # c = 1
                    status_dummy  = 1
                    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_latestSW) + ' ' + str(row_list_name_latestSW), 'To check value of the DID ' + str(row_list_DID_latestSW), '1) Send service 0x22 to the camera for the DID ' +
                        str(row_list_DID_latestSW) + ' using physical addressing', '1) -', '1) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])                                    
                    
            else:
                if status_dummy == 0:
                    if str(row_list_ASCII_values_latestSW) == 'None' and str(row_list_hex_values_latestSW) == 'None':
                        if str(row_list_length_byte_latestSW) != 'None':
                            length_byte = int(row_list_length_byte_latestSW) * 2
                            #print(length_byte)
                            hexvalue_latestSW = str(".{" + str(length_byte) + "}")
                        else:
                            hexvalue_latestSW = str(".*")
                        # print('dung')
                        # print(hexvalue_latestSW)
                        c = 1
                    else:
                        if str(row_list_hex_values_latestSW) == "None":
                            # change ascii sang hex value
                            for i in str(row_list_ASCII_values_latestSW):
                                hexvalue_latestSW += hex(ord(i))[2:]
            
                        # danh cho co length byter thi su dung
                            if str(row_list_length_byte_latestSW) != 'None':
                                # print(hexvalue_latestSW)
                                count_hexvalue_latestSW = len(hexvalue_latestSW)
                                count_hexvalue_latestSW = int(count_hexvalue_latestSW) // 2
                                # print(count_hexvalue_latestSW)
                                # print(type(row_list_length_byte_latestSW))
                                if str(count_hexvalue_latestSW) < row_list_length_byte_latestSW:
                                    # print("Day la do dai byte",row_list_length_byte_latestSW)
                                    length_byte = (int(row_list_length_byte_latestSW) -
                                                int(count_hexvalue_latestSW)) * 2
                                    # print("byte bi thieu", length_byte)
                                    hexvalue_latestSW = hexvalue_latestSW.lower()
                                    hexvalue_latestSW = str(hexvalue_latestSW + ".{" + str(length_byte) + "}")
                                    c = 1
                                else:
                                    c = 0
                            if str(row_list_length_byte_latestSW) == 'None':
                                hexvalue_latestSW = hexvalue_latestSW.lower()
                                hexvalue_latestSW = str(hexvalue_latestSW + ".*")
                                c = 1
                            
                        else:
                        # danh cho co length byter thi su dung
                            if str(row_list_length_byte_latestSW) != 'None':
                                count_hexvalue_latestSW = len(row_list_hex_values_latestSW)
                                count_hexvalue_latestSW = int(count_hexvalue_latestSW) // 2
                                if str(count_hexvalue_latestSW) < row_list_length_byte_latestSW:
                                    length_byte = (int(row_list_length_byte_latestSW) - int(count_hexvalue_latestSW)) * 2
                                    hexvalue_latestSW = row_list_hex_values_latestSW.lower()
                                    hexvalue_latestSW = str(hexvalue_latestSW + ".{" + str(length_byte) + "}")
                                    c = 1
                            else:    
                                hexvalue_latestSW = row_list_hex_values_latestSW.lower()
                                hexvalue_latestSW = str(hexvalue_latestSW + ".*")
                                c = 1

                            # hexvalue_latestSW = row_list_hex_values_latestSW.lower()
                            # hexvalue_latestSW = str(hexvalue_latestSW + ".*")
                            # c = 1

                if  status_dummy  == 1:
                    if str(row_list_hex_values_DummySW) == 'None':
                        if str(row_list_ASCII_values_latestSW) == 'None' and str(row_list_hex_values_latestSW) == 'None':
                            if str(row_list_length_byte_latestSW) != 'None':
                                length_byte = int(row_list_length_byte_latestSW) * 2
                                #print(length_byte)
                                hexvalue_latestSW = str(".{" + str(length_byte) + "}")
                            else:
                                hexvalue_latestSW = str(".*")
                            c = 1
                            
                        else:
                            if str(row_list_hex_values_latestSW) == "None":
                                # change ascii sang hex value
                                for i in str(row_list_ASCII_values_latestSW):
                                    hexvalue_latestSW += hex(ord(i))[2:]
                                    
                            # danh cho co length byter thi su dung
                                if str(row_list_length_byte_latestSW) != 'None':
                                    # print(hexvalue_latestSW)
                                    count_hexvalue_latestSW = len(hexvalue_latestSW)
                                    count_hexvalue_latestSW = int(count_hexvalue_latestSW) // 2
                                    # print(count_hexvalue_latestSW)
                                    # print(type(row_list_length_byte_latestSW))
                                    if str(count_hexvalue_latestSW) < row_list_length_byte_latestSW:
                                        # print("Day la do dai byte",row_list_length_byte_latestSW)
                                        length_byte = (int(row_list_length_byte_latestSW) - int(count_hexvalue_latestSW)) * 2
                                        # print("byte bi thieu", length_byte)
                                        hexvalue_latestSW = hexvalue_latestSW.lower()
                                        hexvalue_latestSW = str(hexvalue_latestSW + ".{" + str(length_byte) + "}")
                                        c = 1
                                    else:
                                        c = 0
                                if str(row_list_length_byte_latestSW) == 'None':
                                    hexvalue_latestSW = hexvalue_latestSW.lower()
                                    hexvalue_latestSW = str(hexvalue_latestSW + ".*")
                                    c = 1
                                else:
                                    c = 0
                            else:
                                if str(row_list_length_byte_latestSW) == 'None':
                                    hexvalue_latestSW = row_list_hex_values_latestSW.lower()
                                    hexvalue_latestSW = str(hexvalue_latestSW + ".*")
                    else:
                        # if str(row_list_length_byte_latestSW) != 'None':
                        #     count_hexvalue_DummySW = len(row_list_hex_values_DummySW)
                        #     count_hexvalue_DummySW = int(count_hexvalue_DummySW) // 2
                        #     if str(count_hexvalue_DummySW) < row_list_length_byte_latestSW:
                        #         length_byte = (int(row_list_length_byte_latestSW) - int(count_hexvalue_DummySW)) * 2
                        #         hexvalue_latestSW = row_list_hex_values_DummySW.lower()
                        #         hexvalue_latestSW = str(hexvalue_latestSW + ".{" + str(length_byte) + "}")
                        #         c = 1
                        # else:
                        #     hexvalue_latestSW = row_list_hex_values_DummySW.lower()
                        #     hexvalue_latestSW = str(hexvalue_latestSW + ".*")
                        #     c = 1

                        hexvalue_latestSW = row_list_hex_values_DummySW.lower()
                        hexvalue_latestSW = str(hexvalue_latestSW + ".*")     
                        c = 1
                    # print(hexvalue_latestSW)
            
            # if str(row_list_name_latestSW) == "None":
            #     row_list_name_latestSW = ""
            
            # if c == 1:
            #     ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_latestSW) + ' ' + str(row_list_name_latestSW), 'To check value of the DID ' + str(row_list_DID_latestSW), '1) Send service 0x22 to the camera for the DID ' +
            #                 str(row_list_DID_latestSW) + ' using physical addressing', '1) -', '1) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
            # if c == 0:
            #     ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_latestSW) + ' ' + str(row_list_name_latestSW), 'To check value of the DID ' + str(row_list_DID_latestSW), '1) Send service 0x22 to the camera for the DID ' +
            #                 str(row_list_DID_latestSW) + ' using physical addressing', '1) -', '1) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Equal)', 'Automated Testcase', 'implemented', latestSW, ''])
            
                if str(row_list_name_latestSW) == "None":
                    row_list_name_latestSW = ""
            
                # if str(row_list_name_baseSW) != "Supplier Software number":
                
                if c == 1:
                    if str(row_list_type_latestSW) == "APPL":

                        ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_latestSW) + ' ' + str(row_list_name_latestSW), 'To check value of the DID ' + str(row_list_DID_latestSW), '1) Send service 0x22 to the camera for the DID ' +
                            str(row_list_DID_latestSW) + ' using physical addressing', '1) -', '1) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

                    if str(row_list_type_latestSW) == "FBL":
                        # ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW) + ' in Programming', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Wait 3s\n4) Check active session should be Extended\n4) Wait 3s\n5) Change to Programming session with Service 0x10 02\n6) Wait 3s\n7) Check active session should be Programming\n8) Wait 3s\n9) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + ' using physical addressing\n10) Wait 3s\n11) Change to Default session with Service 0x10 01\n12) Wait 3s\n13) Check active session should be Default\n14) Wait 3s\n15) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + 'using physical addressing', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -\n13) -\n14) -\n15) -', '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(3000)\n5) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}3, Regexp)\n6) wait(1000)\n7) RequestResponse(1002, 5002.*, Regexp)\n8) wait(3000)\n9) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}2, Regexp)\n10) wait(3000)\n11) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)\n12) RequestResponse(1001, 5001.*, Regexp)\n13) wait(3000)\n14) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}2, Regexp)\n15) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
                        ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_latestSW) + ' ' + str(row_list_name_latestSW), 'To check value of the DID ' + str(row_list_DID_latestSW) + ' in Programming', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Wait 1s\n4) Change to Programming session with Service 0x10 02\n5) Wait 5s\n6) Send service 0x22 to the camera for the DID ' +str(row_list_DID_latestSW) + ' using physical addressing\n7) Wait 3s\n8) Change to Default session with Service 0x10 01\n9) Wait 1s\n10) Send service 0x22 to the camera for the DID ' +str(row_list_DID_latestSW) + 'using physical addressing', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -', '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(1002, 5002.*, Regexp)\n5) wait(5000)\n6) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Regexp)\n7) wait(3000)\n8) RequestResponse(1001, 5001.*, Regexp)\n9) wait(1000)\n10) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

                    if str(row_list_type_latestSW) == "RBEOL":
                        ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_latestSW) + ' ' + str(row_list_name_latestSW), 'To check value of the DID ' + str(row_list_DID_latestSW) + ' in RBEOL', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID ' +
                            str(row_list_DID_latestSW) + ' using physical addressing\n5) Reset ECU\n6) Wait 3s\n7) Send 1001\n8) Wait 3s', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -',
                    '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Regexp)\n5) RequestResponse(1101, 5101, Equal)\n6) wait(3000)\n7) RequestResponse(1001, 5001.*, Regexp)\n8) wait(3000)', 'Automated Testcase', 'implemented', latestSW, ''])

                if c == 0:
                    if str(row_list_type_latestSW) == "APPL":

                        ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_latestSW) + ' ' + str(row_list_name_latestSW), 'To check value of the DID ' + str(row_list_DID_latestSW), '1) Send service 0x22 to the camera for the DID ' +
                            str(row_list_DID_latestSW) + ' using physical addressing', '1) -', '1) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Equal)', 'Automated Testcase', 'implemented', latestSW, ''])

                    if str(row_list_type_latestSW) == "FBL":
                        # ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW) + ' in Programming', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Wait 3s\n4) Check active session should be Extended\n4) Wait 3s\n5) Change to Programming session with Service 0x10 02\n6) Wait 3s\n7) Check active session should be Programming\n8) Wait 3s\n9) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + ' using physical addressing\n10) Wait 3s\n11) Change to Default session with Service 0x10 01\n12) Wait 3s\n13) Check active session should be Default\n14) Wait 3s\n15) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + 'using physical addressing', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -\n13) -\n14) -\n15) -', '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(3000)\n5) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}3, Regexp)\n6) wait(1000)\n7) RequestResponse(1002, 5002.*, Regexp)\n8) wait(3000)\n9) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}2, Regexp)\n10) wait(3000)\n11) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)\n12) RequestResponse(1001, 5001.*, Regexp)\n13) wait(3000)\n14) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}2, Regexp)\n15) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
                        ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_latestSW) + ' ' + str(row_list_name_latestSW), 'To check value of the DID ' + str(row_list_DID_latestSW) + ' in Programming', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Wait 1s\n4) Change to Programming session with Service 0x10 02\n5) Wait 5s\n6) Send service 0x22 to the camera for the DID ' +str(row_list_DID_latestSW) + ' using physical addressing\n7) Wait 3s\n8) Change to Default session with Service 0x10 01\n9) Wait 1s\n10) Send service 0x22 to the camera for the DID ' +str(row_list_DID_latestSW) + 'using physical addressing', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -', '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(1002, 5002.*, Regexp)\n5) wait(5000)\n6) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Equal)\n7) wait(3000)\n8) RequestResponse(1001, 5001.*, Regexp)\n9) wait(1000)\n10) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Equal)', 'Automated Testcase', 'implemented', latestSW, ''])

                    if str(row_list_type_latestSW) == "RBEOL":
                        ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_latestSW) + ' ' + str(row_list_name_latestSW), 'To check value of the DID ' + str(row_list_DID_latestSW) + ' in RBEOL', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID ' +
                            str(row_list_DID_latestSW) + ' using physical addressing\n5) Reset ECU\n6) Wait 3s\n7) Send 1001\n8) Wait 3s', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -',
                    '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Equal)\n5) RequestResponse(1101, 5101, Equal)\n6) wait(3000)\n7) RequestResponse(1001, 5001.*, Regexp)\n8) wait(3000)', 'Automated Testcase', 'implemented', latestSW, ''])
            
            number4 += 1
            o += 1
            j += 1
            k += 1
        else:
            return id
    return id


def variant_base_sw(id, number1, number2, number3, number4, tasks):
    direct = Input_path_text.get()
    if direct == '':
        wb_RF_BaseSW = load_workbook('BSWAvalue.xlsx')
        wb_General_Infomation = load_workbook('BSWAvalue.xlsx')
    else:
        wb_RF_BaseSW = load_workbook(str(direct))
        wb_General_Infomation = load_workbook(str(direct))

    ws_RF_BaseSW = wb_RF_BaseSW.active
    ws_General_Infomation = wb_General_Infomation.active

    ws_RF_BaseSW = wb_RF_BaseSW['RFvalue_baseSW']
    ws_General_Infomation = wb_General_Infomation['General Information']

    for row in range(1, 2):
        for col in range(2, 3):
            char = get_column_letter(col)
            baseSW = ws_General_Infomation[char + str(row)].value
            # print(baseSW)

    for row in range(3, 4):
        for col in range(7, 8):
            char = get_column_letter(col)
            row_Variant_BaseSW = ws_RF_BaseSW[char + str(row)].value
            row_Variant_BaseSW_lowercase = str(row_Variant_BaseSW).lower()
            # print(row_Variant_BaseSW)
    for row in range(4, 5):
        for col in range(7, 8):
            char = get_column_letter(col)
            row_DID_check_variant_BaseSW = ws_RF_BaseSW[char + str(row)].value
            row_DID_check_variant_BaseSW_lowercase = str(
                row_DID_check_variant_BaseSW).lower()
            # print(row_DID_check_variant_BaseSW)
    
    for row in range(5, 6):
        for col in range(7, 8):
            char = get_column_letter(col)
            row_Security_level_BaseSW = ws_RF_BaseSW[char + str(row)].value
            if str(row_Security_level_BaseSW) == "None":
                messagebox.showerror("ERROR", "No Value SECURITY LEVEL, Please add SECURITY LEVEL")
                break            
            # row_DID_check_variant_BaseSW_lowercase = str(
            #     row_DID_check_variant_BaseSW).lower()
    
    if str(row_Security_level_BaseSW) == '0':
        if str(row_Variant_BaseSW) != "None" and str(row_DID_check_variant_BaseSW) != "None":
            ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) wait\n4) Select variant\n5) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(2e' + str(row_Variant_BaseSW_lowercase) + ', 6e' + str(row_DID_check_variant_BaseSW_lowercase) + ', Equal)\n5) RequestResponse(22' + str(row_DID_check_variant_BaseSW_lowercase) + ', 62' + str(row_Variant_BaseSW_lowercase) + ', Equal)', 'Automated Testcase', 'implemented', baseSW, ''])
        else:
            ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check_variant (Variant is default)', 'To check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) wait\n4) Check variant', '1) -\n2) -\n3) -\n4) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(22' + str(row_DID_check_variant_BaseSW_lowercase) + ', 62.*, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    if str(row_Security_level_BaseSW) == '1':
        if str(row_Variant_BaseSW) != "None" and str(row_DID_check_variant_BaseSW) != "None":
            ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) wait\n7) Select variant\n8) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel'+str(row_Security_level_BaseSW)+'(1;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel'+str(row_Security_level_BaseSW)+'(0;0))\n6) wait(1000)\n7) RequestResponse(2e' + str(row_Variant_BaseSW_lowercase) + ', 6e' + str(row_DID_check_variant_BaseSW_lowercase) + ', Equal)\n8) RequestResponse(22' + str(row_DID_check_variant_BaseSW_lowercase) + ', 62' + str(row_Variant_BaseSW_lowercase) + ', Equal)', 'Automated Testcase', 'implemented', baseSW, ''])
        else:
            ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check_variant (Variant is default)', 'To check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) wait\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel'+str(row_Security_level_BaseSW)+'(1;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel'+str(row_Security_level_BaseSW)+'(0;0))\n6) wait(1000)\n7) RequestResponse(22' + str(row_DID_check_variant_BaseSW_lowercase) + ', 62.*, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    if str(row_Security_level_BaseSW) == '0':
        if str(row_Variant_BaseSW) != "None" and str(row_DID_check_variant_BaseSW) != "None":
            ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) wait\n4) Select variant\n5) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(2e' + str(row_Variant_BaseSW_lowercase) + ', 6e' + str(row_DID_check_variant_BaseSW_lowercase) + ', Equal)\n5) RequestResponse(22' + str(row_DID_check_variant_BaseSW_lowercase) + ', 62' + str(row_Variant_BaseSW_lowercase) + ', Equal)', 'Automated Testcase', 'implemented', baseSW, ''])
        else:
            ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check_variant (Variant is default)', 'To check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) wait\n4) Check variant', '1) -\n2) -\n3) -\n4) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(22' + str(row_DID_check_variant_BaseSW_lowercase) + ', 62.*, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    if str(row_Security_level_BaseSW) == '1':
        if str(row_Variant_BaseSW) != "None" and str(row_DID_check_variant_BaseSW) != "None":
            ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) wait\n7) Select variant\n8) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel'+str(row_Security_level_BaseSW)+'(1;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel'+str(row_Security_level_BaseSW)+'(0;0))\n6) wait(1000)\n7) RequestResponse(2e' + str(row_Variant_BaseSW_lowercase) + ', 6e' + str(row_DID_check_variant_BaseSW_lowercase) + ', Equal)\n8) RequestResponse(22' + str(row_DID_check_variant_BaseSW_lowercase) + ', 62' + str(row_Variant_BaseSW_lowercase) + ', Equal)', 'Automated Testcase', 'implemented', baseSW, ''])
        else:
            ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check_variant (Variant is default)', 'To check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) wait\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel'+str(row_Security_level_BaseSW)+'(1;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel'+str(row_Security_level_BaseSW)+'(0;0))\n6) wait(1000)\n7) RequestResponse(22' + str(row_DID_check_variant_BaseSW_lowercase) + ', 62.*, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    return id


def variant_latest_sw(id, number1, number2, number3, number4, tasks):
    direct = Input_path_text.get()
    if direct == '':
        wb_RF_LatestSW = load_workbook('BSWAvalue.xlsx')
        wb_General_Infomation = load_workbook('BSWAvalue.xlsx')
    else:
        wb_RF_LatestSW = load_workbook(str(direct))
        wb_General_Infomation = load_workbook(str(direct))

    ws_RF_LatestSW = wb_RF_LatestSW.active
    ws_General_Infomation = wb_General_Infomation.active

    ws_RF_LatestSW = wb_RF_LatestSW['RFvalue_latestSW']
    ws_General_Infomation = wb_General_Infomation['General Information']
    sheet2 = wb_RF_LatestSW.worksheets[1]
    
    row_count = sheet2.max_row
    for row in range(2, 3):
        for col in range(2, 3):
            char = get_column_letter(col)
            latestSW = ws_General_Infomation[char + str(row)].value
            # print(latestSW)

    for row in range(3, 4):
        for col in range(8, 9):
            char = get_column_letter(col)
            row_Variant_LatestSW = ws_RF_LatestSW[char + str(row)].value
            row_Variant_LatestSW_lowercase = str(row_Variant_LatestSW).lower()
            # print(row_Variant_LatestSW)
    for row in range(4, 5):
        for col in range(8, 9):
            char = get_column_letter(col)
            row_DID_check_variant_LatestSW = ws_RF_LatestSW[char + str(row)].value
            row_DID_check_variant_LatestSW_lowercase = str(
                row_DID_check_variant_LatestSW).lower()
            # print(row_DID_check_variant_LatestSW)

    for row in range(5, 6):
        for col in range(8, 9):
            char = get_column_letter(col)
            row_Security_level_latestSW = ws_RF_LatestSW[char + str(row)].value
            # row_DID_check_variant_BaseSW_lowercase = str(
            #     row_DID_check_variant_BaseSW).lower()
    
    if str(row_Security_level_latestSW) == '0':
        if str(row_Variant_LatestSW) != "None" and str(row_DID_check_variant_LatestSW) != "None":
            ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) wait\n4) Select variant\n5) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(2e' + str(row_Variant_LatestSW_lowercase) + ', 6e' + str(row_DID_check_variant_LatestSW_lowercase) + ', Equal)\n5) RequestResponse(22' + str(row_DID_check_variant_LatestSW_lowercase) + ', 62' + str(row_Variant_LatestSW_lowercase) + ', Equal)', 'Automated Testcase', 'implemented', latestSW, ''])
        else:
            ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check_variant (Variant is default)', 'To check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) wait\n4) Check variant', '1) -\n2) -\n3) -\n4) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(22' + str(row_DID_check_variant_LatestSW_lowercase) + ', 62.*, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    if str(row_Security_level_latestSW) != '0':
        if str(row_Variant_LatestSW) != "None" and str(row_DID_check_variant_LatestSW) != "None":
            ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) wait\n7) Select variant\n8) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel'+str(row_Security_level_latestSW)+'(1;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel'+str(row_Security_level_latestSW)+'(0;0))\n6) wait(1000)\n7) RequestResponse(2e' + str(row_Variant_LatestSW_lowercase) + ', 6e' + str(row_DID_check_variant_LatestSW_lowercase) + ', Equal)\n8) RequestResponse(22' + str(row_DID_check_variant_LatestSW_lowercase) + ', 62' + str(row_Variant_LatestSW_lowercase) + ', Equal)', 'Automated Testcase', 'implemented', latestSW, ''])
        else:
            ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check_variant (Variant is default)', 'To check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) wait\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel'+str(row_Security_level_latestSW)+'(1;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel'+str(row_Security_level_latestSW)+'(0;0))\n6) wait(1000)\n7) RequestResponse(22' + str(row_DID_check_variant_LatestSW_lowercase) + ', 62.*, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    if str(row_Security_level_latestSW) == '0':
        if str(row_Variant_LatestSW) != "None" and str(row_DID_check_variant_LatestSW) != "None":
            ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) wait\n4) Select variant\n5) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(2e' + str(row_Variant_LatestSW_lowercase) + ', 6e' + str(row_DID_check_variant_LatestSW_lowercase) + ', Equal)\n5) RequestResponse(22' + str(row_DID_check_variant_LatestSW_lowercase) + ', 62' + str(row_Variant_LatestSW_lowercase) + ', Equal)', 'Automated Testcase', 'implemented', latestSW, ''])
        else:
            ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check_variant (Variant is default)', 'To check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) wait\n4) Check variant', '1) -\n2) -\n3) -\n4) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(22' + str(row_DID_check_variant_LatestSW_lowercase) + ', 62.*, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    if str(row_Security_level_latestSW) != '0':
        if str(row_Variant_LatestSW) != "None" and str(row_DID_check_variant_LatestSW) != "None":
            ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) wait\n7) Select variant\n8) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel'+str(row_Security_level_latestSW)+'(1;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel'+str(row_Security_level_latestSW)+'(0;0))\n6) wait(1000)\n7) RequestResponse(2e' + str(row_Variant_LatestSW_lowercase) + ', 6e' + str(row_DID_check_variant_LatestSW_lowercase) + ', Equal)\n8) RequestResponse(22' + str(row_DID_check_variant_LatestSW_lowercase) + ', 62' + str(row_Variant_LatestSW_lowercase) + ', Equal)', 'Automated Testcase', 'implemented', latestSW, ''])
        else:
            ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check_variant (Variant is default)', 'To check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) wait\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel'+str(row_Security_level_latestSW)+'(1;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel'+str(row_Security_level_latestSW)+'(0;0))\n6) wait(1000)\n7) RequestResponse(22' + str(row_DID_check_variant_LatestSW_lowercase) + ', 62.*, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])


    return id

# SOURCE

def TC_RF():
    run_btn_text.set("Loading...")

    direct = Input_path_text.get()
    locate_save = Output_path_text.get()


    if direct == '':
        wb_RF_BaseSW = load_workbook('BSWAvalue.xlsx')
        wb_RF_LatestSW = load_workbook('BSWAvalue.xlsx')
        wb_General_Infomation = load_workbook('BSWAvalue.xlsx')
        wb_TC_RF_Clear = load_workbook('TC_RF.xlsx')
    else:
        wb_RF_BaseSW = load_workbook(direct)
        wb_RF_LatestSW = load_workbook(direct)
        wb_General_Infomation = load_workbook(direct)
        wb_TC_RF_Clear = load_workbook(locate_save + '/'+'TC_RF.xlsx')
    # wb_RF_LatestSW = load_workbook(direct)

    ws_RF_BaseSW = wb_RF_BaseSW.active
    ws_RF_LatestSW = wb_RF_LatestSW.active
    ws_General_Infomation = wb_General_Infomation.active

    ws_RF_BaseSW = wb_RF_BaseSW['RFvalue_baseSW']
    ws_RF_LatestSW = wb_RF_LatestSW['RFvalue_latestSW']
    ws_General_Infomation = wb_General_Infomation['General Information']
    
    sheet_RF_BaseSW = wb_RF_BaseSW.worksheets[0]
    sheet_RF_LatestSW = wb_RF_BaseSW.worksheets[1]
    row_count_RF_BaseSW = sheet_RF_BaseSW.max_row
    row_count_RF_LatestSW = sheet_RF_LatestSW.max_row
    tasks = row_count_RF_LatestSW + row_count_RF_BaseSW + 160

    ws_TC_RF_Clear = wb_TC_RF_Clear.active
    ws_TC_RF_Clear = wb_TC_RF_Clear['TC_RF']
    sheet_TC_RF_Clear = wb_TC_RF_Clear.worksheets[0]
    row_count_TC_RF_Clear = sheet_TC_RF_Clear.max_row

    print(row_count_TC_RF_Clear)
    print(tasks)

#  clear old TC_RF
    if row_count_TC_RF_Clear >= tasks:
        n = 0
        while n < row_count_TC_RF_Clear:
            ws_TC_RF.delete_rows(1)
            n += 1
            print("dang xoa")
        print("done")
        n = 0

# baseSW name
    for row in range(1, 2):
        for col in range(2, 3):
            char = get_column_letter(col)
            baseSW = ws_General_Infomation[char + str(row)].value

# latestSW name
    for row in range(2, 3):
        for col in range(2, 3):
            char = get_column_letter(col)
            latestSW = ws_General_Infomation[char + str(row)].value

# ticket
    for row in range(3, 4):
        for col in range(2, 3):
            char = get_column_letter(col)
            ticket_baseSW = ws_General_Infomation[char + str(row)].value
            ticket_latestSW = ws_General_Infomation[char + str(row)].value

# Programming_counter_DID
    for row in range(4, 5):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_counter_DID = ws_General_Infomation[char + str(row)].value

# Programming_Attempt_counter_DID
    for row in range(5, 6):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_Attempt_counter_DID = ws_General_Infomation[char + str(row)].value

# Programming_counter_step
    for row in range(6, 7):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_counter_step = ws_General_Infomation[char + str(row)].value

# Programming_Attempt_counter_step
    for row in range(7, 8):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_Attempt_counter_step = ws_General_Infomation[char + str(row)].value

    for row in range(8, 9):
        for col in range(2, 3):
            char = get_column_letter(col)
            row_DID_Check_Active_Session = ws_General_Infomation[char + str(row)].value
            row_DID_Check_Active_Session_lowercase = str(row_DID_Check_Active_Session).lower()

    # script begin
    id = 2
    number1 = 1
    number2 = 1
    number3 = 1
    number4 = 1

    border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(
        border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
    font_text_header = Font(name="Calibri", size=13,
                            color='00FFFFFF', bold=True)
    font_text = Font(name="Calibri", size=11, color='00000000', bold=False)
    alignment = Alignment(horizontal='center', vertical='center')
    ws_TC_RF.append(['ID', 'XXX Component',  'Test Description', 'Test Steps',  'Test Response','Teststep keywords', 'ObjectType', 'TestStatus', 'Project', 'TestResult'])
    
    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(1, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text_header
        # cell_header.alignment = alignment
    ws_TC_RF.append(['ID_'+str(id),  '1 REFFLASH', '','', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(2, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        # cell_header.alignment = alignment
    # ------------------------------------------------------------------------------------------------------
    # BEGIN TEST CASE 1
    # TEST CASE 1 base SW to latestSW M3
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) +' Base SW to Latest SW M3', '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH BASE_SW VIA UART script
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) +' Flash Base SW via UART', '', '', '', '', 'Test group', '', '', ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])

    for row in range(id - 2, id + 1):
        for col in range(1, 11):
            cell_header = ws_TC_RF.cell(row, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART flash ' + str(baseSW), 'Detail information is mentioned in the ticket: ' +str(ticket_baseSW), '1) Flash Base Software '+ str(baseSW) +' via UART successful', '1) -', "1) TesterConfirm('Do you flash base software "+str(baseSW)+" via UART ?')", 'Automated Testcase', 'implemented', str(baseSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification','', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_base_sw(id, number1, number2, number3, number4, tasks)

    # id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    # number4 += 1
    direct = Input_path_text.get()
    id = DID_baseSW(ws_TC_RF, wb_RF_BaseSW, id, number1, number2, number3, number4, direct)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])
    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    # print(id)

    PC_step = 0
    PAC_step = 0
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL','', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # ws_TC_RF.title = "TC_RF"
    # wb_TC_RF.save('TC_RF.xlsx')

    # Step2 FLASH LATEST_SW M3 VIA Xflash TOOLS

    ws_RF_LatestSW = wb_RF_LatestSW.active
    ws_RF_LatestSW = wb_RF_LatestSW['RFvalue_latestSW']

    sheet_RF_LatestSW = wb_RF_LatestSW.worksheets[1]
    row_count_RF_LatestSW = sheet_RF_LatestSW.max_row

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Latest SW M3 via X-Flash 1st',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via IFLASH successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via IFLASH ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    dummy  = ""
    id = DID_latestSW(ws_TC_RF, wb_RF_LatestSW, id, number1, number2, number3, number4, direct, dummy)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = Programming_counter_step
    PAC_step = Programming_counter_step
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL','', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # ws_TC_RF.title = "TC_RF"
    # wb_TC_RF.save('TC_RF.xlsx')

    # END TEST CASE 1

    # ------------------------------------------------------------------------------------------------------
    # # BEGIN TEST CASE 2
    # # TEST CASE 2 base SW to latestSW M5
    number1 += 1
    number2 = 1
    number3 = 1
    number4 = 1
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) +
                ' Base SW to Latest SW M5', '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH BASE_SW VIA UART script
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Flash Base SW via UART',
                '', '', '', '', 'Test group', '', '', ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Flash SW', '',
                '', '', '', 'Test group', '', '', ''])

    for row in range(id - 2, id + 1):
        for col in range(1, 11):
            cell_header = ws_TC_RF.cell(row, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART flash ' + str(baseSW), 'Detail information is mentioned in the ticket: ' +str(ticket_baseSW), '1) Flash Base Software '+ str(baseSW) +' via UART successful', '1) -', "1) TesterConfirm(''Do you flash base software "+str(baseSW)+" via UART ?')", 'Automated Testcase', 'implemented', str(baseSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_base_sw(id, number1, number2, number3, number4, tasks)

    # id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    # number4 += 1
    direct = Input_path_text.get()
    id = DID_baseSW(ws_TC_RF, wb_RF_BaseSW, id, number1, number2, number3, number4, direct)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = 0
    PAC_step = 0
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # # ws_TC_RF.title = "TC_RF"
    # # wb_TC_RF.save('TC_RF.xlsx')

    # # Step2 FLASH LATEST_SW M5 1st VIA Xflash TOOLS

    # # # Reflash latest SW M5 via xflash tool

    ws_RF_LatestSW = wb_RF_LatestSW.active
    ws_RF_LatestSW = wb_RF_LatestSW['RFvalue_latestSW']

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Latest SW M5 via X-Flash 1st','', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via IFLASH successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via IFLASH ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    # number4 += 1
    direct = Input_path_text.get()
    dummy = ""
    id = DID_latestSW(ws_TC_RF, wb_RF_LatestSW, id, number1, number2, number3, number4, direct, dummy)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = Programming_counter_step
    PAC_step = Programming_Attempt_counter_step
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text


    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*5,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # # Step2 FLASH LATEST_SW M3 2nd VIA Xflash TOOLS

    # # # Reflash latest SW M3 via xflash tool

    ws_RF_LatestSW = wb_RF_LatestSW.active
    ws_RF_LatestSW = wb_RF_LatestSW['RFvalue_latestSW']

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Latest SW M3 via X-Flash 2nd',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via IFLASH successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via IFLASH ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    dummy = ""
    id = DID_latestSW(ws_TC_RF, wb_RF_LatestSW, id, number1, number2, number3, number4, direct, dummy)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = int(Programming_counter_step) + int(Programming_counter_step)
    PAC_step = int(Programming_Attempt_counter_step) + int(Programming_Attempt_counter_step)
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # # ws_TC_RF.title = "TC_RF"
    # # wb_TC_RF.save('TC_RF.xlsx')

    # # END TEST CASE 2

    # # ------------------------------------------------------------------------------------------------------

    # # BEGIN TEST CASE 3
    # # TEST CASE 3 latest SW to DummySW M3
    ws_RF_LatestSW = wb_RF_LatestSW.active
    ws_RF_LatestSW = wb_RF_LatestSW['RFvalue_latestSW']

    number1 += 1
    number2 = 1
    number3 = 1
    number4 = 1
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + ' latest SW to Dummy SW M3',
                '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH latest_SW VIA UART script

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Flash latest SW via UART',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Flash SW', '',
                '', '', '', 'Test group', '', '', ''])

    for row in range(id - 2, id + 1):
        for col in range(1, 11):
            cell_header = ws_TC_RF.cell(row, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART Flash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via UART successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via UART ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    dummy = ""
    id = DID_latestSW(ws_TC_RF, wb_RF_LatestSW, id, number1, number2, number3, number4, direct,dummy)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = 0
    PAC_step = 0
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # # ws_TC_RF.title = "TC_RF"
    # # wb_TC_RF.save('TC_RF.xlsx')

    # # Step2 FLASH DUMMY_SW M3 VIA Xflash TOOLS

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Dummy SW M3 via X-Flash 1st',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via IFLASH successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via IFLASH ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    dummy = "dummy"
    id = DID_latestSW(ws_TC_RF, wb_RF_LatestSW, id, number1, number2, number3, number4, direct, dummy)
    # print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = Programming_counter_step
    PAC_step = Programming_Attempt_counter_step
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # # ws_TC_RF.title = "TC_RF"
    # # wb_TC_RF.save('TC_RF.xlsx')

    # # END TEST CASE 3

    # # ------------------------------------------------------------------------------------------------------
    # # BEGIN TEST CASE 4
    # # TEST CASE 4 latest SW to DummySW M5
    number1 += 1
    number2 = 1
    number3 = 1
    number4 = 1
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + ' latest SW to Dummy SW M5',
                '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH latest_SW VIA UART script

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Flash latest SW via UART',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Flash SW', '',
                '', '', '', 'Test group', '', '', ''])

    for row in range(id - 2, id + 1):
        for col in range(1, 11):
            cell_header = ws_TC_RF.cell(row, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART Flash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via UART successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via UART ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    # number4 += 1
    direct = Input_path_text.get()
    dummy = ""
    id = DID_latestSW(ws_TC_RF, wb_RF_LatestSW, id, number1, number2, number3, number4, direct,dummy)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = 0
    PAC_step = 0
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # ws_TC_RF.title = "TC_RF"
    # wb_TC_RF.save('TC_RF.xlsx')

# --------------------------------------------------------------------------------------------------------
    # Step2 FLASH DUMMY_SW M5 1st VIA Xflash TOOLS
    # # Reflash Dummy SW M5 via xflash tool

    ws_RF_LatestSW = wb_RF_LatestSW .active
    ws_RF_LatestSW = wb_RF_LatestSW['RFvalue_latestSW']

    # number = df.shape[0]
    # print(number)
    sheet_RF_LatestSW = wb_RF_LatestSW.worksheets[1]
    row_count_RF_LatestSW = sheet_RF_LatestSW.max_row

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Dummy SW M5 via X-Flash 1st',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via IFLASH successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via IFLASH ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    # DID_latestSW(ws_RF_LatestSW, wb, id, number1, number2, number3, number4)
    dummy = "dummy"
    id = DID_latestSW(ws_TC_RF, wb_RF_LatestSW, id, number1, number2, number3, number4, direct, dummy)
    # print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = Programming_counter_step
    PAC_step = Programming_Attempt_counter_step
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*5,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # Step2 FLASH DUMMY_SW M3 2nd VIA Xflash TOOLS

    # # Reflash Dummy SW M3 via xflash tool

    ws_RF_LatestSW = wb_RF_LatestSW .active
    ws_RF_LatestSW = wb_RF_LatestSW['RFvalue_latestSW']

    # number = df.shape[0]
    # print(number)
    sheet = wb_RF_LatestSW.worksheets[1]
    row_count_RF_LatestSW = sheet_RF_LatestSW.max_row

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Dummy SW M3 via X-Flash 2nd',
                '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via IFLASH successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via IFLASH ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    # DID_latestSW(ws_RF_LatestSW, wb, id, number1, number2, number3, number4)
    dummy = "dummy"
    id = DID_latestSW(ws_TC_RF, wb_RF_LatestSW, id, number1, number2, number3, number4, direct, dummy)
    # print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = int(Programming_counter_step) + int(Programming_counter_step)
    PAC_step = int(Programming_Attempt_counter_step) + int(Programming_Attempt_counter_step)
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_RF.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    # Check  Check DID in RBEOL
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws_TC_RF.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])
    
    for row in range(id - id + 1, id + 1):
        for col in range(1, 11):
            cell_header = ws_TC_RF.cell(row, col)
            # used hex code for red color
            cell_header.border = border
    # END TEST CASE 4

    print(str(bar['value']))

    locate_save = Output_path_text.get()
    ws_TC_RF.title = "TC_RF"
    print(locate_save)
    if locate_save == '':
        # print('Ok')
        wb_TC_RF.save('TC_RF.xlsx')
    else:
        wb_TC_RF.save(locate_save + '/'+'TC_RF.xlsx')

    if locate_save == '':
        wb_TC_RF_Rebuild_format = load_workbook('TC_RF.xlsx')
    else:
        wb_TC_RF_Rebuild_format = load_workbook(locate_save + '/'+'TC_RF.xlsx')

    ws_TC_RF_Rebuild_format = wb_TC_RF_Rebuild_format['TC_RF']
    sheet_TC_RF_Rebuild_format = wb_TC_RF_Rebuild_format.worksheets[0]
    row_count_TC_RF_Rebuild_format = sheet_TC_RF_Rebuild_format.max_row

    # sheet_TC_RF_Rebuild_format = wb6.worksheets[0]
    sheet_TC_RF_Rebuild_format.column_dimensions['A'].width = 30
    sheet_TC_RF_Rebuild_format.column_dimensions['B'].width = 30
    sheet_TC_RF_Rebuild_format.column_dimensions['C'].width = 30
    sheet_TC_RF_Rebuild_format.column_dimensions['D'].width = 30
    sheet_TC_RF_Rebuild_format.column_dimensions['E'].width = 30
    sheet_TC_RF_Rebuild_format.column_dimensions['F'].width = 30
    sheet_TC_RF_Rebuild_format.column_dimensions['G'].width = 30
    sheet_TC_RF_Rebuild_format.column_dimensions['H'].width = 30
    sheet_TC_RF_Rebuild_format.column_dimensions['I'].width = 30
    sheet_TC_RF_Rebuild_format.column_dimensions['J'].width = 30
    print(row_count_TC_RF_Rebuild_format)
    print(tasks)


    if locate_save == '':
        wb_TC_RF_Rebuild_format.save('TC_RF.xlsx')
    else:
        wb_TC_RF_Rebuild_format.save(locate_save + '/'+'TC_RF.xlsx')
    
    # wb6.close()
    # print(id)
    # print(percent)
    # print(tasks)
    # print(row_count)
    # print(row_count2)
    run_btn_text.set("RUN")
    tkinter.messagebox.showinfo("GREAT!", "Test case RFlash tool created successfully")

def DID_In_DPT(ws_DID_In_DPT, wb_DID_In_DPT, id, number1, number2, number3, number4, direct, status):
    p = 0
    if direct == '':
        wb_DID_In_DPT = load_workbook('BSWAvalue.xlsx')
        wb_General_Infomation = load_workbook('BSWAvalue.xlsx')
    else:
        wb_DID_In_DPT = load_workbook(str(direct))
        wb_General_Infomation = load_workbook(str(direct))

    ws_DID_In_DPT = wb_DID_In_DPT.active
    ws_General_Infomation = wb_General_Infomation.active
    
    ws_DID_In_DPT = wb_DID_In_DPT['DID_DPT']
    ws_General_Infomation = wb_General_Infomation['General Information']
    

    sheet_DID_In_DPT = wb_DID_In_DPT.worksheets[2]
    sheet_General_Infomation = wb_General_Infomation.worksheets[3]
    
    row_count_DID_In_DPT = sheet_DID_In_DPT.max_row
    row_count_General_Infomation = sheet_General_Infomation.max_row

    # print(row_count_DID_In_DPT)

    for row in range(2, 3):
        for col in range(2, 3):
            char = get_column_letter(col)
            latestSW = ws_General_Infomation[char + str(row)].value
            # print(latestSW)
    
    for row in range(8, 9):
        for col in range(2, 3):
            char = get_column_letter(col)
            row_DID_Check_Active_Session = ws_General_Infomation[char + str(row)].value
            row_DID_Check_Active_Session_lowercase = str(row_DID_Check_Active_Session).lower()
            # print(row_DID_Check_Active_Session_lowercase)

    i = 0
    o = 2
    j = 3

    k = 1
    number2 += 1

    c = 0

    i = 0
    count_string_number = 0
    hexvalue_DID_DPT = ""
    while k < row_count_DID_In_DPT:
        for row in range(o, j):
            for col in range(1, 2):
                char = get_column_letter(col)
                row_list_DID_DID_DPT = ws_DID_In_DPT[char + str(row)].value
                row_list_DID_DID_DPT_lowercase = str(row_list_DID_DID_DPT).lower()
                # print(row_list_DID_DID_DPT_lowercase)



        for row in range(o, j):
            for col in range(2, 3):
                char = get_column_letter(col)
                row_list_name_DID_DPT = ws_DID_In_DPT[char + str(row)].value
                # print(row_list_name_DID_DPT)
                if str(row_list_DID_DID_DPT) == "None":
                    messagebox.showerror("ERROR", "No DID, Please add DID")
                    break
                # print(ws[char + str(row)].value)
        for row in range(o, j):
            for col in range(3, 4):
                char = get_column_letter(col)
                row_list_length_byte_DID_DPT = ws_DID_In_DPT[char + str(row)].value
                # print(row_list_length_byte_DID_DPT)

        for row in range(o, j):
            for col in range(4, 5):
                char = get_column_letter(col)
                row_list_ASCII_values_DID_DPT = ws_DID_In_DPT[char + str(row)].value
                # print(row_list_ASCII_values_DID_DPT)
        
        for row in range(o, j):
            for col in range(5, 6):
                char = get_column_letter(col)
                row_list_hex_values_DID_DPT = ws_DID_In_DPT[char + str(row)].value
                # print(row_list_hex_values_DID_DPT)

        for row in range(o, j):
            for col in range(6, 7):
                char = get_column_letter(col)
                row_list_type_DID_DPT = ws_DID_In_DPT[char + str(row)].value
                # print(row_list_type_DID_DPT)
        
        for row in range(o, j):
            for col in range(7, 8):
                char = get_column_letter(col)
                row_list_programming_type_DID_DPT = ws_DID_In_DPT[char + str(row)].value
                # print(row_list_programming_type_DID_DPT)

        id += 1
        # check lenghth byte
        count_hexvalue_DID_DPT = 0
        hexvalue_DID_DPT = ""
        length_byte = 0
        if str(row_list_ASCII_values_DID_DPT) == 'None'  and str(row_list_hex_values_DID_DPT) == 'None':
            
            length_byte = int(row_list_length_byte_DID_DPT) * 2
            
            hexvalue_DID_DPT = str(".{" + str(length_byte) + "}")
            
            # print(hexvalue_DID_DPT)
            c = 1
        else:
            if str(row_list_hex_values_DID_DPT) == "None":
                # change ascii sang hex value
                for i in str(row_list_ASCII_values_DID_DPT):
                    hexvalue_DID_DPT += hex(ord(i))[2:]
                # print(hexvalue_DID_DPT)
                count_hexvalue_DID_DPT = len(hexvalue_DID_DPT)
                count_hexvalue_DID_DPT = int(count_hexvalue_DID_DPT) // 2
                # print(count_hexvalue_DID_DPT)
                # print(type(row_list_length_byte_DID_DPT))
                if str(count_hexvalue_DID_DPT) < row_list_length_byte_DID_DPT:
                    # print("Day la do dai byte",row_list_length_byte_DID_DPT)
                    length_byte = (int(row_list_length_byte_DID_DPT) - int(count_hexvalue_DID_DPT)) * 2
                    # print("byte bi thieu", length_byte)
                    hexvalue_DID_DPT = hexvalue_DID_DPT.lower()
                    hexvalue_DID_DPT = str(hexvalue_DID_DPT + ".{" + str(length_byte) + "}")
                    # print(hexvalue_DID_DPT)
                    c = 1
                else :
                    c = 0
            else:
                count_hexvalue_DID_DPT = len(row_list_hex_values_DID_DPT)
                count_hexvalue_DID_DPT = int(count_hexvalue_DID_DPT) // 2
                if str(count_hexvalue_DID_DPT) < row_list_length_byte_DID_DPT:
                    length_byte = (int(row_list_length_byte_DID_DPT) - int(count_hexvalue_DID_DPT)) * 2
                    hexvalue_DID_DPT = row_list_hex_values_DID_DPT.lower()
                    hexvalue_DID_DPT = str(hexvalue_DID_DPT + ".{" + str(length_byte) + "}")
                    c = 1
                else:
                    value_check_special = row_list_hex_values_DID_DPT.partition(".{")[1]
                    value_check_special2 = row_list_hex_values_DID_DPT.partition(".*")[1]
                    if value_check_special == '.{' or value_check_special2 == '.*':
                        c = 1
                    else:
                        c = 0
                    hexvalue_DID_DPT = row_list_hex_values_DID_DPT.lower()
                    
                    
            # print("hoan thanh",hexvalue_DID_DPT)
        if str(row_list_name_DID_DPT) == "None":
            row_list_name_DID_DPT = ""
            
        # if str(row_list_name_DID_DPT) != "Supplier Software number":

        if str(status) == 'Default':
            ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Access to Default Session' , 'Access to Default Session', '1) Access to Default Session\n2) Wait 3s\n3) Check active session should be Default', '1) -\n2) -\n3) -', '1) RequestResponse(1001,5001.*, Regexp)\n2) wait(3000)\n3) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*1' + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
            id += 1
            number2 += 1
            status = ''
        if str(status) == 'Extended':
            ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Access to Extended Session' , 'Access to Extended Session', '1) Access to Default Session\n2) Wait 3s\n3) Check active session should be Default\n4) Wait 3s\n5) Access to Extended Session\n6) Tester Present ON\n7) Wait 5s\n8) Check active session should be Extended', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -', '1) RequestResponse(1001,5001.*, Regexp)\n2) wait(3000)\n3) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*1' + ', Regexp)\n4) wait(3000)\n5) RequestResponse(1003,5003.*, Regexp)\n6) envvar(EnvTesterPresentOnOff(1;0))\n7) wait(5000)\n8) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*3' + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
            id += 1
            number2 += 1
            status = ''
        if str(status) == 'Programming':
            ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Access to Extended Session' , 'Access to Extended Session', '1) Access to Default Session\n2) Wait 3s\n3) Check active session should be Default\n4) Wait 3s\n5) Access to Extended Session\n6) Tester Present ON\n7) Wait 5s\n8) Check active session should be Extended\n9) Access to Programming Session\n10) Wait 3s\n11) Check active session should be Programming', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -', '1) RequestResponse(1001,5001.*, Regexp)\n2) wait(3000)\n3) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*1' + ', Regexp)\n4) wait(3000)\n5) RequestResponse(1003,5003.*, Regexp)\n6) envvar(EnvTesterPresentOnOff(1;0))\n7) wait(5000)\n8) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*3' + ', Regexp)\n9) RequestResponse(1002, 5002.*, Regexp)\n10) wait(3000)\n11) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*2' + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
            id += 1
            number2 += 1
            status = ''
            p = 1
            
        if p == 1:
            
            if str(row_list_programming_type_DID_DPT) == 'APL and FCN':
                if c == 1:
                    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' ' + str(row_list_DID_DID_DPT) + ' ' + str(row_list_name_DID_DPT), 'To check value of the DID ' + str(row_list_DID_DID_DPT), '1) Send service 0x22 to the camera for the DID ' +str(row_list_DID_DID_DPT) + ' using physical addressing\n2) Send service 0x22 to the camera for the DID ' + str(row_list_DID_DID_DPT) + ' using functional addressing', '1) -\n2) -', '1) RequestResponse(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+'62'+str(row_list_DID_DID_DPT_lowercase) + str(hexvalue_DID_DPT) + ', Regexp)\n2) FunctionalMessage(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+'62'+str(row_list_DID_DID_DPT_lowercase) + str(hexvalue_DID_DPT) + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
                
                if c == 0:
                    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' ' + str(row_list_DID_DID_DPT) + ' ' + str(row_list_name_DID_DPT), 'To check value of the DID ' + str(row_list_DID_DID_DPT), '1) Send service 0x22 to the camera for the DID ' +
                                str(row_list_DID_DID_DPT) + ' using physical addressing\n2) Send service 0x22 to the camera for the DID ' + str(row_list_DID_DID_DPT) + ' using functional addressing', '1) -\n2) -', '1) RequestResponse(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+'62'+str(row_list_DID_DID_DPT_lowercase) + str(hexvalue_DID_DPT) + ', Equal)\n2) FunctionalMessage(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+'62'+str(row_list_DID_DID_DPT_lowercase) + str(hexvalue_DID_DPT) + ', Equal)', 'Automated Testcase', 'implemented', latestSW, ''])

            if str(row_list_programming_type_DID_DPT) == 'APL':
                if c == 1:
                    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' ' + str(row_list_DID_DID_DPT) + ' ' + str(row_list_name_DID_DPT), 'To check value of the DID ' + str(row_list_DID_DID_DPT), '1) Send service 0x22 to the camera for the DID ' +
                                str(row_list_DID_DID_DPT) + ' using physical addressing\n2) Send service 0x22 to the camera for the DID ' + str(row_list_DID_DID_DPT) + ' using functional addressing', '1) -\n2) -', '1) RequestResponse(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+'62'+str(row_list_DID_DID_DPT_lowercase) + str(hexvalue_DID_DPT) + ', Regexp)\n2) FunctionalMessage(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+',None)', 'Automated Testcase', 'implemented', latestSW, ''])
                if c == 0:
                    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' ' + str(row_list_DID_DID_DPT) + ' ' + str(row_list_name_DID_DPT), 'To check value of the DID ' + str(row_list_DID_DID_DPT), '1) Send service 0x22 to the camera for the DID ' +
                                str(row_list_DID_DID_DPT) + ' using physical addressing\n2) Send service 0x22 to the camera for the DID ' + str(row_list_DID_DID_DPT) + ' using functional addressing', '1) -\n2) -', '1) RequestResponse(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+'62'+str(row_list_DID_DID_DPT_lowercase) + str(hexvalue_DID_DPT) + ', Equal)\n2) FunctionalMessage(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+',None)', 'Automated Testcase', 'implemented', latestSW, ''])

            if str(row_list_programming_type_DID_DPT) == 'FCN':
                if c == 1:
                    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' ' + str(row_list_DID_DID_DPT) + ' ' + str(row_list_name_DID_DPT), 'To check value of the DID ' + str(row_list_DID_DID_DPT), '1) Send service 0x22 to the camera for the DID ' +
                                str(row_list_DID_DID_DPT) + ' using physical addressing\n2) Send service 0x22 to the camera for the DID ' + str(row_list_DID_DID_DPT) + ' using functional addressing', '1) -\n2) -', '1) RequestResponse(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+',None)\n2) FunctionalMessage(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+'62'+str(row_list_DID_DID_DPT_lowercase) + str(hexvalue_DID_DPT) + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
                if c == 0:
                    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' ' + str(row_list_DID_DID_DPT) + ' ' + str(row_list_name_DID_DPT), 'To check value of the DID ' + str(row_list_DID_DID_DPT), '1) Send service 0x22 to the camera for the DID ' +
                                str(row_list_DID_DID_DPT) + ' using physical addressing\n2) Send service 0x22 to the camera for the DID ' + str(row_list_DID_DID_DPT) + ' using functional addressing', '1) -\n2) -', '1) RequestResponse(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+',None)\n2) FunctionalMessage(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+'62'+str(row_list_DID_DID_DPT_lowercase) + str(hexvalue_DID_DPT) + ', Equal)', 'Automated Testcase', 'implemented', latestSW, ''])

            if str(row_list_programming_type_DID_DPT) == 'NONE':
                if c == 1:
                    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' ' + str(row_list_DID_DID_DPT) + ' ' + str(row_list_name_DID_DPT), 'To check value of the DID ' + str(row_list_DID_DID_DPT), '1) Send service 0x22 to the camera for the DID ' +
                                str(row_list_DID_DID_DPT) + ' using physical addressing\n2) Send service 0x22 to the camera for the DID ' + str(row_list_DID_DID_DPT) + ' using functional addressing', '1) -\n2) -', '1) RequestResponse(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+',None)\n2) FunctionalMessage(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+',None)', 'Automated Testcase', 'implemented', latestSW, ''])
                if c == 0:
                    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' ' + str(row_list_DID_DID_DPT) + ' ' + str(row_list_name_DID_DPT), 'To check value of the DID ' + str(row_list_DID_DID_DPT), '1) Send service 0x22 to the camera for the DID ' +
                                str(row_list_DID_DID_DPT) + ' using physical addressing\n2) Send service 0x22 to the camera for the DID ' + str(row_list_DID_DID_DPT) + ' using functional addressing', '1) -\n2) -', '1) RequestResponse(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+',None)\n2) FunctionalMessage(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+',None)', 'Automated Testcase', 'implemented', latestSW, ''])

        else:
            if str(row_list_type_DID_DPT) == 'APL and FCN':
                if c == 1:
                    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' ' + str(row_list_DID_DID_DPT) + ' ' + str(row_list_name_DID_DPT), 'To check value of the DID ' + str(row_list_DID_DID_DPT), '1) Send service 0x22 to the camera for the DID ' +
                                str(row_list_DID_DID_DPT) + ' using physical addressing\n2) Send service 0x22 to the camera for the DID ' + str(row_list_DID_DID_DPT) + ' using functional addressing', '1) -\n2) -', '1) RequestResponse(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+'62'+str(row_list_DID_DID_DPT_lowercase) + str(hexvalue_DID_DPT) + ', Regexp)\n2) FunctionalMessage(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+'62'+str(row_list_DID_DID_DPT_lowercase) + str(hexvalue_DID_DPT) + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
                
                if c == 0:
                    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' ' + str(row_list_DID_DID_DPT) + ' ' + str(row_list_name_DID_DPT), 'To check value of the DID ' + str(row_list_DID_DID_DPT), '1) Send service 0x22 to the camera for the DID ' +
                                str(row_list_DID_DID_DPT) + ' using physical addressing\n2) Send service 0x22 to the camera for the DID ' + str(row_list_DID_DID_DPT) + ' using functional addressing', '1) -\n2) -', '1) RequestResponse(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+'62'+str(row_list_DID_DID_DPT_lowercase) + str(hexvalue_DID_DPT) + ', Equal)\n2) FunctionalMessage(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+'62'+str(row_list_DID_DID_DPT_lowercase) + str(hexvalue_DID_DPT) + ', Equal)', 'Automated Testcase', 'implemented', latestSW, ''])

            if str(row_list_type_DID_DPT) == 'APL':
                if c == 1:
                    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' ' + str(row_list_DID_DID_DPT) + ' ' + str(row_list_name_DID_DPT), 'To check value of the DID ' + str(row_list_DID_DID_DPT), '1) Send service 0x22 to the camera for the DID ' +
                                str(row_list_DID_DID_DPT) + ' using physical addressing\n2) Send service 0x22 to the camera for the DID ' + str(row_list_DID_DID_DPT) + ' using functional addressing', '1) -\n2) -', '1) RequestResponse(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+'62'+str(row_list_DID_DID_DPT_lowercase) + str(hexvalue_DID_DPT) + ', Regexp)\n2) FunctionalMessage(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+',None)', 'Automated Testcase', 'implemented', latestSW, ''])
                if c == 0:
                    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' ' + str(row_list_DID_DID_DPT) + ' ' + str(row_list_name_DID_DPT), 'To check value of the DID ' + str(row_list_DID_DID_DPT), '1) Send service 0x22 to the camera for the DID ' +
                                str(row_list_DID_DID_DPT) + ' using physical addressing\n2) Send service 0x22 to the camera for the DID ' + str(row_list_DID_DID_DPT) + ' using functional addressing', '1) -\n2) -', '1) RequestResponse(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+'62'+str(row_list_DID_DID_DPT_lowercase) + str(hexvalue_DID_DPT) + ', Equal)\n2) FunctionalMessage(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+',None)', 'Automated Testcase', 'implemented', latestSW, ''])

            if str(row_list_type_DID_DPT) == 'FCN':
                if c == 1:
                    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' ' + str(row_list_DID_DID_DPT) + ' ' + str(row_list_name_DID_DPT), 'To check value of the DID ' + str(row_list_DID_DID_DPT), '1) Send service 0x22 to the camera for the DID ' +
                                str(row_list_DID_DID_DPT) + ' using physical addressing\n2) Send service 0x22 to the camera for the DID ' + str(row_list_DID_DID_DPT) + ' using functional addressing', '1) -\n2) -', '1) RequestResponse(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+',None)\n2) FunctionalMessage(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+'62'+str(row_list_DID_DID_DPT_lowercase) + str(hexvalue_DID_DPT) + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
                if c == 0:
                    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' ' + str(row_list_DID_DID_DPT) + ' ' + str(row_list_name_DID_DPT), 'To check value of the DID ' + str(row_list_DID_DID_DPT), '1) Send service 0x22 to the camera for the DID ' +
                                str(row_list_DID_DID_DPT) + ' using physical addressing\n2) Send service 0x22 to the camera for the DID ' + str(row_list_DID_DID_DPT) + ' using functional addressing', '1) -\n2) -', '1) RequestResponse(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+',None)\n2) FunctionalMessage(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+'62'+str(row_list_DID_DID_DPT_lowercase) + str(hexvalue_DID_DPT) + ', Equal)', 'Automated Testcase', 'implemented', latestSW, ''])

            if str(row_list_type_DID_DPT) == 'NONE':
                if c == 1:
                    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' ' + str(row_list_DID_DID_DPT) + ' ' + str(row_list_name_DID_DPT), 'To check value of the DID ' + str(row_list_DID_DID_DPT), '1) Send service 0x22 to the camera for the DID ' +
                                str(row_list_DID_DID_DPT) + ' using physical addressing\n2) Send service 0x22 to the camera for the DID ' + str(row_list_DID_DID_DPT) + ' using functional addressing', '1) -\n2) -', '1) RequestResponse(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+',None)\n2) FunctionalMessage(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+',None)', 'Automated Testcase', 'implemented', latestSW, ''])
                if c == 0:
                    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' ' + str(row_list_DID_DID_DPT) + ' ' + str(row_list_name_DID_DPT), 'To check value of the DID ' + str(row_list_DID_DID_DPT), '1) Send service 0x22 to the camera for the DID ' +
                                str(row_list_DID_DID_DPT) + ' using physical addressing\n2) Send service 0x22 to the camera for the DID ' + str(row_list_DID_DID_DPT) + ' using functional addressing', '1) -\n2) -', '1) RequestResponse(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+',None)\n2) FunctionalMessage(' + '22' + str(row_list_DID_DID_DPT_lowercase) + ','+',None)', 'Automated Testcase', 'implemented', latestSW, ''])


        number2 += 1
        o += 1
        j += 1
        k += 1

    return id

def TC_FBL():
    run_btn_text.set("Loading...")

    direct = Input_path_text.get()
    locate_save = Output_path_text.get()


    if direct == '':
        wb_FBL_BaseSW = load_workbook('BSWAvalue.xlsx')
        wb_FBL_LatestSW = load_workbook('BSWAvalue.xlsx')
        wb_General_Infomation = load_workbook('BSWAvalue.xlsx')
        wb_DID_In_DPT = load_workbook('BSWAvalue.xlsx')
        wb_TC_FBL_Clear = load_workbook('TC_FBL.xlsx')
    else:
        wb_FBL_BaseSW = load_workbook(direct)
        wb_FBL_LatestSW = load_workbook(direct)
        wb_General_Infomation = load_workbook(direct)
        wb_DID_In_DPT = load_workbook(direct)
        wb_TC_FBL_Clear = load_workbook(locate_save + '/'+'TC_FBL.xlsx')
    # wb_FBL_LatestSW = load_workbook(direct)

    ws_FBL_BaseSW = wb_FBL_BaseSW.active
    ws_FBL_LatestSW = wb_FBL_LatestSW.active
    ws_General_Infomation = wb_General_Infomation.active
    ws_DID_In_DPT = wb_DID_In_DPT.active

    ws_FBL_BaseSW = wb_FBL_BaseSW['RFvalue_baseSW']
    ws_FBL_LatestSW = wb_FBL_LatestSW['RFvalue_latestSW']
    ws_General_Infomation = wb_General_Infomation['General Information']
    ws_DID_In_DPT = wb_DID_In_DPT['DID_DPT']
    

    sheet_FBL_BaseSW = wb_FBL_BaseSW.worksheets[0]
    sheet_FBL_LatestSW = wb_FBL_BaseSW.worksheets[1]
    sheet_DID_In_DPT = wb_DID_In_DPT.worksheets[2]

    row_count_FBL_BaseSW = sheet_FBL_BaseSW.max_row
    row_count_FBL_LatestSW = sheet_FBL_LatestSW.max_row
    row_count_DID_In_DPT = sheet_DID_In_DPT.max_row


    tasks = row_count_FBL_LatestSW + row_count_FBL_BaseSW + 151

    ws_TC_FBL_Clear = wb_TC_FBL_Clear.active
    ws_TC_FBL_Clear = wb_TC_FBL_Clear['TC_FBL']
    sheet_TC_FBL_Clear = wb_TC_FBL_Clear.worksheets[0]
    row_count_TC_FBL_Clear = sheet_TC_FBL_Clear.max_row

    # print(row_count_TC_FBL_Clear)
    print(tasks)

#  clear old TC_FBL
    if row_count_TC_FBL_Clear >= tasks:
        n = 0
        while n < row_count_TC_FBL_Clear:
            ws_TC_FBL.delete_rows(1)
            n += 1
            print("dang xoa")
        print("done")
        n = 0

# baseSW name
    for row in range(1, 2):
        for col in range(2, 3):
            char = get_column_letter(col)
            baseSW = ws_General_Infomation[char + str(row)].value

# latestSW name
    for row in range(2, 3):
        for col in range(2, 3):
            char = get_column_letter(col)
            latestSW = ws_General_Infomation[char + str(row)].value

# ticket
    for row in range(3, 4):
        for col in range(2, 3):
            char = get_column_letter(col)
            ticket_baseSW = ws_General_Infomation[char + str(row)].value
            ticket_latestSW = ws_General_Infomation[char + str(row)].value

# Programming_counter_DID
    for row in range(4, 5):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_counter_DID = ws_General_Infomation[char + str(row)].value

# Programming_Attempt_counter_DID
    for row in range(5, 6):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_Attempt_counter_DID = ws_General_Infomation[char + str(row)].value

# Programming_counter_step
    for row in range(6, 7):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_counter_step = ws_General_Infomation[char + str(row)].value

# Programming_Attempt_counter_step
    for row in range(7, 8):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_Attempt_counter_step = ws_General_Infomation[char + str(row)].value

    for row in range(8, 9):
        for col in range(2, 3):
            char = get_column_letter(col)
            row_DID_Check_Active_Session = ws_General_Infomation[char + str(row)].value
            row_DID_Check_Active_Session_lowercase = str(row_DID_Check_Active_Session).lower()

    
    # script begin
    id = 2
    number1 = 1
    number2 = 1
    number3 = 1
    number4 = 1

    border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(
        border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
    font_text_header = Font(name="Calibri", size=13,
                            color='00FFFFFF', bold=True)
    font_text = Font(name="Calibri", size=11, color='00000000', bold=False)
    alignment = Alignment(horizontal='center', vertical='center')
    ws_TC_FBL.append(['ID', 'XXX Component',  'Test Description', 'Test Steps',  'Test Response','Teststep keywords', 'ObjectType', 'TestStatus', 'Project', 'TestResult'])
    
    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(1, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text_header
        # cell_header.alignment = alignment

    ws_TC_FBL.append(['ID_'+str(id),  '1 FBL', '','', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(2, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        # cell_header.alignment = alignment
    # ------------------------------------------------------------------------------------------------------
    # BEGIN TEST CASE 1
    # TEST CASE 1 base SW to latestSW M3
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) +' Base SW to Latest SW M3', '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH BASE_SW VIA UART script
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) +' Flash Base SW via UART', '', '', '', '', 'Test group', '', '', ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])

    for row in range(id - 2, id + 1):
        for col in range(1, 11):
            cell_header = ws_TC_FBL.cell(row, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART flash ' + str(baseSW), 'Detail information is mentioned in the ticket: ' +str(ticket_baseSW), '1) Flash Base Software '+ str(baseSW) +' via UART successful', '1) -', "1) TesterConfirm('Do you flash base software "+str(baseSW)+" via UART ?')", 'Automated Testcase', 'implemented', str(baseSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification','', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_base_sw(id, number1, number2, number3, number4, tasks)

    # id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    # number4 += 1

    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    direct = Input_path_text.get()
    id = DID_baseSW(ws_TC_FBL, wb_FBL_BaseSW, id, number1, number2, number3, number4, direct)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])
    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    # print(id)

    PC_step = 0
    PAC_step = 0
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL','', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # ws_TC_FBL.title = "TC_FBL"
    # wb_TC_FBL.save('TC_FBL.xlsx')

    # Step2 FLASH LATEST_SW M3 VIA Xflash TOOLS

    ws_FBL_LatestSW = wb_FBL_LatestSW.active
    ws_FBL_LatestSW = wb_FBL_LatestSW['RFvalue_latestSW']

    sheet_FBL_LatestSW = wb_FBL_LatestSW.worksheets[1]
    row_count_FBL_LatestSW = sheet_FBL_LatestSW.max_row

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Latest SW M3 via X-Flash 1st',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via IFLASH successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via IFLASH ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    # id += 1
    # bar['value'] += 1
    # percent.set(str((id//tasks)*100)+"%")
    # direct = Input_path_text.get()
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    dummy  = ""
    id = DID_latestSW(ws_FBL_LatestSW, wb_FBL_LatestSW, id, number1, number2, number3, number4, direct, dummy)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = Programming_counter_step
    PAC_step = Programming_counter_step
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL','', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # ws_TC_FBL.title = "TC_FBL"
    # wb_TC_FBL.save('TC_FBL.xlsx')

    # END TEST CASE 1

    # ------------------------------------------------------------------------------------------------------
    # # BEGIN TEST CASE 2
    #  TEST CASE 2 base SW to latestSW M5
    number1 += 1
    number2 = 1
    number3 = 1
    number4 = 1
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) +
                ' Base SW to Latest SW M5', '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH BASE_SW VIA UART script
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Flash Base SW via UART',
                '', '', '', '', 'Test group', '', '', ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Flash SW', '',
                '', '', '', 'Test group', '', '', ''])

    for row in range(id - 2, id + 1):
        for col in range(1, 11):
            cell_header = ws_TC_FBL.cell(row, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART flash ' + str(baseSW), 'Detail information is mentioned in the ticket: ' +str(ticket_baseSW), '1) Flash Base Software '+ str(baseSW) +' via UART successful', '1) -', "1) TesterConfirm(''Do you flash base software "+str(baseSW)+" via UART ?')", 'Automated Testcase', 'implemented', str(baseSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_base_sw(id, number1, number2, number3, number4, tasks)

    # id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    # number4 += 1
    direct = Input_path_text.get()
    id = DID_baseSW(ws_TC_FBL, wb_FBL_BaseSW, id, number1, number2, number3, number4, direct)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = 0
    PAC_step = 0
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # # ws_TC_FBL.title = "TC_FBL"
    # # wb_TC_FBL.save('TC_FBL.xlsx')

    # # Step2 FLASH LATEST_SW M5 1st VIA Xflash TOOLS

    # # # Reflash latest SW M5 via xflash tool

    ws_FBL_LatestSW = wb_FBL_LatestSW.active
    ws_FBL_LatestSW = wb_FBL_LatestSW['RFvalue_latestSW']

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Latest SW M5 via X-Flash 1st','', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via IFLASH successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via IFLASH ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)


    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    direct = Input_path_text.get()
    dummy = ""
    id = DID_latestSW(ws_FBL_LatestSW, wb_FBL_LatestSW, id, number1, number2, number3, number4, direct, dummy)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = Programming_counter_step
    PAC_step = Programming_Attempt_counter_step
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text


    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*5,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # # Step2 FLASH LATEST_SW M3 2nd VIA Xflash TOOLS

    # # # Reflash latest SW M3 via xflash tool

    ws_FBL_LatestSW = wb_FBL_LatestSW.active
    ws_FBL_LatestSW = wb_FBL_LatestSW['RFvalue_latestSW']

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Latest SW M3 via X-Flash 2nd',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via IFLASH successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via IFLASH ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    direct = Input_path_text.get()
    dummy = ""
    id = DID_latestSW(ws_FBL_LatestSW, wb_FBL_LatestSW, id, number1, number2, number3, number4, direct, dummy)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = int(Programming_counter_step) + int(Programming_counter_step)
    PAC_step = int(Programming_Attempt_counter_step) + int(Programming_Attempt_counter_step)
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    #  ws_TC_FBL.title = "TC_FBL"
    #  wb_TC_FBL.save('TC_FBL.xlsx')

    # # END TEST CASE 2

    #  ------------------------------------------------------------------------------------------------------

    # # # BEGIN TEST CASE 3
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number1 += 1
    number2 = 0
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) +' SESSION TRANSITION', '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 = 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Default SS to Extended SS directly and vice versa' , 'Default SS to Extended SS directly and vice versa', '1) Access to Default Session\n2) Wait 3s\n3) Check active session should be Default\n4) Wait 3s\n5) Access to Extended Session\n6) Tester Present ON\n7) Wait 5s\n8) Check active session should be Extended\n9) Wait 3s\n10) Access to Default Session\n11) Wait 3s\n12) Check active session should be Default', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -', '1) RequestResponse(1001,5001.*, Regexp)\n2) wait(3000)\n3) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*1' + ', Regexp)\n4) wait(3000)\n5) RequestResponse(1003,5003.*, Regexp)\n6) envvar(EnvTesterPresentOnOff(1;0))\n7) wait(5000)\n8) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*3' + ', Regexp)\n9) wait(3000)\n10) RequestResponse(1001,5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*1' + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Default SS to Programming SS directly and vice versa' , 'Default SS to Programming SS directly and vice versa', '1) Access to Default Session\n2) Wait 3s\n3) Check active session should be Default\n4) Wait 3s\n5) Access to Programming Session (can not access directly and NRC 7E is responsed)\n6) Wait 3s\n7) Check active session should be Default\n8) Wait 3s\n9) Access to Extended Session before access to Programming session\n10) Tester Present ON\n11) Wait 5s\n12) Check active session should be Extended\n13) Access to Programming Session\n14) Wait 3s\n15) Check active session should be Programming\n16) Access to Default Session\n17) Wait 3s\n18) Check active session should be Default', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -\n13) -\n14) -\n15) -\n16) -\n17) -\n18) -', '1) RequestResponse(1001,5001.*, Regexp)\n2) wait(3000)\n3) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*1' + ', Regexp)\n4) wait(3000)\n5) RequestResponse(1002, 7f107e, Equal)\n6) wait(3000)\n7) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*1' + ', Regexp)\n8) wait(3000)\n9) RequestResponse(1003, 5003.*, Regexp)\n10) envvar(EnvTesterPresentOnOff(1;0))\n11) wait(5000)\n12) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*3' + ', Regexp)\n13) RequestResponse(1002, 5002.*, Regexp)\n14) wait(3000)\n15) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*2' + ', Regexp)\n16) RequestResponse(1001, 5001.*, Regexp)\n17) wait(3000)\n18) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*1' + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    
    # id += 1
    # bar['value'] += 1
    # percent.set(str((id//tasks)*100)+"%")
    # app.update_idletasks()
    # number2 +=1
    # ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Default SS to Programming SS directly and vice versa' , 'Default SS to Programming SS directly and vice versa', "1) Access to Default Session\n2) Wait 3s\n3) Check active session should be Default\n4) Wait 3s\n5) Access to Programming Session (can't access directly and NRC 7E is responded)\n6) Wait 3s\n7) Check active session should be Default\n8) Wait 3s\n9) Access to Extended before access to Programming Session\n10) Tester Present ON\n11 Wait 5s\n12) Check active session should be Extended\n13) Wait 3s\n14) Access to Progamming Session\n15) Wait 3s\n16) Check active session should be Progamming Session\n17) Access to Default Session\n18) Wait 3s\n19) Check active session should be Progamming Session", "1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -\n13) -\n14) -\n15) -\n16) -\n17) -\n18) -\n19) -", '1) RequestResponse(1001, 5001.*, Regexp)\n2) wait(3000)\n3) RequestResponse(22' + str(row_DID_Check_Active_Session_lowercase) + ', '+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*1' + ', Regexp)\n4) wait(3000)\n5) RequestResponse(1002, 7f107e, Equal)\n6) wait(3000)\n7) RequestResponse(22' + str(row_DID_Check_Active_Session_lowercase) + ', '+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*1' + ', Regexp)\n8) wait(3000)\n9) RequestResponse(1003, 5003.*, Regexp)\n10) envvar(EnvTesterPresentOnOff(1;0))\n11) wait(5000)\n12) RequestResponse(22' + str(row_DID_Check_Active_Session_lowercase) + ', '+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*3' + ', Regexp)\n13) RequestResponse(1002, 5002.*, Regexp)\n14) wait(3000)\n15) RequestResponse(22' + str(row_DID_Check_Active_Session_lowercase) + ', '+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*2' + ', Regexp)\n16) wait(3000)\n17) RequestResponse(1001, 5001.*, Regexp)\n18) wait(3000)\n19) RequestResponse(22' + str(row_DID_Check_Active_Session_lowercase) + ', '+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*1' + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])  

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Programming SS to Extended SS directly and vice versa' , 'Programming SS to Extended SS directly and vice versa', '1) Access to Default Session\n2) Wait 3s\n3) Check active session should be Default\n4) Wait 3s\n5) Access to Extended Session before access to Programming session\n6) Tester Present ON\n7) Wait 5s\n8) Check active session should be Extended\n9) Access to Programming Session\n10) Wait 3s\n11) Check active session should be Programming\n12) Access to Extended Session (can not access directly and NRC 7E is responsed)\n13) Wait 3s\n14) Check active session should be Programming\n15) Access to Default Session\n16) Wait 3s\n17) Check active session should be Default', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -\n13) -\n14) -\n15) -\n16) -\n17) -', '1) RequestResponse(1001,5001.*, Regexp)\n2) wait(3000)\n3) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*1' + ', Regexp)\n4) wait(3000)\n5) RequestResponse(1003, 5003.*, Regexp)\n6) envvar(EnvTesterPresentOnOff(1;0))\n7) wait(5000)\n8) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*3' + ', Regexp)\n9) RequestResponse(1002, 5002.*, Regexp)\n10) wait(3000)\n11) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*2' + ', Regexp)\n12) RequestResponse(1003,7f107e, Equal)\n13) wait(3000)\n14) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*2' + ', Regexp)\n15) RequestResponse(1001, 5001.*, Regexp)\n16) wait(3000)\n17) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*1' + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    

    # id += 1
    # bar['value'] += 1
    # percent.set(str((id//tasks)*100)+"%")
    # app.update_idletasks()
    # number2 +=1
    # ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Programming SS to Extended SS directly and vice versa' , 'Programming SS to Extended SS directly and vice versa', "1) Access to Default Session\n2) Wait 3s\n3) Check active session should be Default\n4) Wait 3s\n5) Access to Extended Session before access to Programming Session\n6) Tester Present ON\n7) Wait 5s\n8) Check active session should be Extended\n9) Access to Programming Session\n10) Wait 3s\n11) Check active session should be Programming Session\n12) Wait 3s\n13) Access to Extended Session (can't access directly and NRC 7E is responded)\n14) Wait 3s\n15) Check active session should be Programming Session\n16) Access to Default Session\n17) Wait 3s\n18) Check active session should be Default Session", '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -\n13) -\n14) -\n15) -\n16) -\n17) -\n18) -', '1) RequestResponse(1001,5001.*, Regexp)\n2) wait(3000)\n3) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*1' + ', Regexp)\n4) wait(3000)\n5) RequestResponse(1003,5003.*, Regexp)\n6) envvar(EnvTesterPresentOnOff(1;0))\n7) wait(5000)\n8) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*3' + ', Regexp)\n9) RequestResponse(1002,5002.*, Regexp)\n10) wait(3000)\n11) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*2' + ', Regexp)\n12) wait(3000)\n13) RequestResponse(1003,7f107e,Equal)\n14) wait(3000)\n15) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*2' + ', Regexp)\n16) RequestResponse(1001,5001.*, Regexp)\n17) wait(3000)\n18) RequestResponse(' + '22' + str(row_DID_Check_Active_Session_lowercase) + ','+'62'+str(row_DID_Check_Active_Session_lowercase) + '.*1' + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])



    # # # END TEST CASE 3

    # # # # BEGIN TEST CASE 4
        # DID in DPT
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number1 += 1
    number2 = 0
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + ' Read DID DPT in Default section','', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    status = 'Default'
    id = DID_In_DPT(ws_DID_In_DPT, wb_DID_In_DPT, id, number1, number2, number3, number4, direct, status)
    # print(id)

    # # #  END TEST CASE 4

    # -------------------------------------------------------------------------------------------------------

    # # # # BEGIN TEST CASE 5
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number1 += 1
    number2 = 0
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + ' Read DID DPT in Extended section',
                '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    status = 'Extended'
    id = DID_In_DPT(ws_DID_In_DPT, wb_DID_In_DPT, id, number1, number2, number3, number4, direct, status)
    # print(id)
    # # # # END TEST CASE 5

    # # # # # BEGIN TEST CASE 6
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number1 += 1
    number2 = 0
    ws_TC_FBL.append(['ID_'+str(id),  '1.' + str(number1) + ' Read DID DPT in Programming section',
                '', '', '', '', 'Test group', '', '', ''])
    

    for col in range(1, 11):
        cell_header = ws_TC_FBL.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    status = 'Programming'
    id = DID_In_DPT(ws_DID_In_DPT, wb_DID_In_DPT, id, number1, number2, number3, number4, direct, status)
    print(id)
    # # # # END TEST CASE 6

    bar['value'] += 36
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    # print(id)
    # print(str((id//tasks)*100)+"%")
    # print(bar['value'])

    locate_save = Output_path_text.get()
    ws_TC_FBL.title = "TC_FBL"
    # print(locate_save)
    if locate_save == '':
        # print('Ok')
        wb_TC_FBL.save('TC_FBL.xlsx')
    else:
        wb_TC_FBL.save(locate_save + '/'+'TC_FBL.xlsx')

    if locate_save == '':
        wb_TC_FBL_Rebuild_format = load_workbook('TC_FBL.xlsx')
    else:
        wb_TC_FBL_Rebuild_format = load_workbook(locate_save + '/'+'TC_FBL.xlsx')

    ws_TC_FBL_Rebuild_format = wb_TC_FBL_Rebuild_format['TC_FBL']
    sheet_TC_FBL_Rebuild_format = wb_TC_FBL_Rebuild_format.worksheets[0]
    row_count_TC_FBL_Rebuild_format = sheet_TC_FBL_Rebuild_format.max_row

    # sheet_TC_FBL_Rebuild_format = wb6.worksheets[0]
    sheet_TC_FBL_Rebuild_format.column_dimensions['A'].width = 30
    sheet_TC_FBL_Rebuild_format.column_dimensions['B'].width = 30
    sheet_TC_FBL_Rebuild_format.column_dimensions['C'].width = 30
    sheet_TC_FBL_Rebuild_format.column_dimensions['D'].width = 30
    sheet_TC_FBL_Rebuild_format.column_dimensions['E'].width = 30
    sheet_TC_FBL_Rebuild_format.column_dimensions['F'].width = 30
    sheet_TC_FBL_Rebuild_format.column_dimensions['G'].width = 30
    sheet_TC_FBL_Rebuild_format.column_dimensions['H'].width = 30
    sheet_TC_FBL_Rebuild_format.column_dimensions['I'].width = 30
    sheet_TC_FBL_Rebuild_format.column_dimensions['J'].width = 30
    # print(row_count_TC_FBL_Rebuild_format)
    # print(tasks)


    if locate_save == '':
        wb_TC_FBL_Rebuild_format.save('TC_FBL.xlsx')
    else:
        wb_TC_FBL_Rebuild_format.save(locate_save + '/'+'TC_FBL.xlsx')
    

    run_btn_text.set("RUN")
    tkinter.messagebox.showinfo("GREAT!", "Test case FBL tool created successfully")

def run_program():
    if selected_test_option.get() == 'ReFlash':
        # tkinter.messagebox.showinfo("GREAT!", "ReFlash")
        TC_RF()
    if selected_test_option.get() == 'FBL':
        # tkinter.messagebox.showinfo("GREAT!", "FBL")
        TC_FBL()
    if selected_test_option.get() == 'SRcheck':
        tkinter.messagebox.showinfo("GREAT!", "SRcheck will be update in future")

    # if selected_test_option.get() == 'EOL':
    #     tkinter.messagebox.showinfo("GREAT!", "EOL")

def start_program():
    try:
        os.remove('save_input.txt')
    except:
        print('')
    direct = Input_path_text.get()
    locate_save = Output_path_text.get()

    data_input_save = ',' + direct + ',' + locate_save

    with open('save_input.txt', 'w') as f:
        f.write(data_input_save)

    if direct == '':
        wb_RF_BaseSW = load_workbook('BSWAvalue.xlsx')
        wb_RF_LatestSW = load_workbook('BSWAvalue.xlsx')
        wb_General_Infomation = load_workbook('BSWAvalue.xlsx')
    else:
        wb_RF_BaseSW = load_workbook(direct)
        wb_RF_LatestSW = load_workbook(direct)
        wb_General_Infomation = load_workbook(direct)
        
    
    ws_RF_BaseSW = wb_RF_BaseSW.active
    ws_RF_BaseSW = wb_RF_BaseSW['RFvalue_baseSW']
    ws_RF_LatestSW = wb_RF_LatestSW.active
    ws_RF_LatestSW = wb_RF_LatestSW['RFvalue_latestSW']
    ws_General_Infomation = wb_General_Infomation.active
    ws_General_Infomation = wb_General_Infomation['General Information']
    sheet_RF_LatestSW = wb_RF_LatestSW.worksheets[1]
    sheet_RF_BaseSW = wb_RF_BaseSW.worksheets[0]
    row_count_RF_BaseSW = sheet_RF_BaseSW.max_row

    for row in range(2, 3):
        for col in range(1, 2):
            char = get_column_letter(col)
            row_list_DID_baseSW = ws_RF_BaseSW[char + str(row)].value
            row_list_DID_latestSW = ws_RF_LatestSW[char + str(row)].value
    
    for row in range(1, 2):
        for col in range(2, 3):
            char = get_column_letter(col)
            baseSW = ws_General_Infomation[char + str(row)].value
    for row in range(2, 3):
        for col in range(2, 3):
            char = get_column_letter(col)
            latestSW = ws_General_Infomation[char + str(row)].value

    for row in range(3, 4):
        for col in range(2, 3):
            char = get_column_letter(col)
            ticket_baseSW = ws_General_Infomation[char + str(row)].value
            ticket_latestSW = ws_General_Infomation[char + str(row)].value

    for row in range(4, 5):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_counter_DID = ws_General_Infomation[char + str(row)].value


    for row in range(5, 6):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_Attempt_counter_DID = ws_General_Infomation[char + str(row)].value

    for row in range(7, 8):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_counter_step = ws_General_Infomation[char + str(row)].value


    for row in range(8, 9):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_Attempt_counter_step = ws_General_Infomation[char + str(row)].value




    if str(row_list_DID_baseSW) == "None":
        messagebox.showerror(
            "ERROR", "No DID in row 1 at sheet RFvalue_baseSW, Please add DID")

    if str(row_list_DID_latestSW) == "None":
        messagebox.showerror(
            "ERROR", "No DID in row 1 at sheet RFvalue_latestSW, Please add DID")

    if str(baseSW) == "None":
        messagebox.showerror("ERROR", "BaseSW Name is invalid")

    if str(ticket_baseSW) == "None":
        messagebox.showerror("ERROR", "ticket_baseSW Name is invalid")

    if str(latestSW) == "None":
        messagebox.showerror("ERROR", "latestSW Name is invalid")

    if str(ticket_latestSW) == "None":
        messagebox.showerror("ERROR", "ticket_latestSW Name is invalid")

    if str(Programming_counter_DID) == "None":
        messagebox.showerror("ERROR", "Programming_counter_DID is invalid")

    if str(Programming_Attempt_counter_DID) == "None":
        messagebox.showerror("ERROR", "Programming_Attempt_counter_DID is invalid")

    if str(Programming_counter_step) == "None":
        messagebox.showerror("ERROR", "Programming_counter_step is invalid")

    if str(Programming_Attempt_counter_step) == "None":
        messagebox.showerror("ERROR", "Programming_Attempt_counter_step is invalid")


    if str(baseSW) != "None" and str(ticket_baseSW) != "None" and str(latestSW) != "None" and str(ticket_latestSW) != "None" and str(row_list_DID_baseSW) != "None" and str(row_list_DID_latestSW) != "None" and str(Programming_counter_DID) != "None" and str(Programming_Attempt_counter_DID) != "None" and str(Programming_counter_step) != "None" and str(Programming_Attempt_counter_step) != "None":
        print('run_program')
        # print(Programing_counter_text.get())
        if locate_save == '':
            try:
                # with open("TC_RF.xlsx", "r") as file:
                    # print("File was already have")
                fd = os.open("TC_RF.xlsx", os.O_RDWR)
                os.close(fd)
                
                run_program()
            except IOError:
                messagebox.showerror("ERROR","File TC_RF.xlsx has opened already. Please close the file in the same folder before push run button")
                # print("File has opened already. Please close the TC_RF.xlsx before push run button")
            
        else:
            try:
                # with open(locate_save + '/'+'TC_RF.xlsx', "r") as file:
                #     print("File has been remove")
                fd2 = os.open(locate_save + '/'+'TC_RF.xlsx', os.O_RDWR)
                os.close(fd2)
                
                run_program()
            except IOError:
                messagebox.showerror('ERROR','File TC_RF.xlsx has opened already. Please close the file at :'+ locate_save + '/'+'TC_RF.xlsx'+' before push run button')
                # print("File has opened already. Please close the TC_RF.xlsx before push run button")
        


# app front end
app = tk.Tk()

app.title('Basic SW Allide Tool')
app.geometry('700x420')


create_value_file()


def open_file():
    browse_input_path_text.set("loading...")
    file_path = askopenfile(parent=app, mode='rb', title="Choose location take file", filetype=[
                            ("excel file", ".xlsx")])
    print("Original string: " + str(file_path))

    result_str = ""
    final_str = ""
    for i in range(0, len(str(file_path))):
        if i >= 26:
            result_str = result_str + str(file_path)[i]
    reverse_str = result_str[::-1]
    for i in range(0, len(reverse_str)):
        if i >= 2:
            final_str = final_str + reverse_str[i]
    complete_str = final_str[::-1]
    print(type(complete_str))
    print(complete_str)
    if file_path:
        Input_path_text.set(str(complete_str))
        browse_input_path_text.set("Browse")
    return complete_str


def save_file():
    # print("is this working??")
    browse_output_path_text.set("loading...")
    file_path2 = filedialog.askdirectory()
    print(file_path2)
    if file_path2:
        Output_path_text.set(str(file_path2))
        browse_output_path_text.set("Browse")
    return str(file_path2)


file_path = ""

# Part Base SW
frameall = tk.Frame(app)
frame1 = tk.Frame(frameall)
frame2 = tk.Frame(frameall)
frame3 = tk.Frame(frameall)
    
    

Input_path_text = tk.StringVar()
Input_path_label = tk.Label(frame1, text='Input path', font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=0, column=0, sticky='w')
Input_path_entry = tk.Entry(frame1, textvariable=Input_path_text,
                            font='large_font', width=55).grid(row=1, column=0, sticky='w')

Output_path_text = tk.StringVar()
Output_path_label = tk.Label(frame1, text='Output path', font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=2, column=0, sticky='w')
Output_path_entry = tk.Entry(frame1, textvariable=Output_path_text,
                            font='large_font', width=55).grid(row=3, column=0, sticky='w')

# browse button open file
browse_input_path_text = tk.StringVar()
browse_btn_input_path = tk.Button(frame1, textvariable=browse_input_path_text, command=lambda: open_file(
), font="bold", width=7, height=1).grid(row=1, column=1, pady=5, padx=10)
browse_input_path_text.set("Browse")

# browse button save file
browse_output_path_text = tk.StringVar()
browse_btn_output_path = tk.Button(frame1, textvariable=browse_output_path_text, command=lambda: save_file(
), font="bold", width=7, height=1).grid(row=3, column=1, pady=5, padx=10)
browse_output_path_text.set("Browse")

# frame1.pack()

# Run program Buttons
run_btn_text = tk.StringVar()
run_btn = tk.Button(frame1, textvariable=run_btn_text, command=start_program,
                    font="bold", width=15).grid(row=4, column=0, columnspan=2, pady=20)
run_btn_text.set("RUN")

# frame1.pack()
# frame2.pack()

noneFill = tk.StringVar()
noneLabel = tk.Label(frame1, textvariable=noneFill, font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=5, column=0, columnspan=2)
# frame3.pack()

# frame1.pack()

# progress bar
bar = ttk.Progressbar(app, orient='horizontal', length=583, mode='determinate')

bar.place(relx=0.5, rely=0.6, anchor=CENTER)

# frame1.pack()

# frame4.pack()
frame1.pack()

percent = tk.StringVar()
percentLabel = tk.Label(frame2, textvariable=percent, font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=5, column=0, columnspan=3)

noneFill = tk.StringVar()
noneLabel = tk.Label(frame2, textvariable=noneFill, font=(
    'bold', 14), bg="#20bebe", fg="black" ,).grid(row=7, column=0, columnspan=2)

noneFill = tk.StringVar()
noneLabel = tk.Label(frame2, textvariable=noneFill, font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=8, column=0, columnspan=2)
# frame2.pack()

noneFill = tk.StringVar()
noneLabel = tk.Label(frame2, textvariable=noneFill, font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=9, column=0, columnspan=2)

# noneFill = tk.StringVar()
# noneLabel = tk.Label(frame2, textvariable=noneFill, font=(
#     'bold', 14), bg="#20bebe", fg="black").grid(row=10, column=0, columnspan=2)

# frame2.pack()

selected_test_option = tk.StringVar()
Test_option = [('ReFlash test', 'ReFlash'),
            ('FBL test', 'FBL'),
            ('SRcheck test', 'SRcheck')
            # ('EOL test', 'EOL')
            ]

myColor = '#20bebe'
s = ttk.Style()                     # Creating style element
s.configure('Wild.TRadiobutton',    # First argument is the name of style. Needs to end with: .TRadiobutton
        background=myColor,         # Setting background to our specified color above
        font=('Arial', 12,),
        foreground='black') 

for idx, (Test_list, text) in enumerate(Test_option):
    ttk.Radiobutton(frame2, text = Test_list, variable = selected_test_option,width = 13, value = text,style = 'Wild.TRadiobutton').grid(row=7, column = idx)
# ReFlash_option = ttk.Radiobutton(frame2,text='ReFlash test',value='ReFlash',variable=selected_test_option,width = 15,style = 'Wild.TRadiobutton').grid(row=7, column = 0, columnspan=10)
# FBL_option = ttk.Radiobutton(frame2,text='FBL test',value='FBL',variable=selected_test_option,width = 12,style = 'Wild.TRadiobutton').grid(row=7, column=2, columnspan=10)
# SRcheck_option= ttk.Radiobutton(frame2,text='SRcheck test',value='SRcheck',variable=selected_test_option,width = 15,style = 'Wild.TRadiobutton').grid(row=7, column=6, columnspan=10)
# EOL_option = ttk.Radiobutton(frame2,text='EOL test',value='EOL',variable=selected_test_option,width = 15,style = 'Wild.TRadiobutton').grid(row=7, column=10, columnspan=10)


frame2.pack()


frameall.place(relx=0.5, rely=0.5, anchor=CENTER)

# instruction

instruction = tk.Label(
    app, text="    Welcome to Basic Software Allied tool create by dev Huynh Minh Dang", font=("helvetica", 14))
instruction_version = tk.Label(app, text="R1.3.1", font=("helvetica", 14))
instruction_version.pack(side="right", anchor='s')
instruction.pack(side="bottom", fill='both', anchor=CENTER)

frameall.configure(background="#20bebe")
frame1.configure(background="#20bebe")
frame2.configure(background="#20bebe")

app.configure(background="#20bebe")

try:
    with open('save_input.txt', 'r') as f:
        data_input = f.read()
        # print(data_input)
    
        # input
        data_input = data_input.partition(",")[2] 
        direct_input = data_input
        direct_input = direct_input.split(',', 1)[0]
        Input_path_text.set(str(direct_input))
    
        # output
        data_input = data_input.partition(",")[2] 
        direct_output = data_input
        direct_output = direct_output.split(',', 1)[0]
        Output_path_text.set(direct_output)

except:
    print('chua co file data input')

# Start program
app.mainloop()
