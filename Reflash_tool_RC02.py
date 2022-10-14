import xlrd
import matplotlib
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import time

# base SW script
# chon sheet
# n = 0
wb = load_workbook('RFvalue2.xlsx')
ws = wb .active
ws = wb['RFvalue_baseSW']

sheet = wb.worksheets[0]
row_count = sheet.max_row
column_count = sheet.max_column


i = 0
o = 2
j = 3
id = 2
number = 1
count = 0
baseSW = "CA_CD569ICA_BL03_V4"
tiket_baseSW = "abc_323"


ws.append(['ID', 'XXX Component',  'Test Description', 'Test Steps',  'Test Response','Teststep keywords', 'ObjectType', 'TestStatus', 'Project', 'TestResult'])
ws.append(['ID_'+str(id),  '1 REFFLASH', '','', '', '', '', 'Test group', '', ''])
id += 1
# base SW script
ws.append(['ID_'+str(id),  '1.1 Base SW to Latest SW M3','', '', '', '', '', 'Test group', '', ''])
id += 1
ws.append(['ID_'+str(id),  '1.1.1 Flash Base SW via UART','', '', '', '', '', 'Test group', '', ''])
id += 1
ws.append(['ID_'+str(id),  '1.1.1.1 Flash SW', '','', '', '', '', 'Test group', '', ''])
id += 1
ws.append(['ID_'+str(id),  '1.1.1.1.1 UART flash'+baseSW, 'Detail information is mentioned in the ticket: '+tiket_baseSW,'Screen shot the successful flash procress', 'Screen shot the successful flash procress', 'Manual Testcase', 'Manual Testcase', 'implemented', baseSW, ''])
id += 1
ws.append(['ID_'+str(id),  '1.1.1.2 Variant and Software  Identification','', '', '', '', '', 'Test group', '', ''])
id += 1
ws.append(['ID_'+str(id),  '1.1.1.2.1 Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Select variant\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -','1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(2e014001, 6e0140, Equal)\n7) RequestResponse(22F1F0, 62014001, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])


print(row_count)
k = 1
# gap doi so dong  de xoa cac du lieu cu
dbrow = row_count + row_count
i = 0
hexvalue = ""
while k < row_count:
    for row in range(o, j):
        for col in range(1, 2):
            char = get_column_letter(col)
            row_list_DID = ws[char + str(row)].value
            # print(ws[char + str(row)].value)
    for row in range(o, j):
        for col in range(2, 3):
            char = get_column_letter(col)
            row_list_name = ws[char + str(row)].value
            # print(ws[char + str(row)].value)
    for row in range(o, j):
        for col in range(4, 5):
            char = get_column_letter(col)
            row_list_values = ws[char + str(row)].value

            # print(ws[char + str(row)].value)
    id += 1
    hexvalue = ""
    for i in str(row_list_values):
        hexvalue += hex(ord(i))[2:]
    print(hexvalue)
    ws.append(['ID_'+str(id),  '1.1.1.2.' + str(number) + ' ' + str(row_list_DID) + ' ' + str(row_list_name), 'To check value of the DID ' + str(row_list_DID), '1) Send service 0x22 to the camera for the DID ' +str(row_list_DID) + ' using physical addressing', '1) -', '1) RequestResponse(' + str(row_list_DID) + ','+str(hexvalue) + ', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    number += 1
    o += 1
    j += 1
    k += 1
# programing couter
id += 1
ws.append(['ID_'+str(id),  '1.1.1.3 Programming Counter and Programming Attempt Counter','', '', '', '', '', 'Test group', '', ''])
id += 1
ws.append(['ID_'+str(id),  '1.1.1.3.1 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing','1) -', '1) RequestResponse(220200, 620200.{3}0, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
id += 1
ws.append(['ID_'+str(id),  '1.1.1.3.1 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing','1) -', '1) RequestResponse(220201, 620201.{7}0, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
# RBEOL read
id += 1
ws.append(['ID_'+str(id),  '1.1.1.4 DID in RBEOL','', '', '', '', '', 'Test group', '', ''])
id += 1
ws.append(['ID_'+str(id),  '1.1.1.4.1 F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -','1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.{5}3 ,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
id += 1
ws.append(['ID_'+str(id),  '1.1.1.4.2 F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -','1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
id += 1
ws.append(['ID_'+str(id),  '1.1.1.4.2 4255', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -','1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
id += 1
ws.append(['ID_'+str(id),  '1.1.1.4.2 4259', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -','1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

ws.title = "TC_RF"
wb.save('TC_RF.xlsx')

# time.sleep(2)
# wb2 = load_workbook('TC_RF.xlsx')
# # ws2 = wb2 .active
# ws2 = wb2['TC_RF']
# dlrow = 1
# range = row_count + 1
# print(k)
# print(dbrow)

# if k >= row_count:
#     while k < dbrow:
#         ws2.delete_rows(1)
#         # dlrow += 1
#         k += 1
#         # print(dlrow)
#         print("dang xoa")
#     wb2.save('TC_RF.xlsx')
#     print("done")
#     k = 1

# ---------------------------------------------------------------------------------------------------------------

# latest SW script
wb3 = load_workbook('RFvalue2.xlsx')
ws3 = wb3 .active
ws3 = wb3['RFvalue_latestSW']

sheet = wb3.worksheets[1]
row_count2 = sheet.max_row
column_count = sheet.max_column

m = 0
a = 2
b = 3
# id = 2
number = 1
count = 0
latestSW = "CA_CD569ICA_BL03_RC05"
tiket_latestSW = "abc_779"

ws3.append(['ID', 'XXX Component',  'Test Description', 'Test Steps',  'Test Response','Teststep keywords', 'ObjectType', 'TestStatus', 'Project', 'TestResult'])
ws3.append(['ID_'+str(id),  '1 REFFLASH', '','', '', '', '', 'Test group', '', ''])
id += 1
# base SW script
ws3.append(['ID_'+str(id),  '1.1 Base SW to Latest SW M3','', '', '', '', '', 'Test group', '', ''])
id += 1
ws3.append(['ID_'+str(id),  '1.1.1 Flash Base SW via UART','', '', '', '', '', 'Test group', '', ''])
id += 1
ws3.append(['ID_'+str(id),  '1.1.1.1 Flash SW', '','', '', '', '', 'Test group', '', ''])
id += 1
ws3.append(['ID_'+str(id),  '1.1.1.1.1 UART flash'+latestSW, 'Detail information is mentioned in the ticket: '+tiket_latestSW,'Screen shot the successful flash procress', 'Screen shot the successful flash procress', 'Manual Testcase', 'Manual Testcase', 'implemented', latestSW, ''])
id += 1
ws3.append(['ID_'+str(id),  '1.1.1.2 Variant and Software  Identification','', '', '', '', '', 'Test group', '', ''])
id += 1
ws3.append(['ID_'+str(id),  '1.1.1.2.1 Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Select variant\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -','1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(2e014001, 6e0140, Equal)\n7) RequestResponse(22F1F0, 62014001, Equal)', 'Automated Testcase', 'implemented', latestSW, ''])


print(row_count2)
k = 1
# gap doi so dong  de xoa cac du lieu cu
dbrow2 = row_count2 + row_count2
m = 0
hexvalue = ""
row2 = 2
while k < row_count2:

    char = 'A'
    row_list_DID = ws3[char + str(row2)].value
    # print(ws3[char + str(row2)].value)

    char = 'B'
    row_list_name = ws3[char + str(row2)].value
    # print(ws3[char + str(row2)].value)

    char = 'D'
    row_list_values = ws3[char + str(row2)].value

            # print(ws3[char + str(row2)].value)
    id += 1
    hexvalue = ""
    for m in str(row_list_values):
        hexvalue += hex(ord(m))[2:]
    print(hexvalue)
    ws3.append(['ID_'+str(id),  '1.1.1.2.' + str(number) + ' ' + str(row_list_DID) + ' ' + str(row_list_name), 'To check value of the DID ' + str(row_list_DID), '1) Send service 0x22 to the camera for the DID ' +str(row_list_DID) + ' using physical addressing', '1) -', '1) RequestResponse(' + str(row_list_DID) + ','+str(hexvalue) + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    number += 1
    row2 += 1
    k += 1
# programing couter
id += 1
ws3.append(['ID_'+str(id),  '1.1.1.3 Programming Counter and Programming Attempt Counter','', '', '', '', '', 'Test group', '', ''])
id += 1
ws3.append(['ID_'+str(id),  '1.1.1.3.1 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing','1) -', '1) RequestResponse(220200, 620200.{3}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
id += 1
ws3.append(['ID_'+str(id),  '1.1.1.3.1 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing','1) -', '1) RequestResponse(220201, 620201.{7}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
# RBEOL read
id += 1
ws3.append(['ID_'+str(id),  '1.1.1.4 DID in RBEOL','', '', '', '', '', 'Test group', '', ''])
id += 1
ws3.append(['ID_'+str(id),  '1.1.1.4.1 F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -','1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.{5}3 ,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
id += 1
ws3.append(['ID_'+str(id),  '1.1.1.4.2 F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -','1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
id += 1
ws3.append(['ID_'+str(id),  '1.1.1.4.2 4255', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -','1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
id += 1
ws3.append(['ID_'+str(id),  '1.1.1.4.2 4259', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -','1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

ws3.title = "TC_RF"

wb3.save('TC_RF.xlsx')

# time.sleep(2)
wb4 = load_workbook('TC_RF.xlsx')
# ws2 = wb2 .active
ws4 = wb4['TC_RF']
dlrow = 1
range = row_count2 + 1
print(k)
print(dbrow2)

if k >= row_count2:
    while k < dbrow2:
        ws4.delete_rows(1)
        # dlrow += 1
        k += 1
        # print(dlrow)
        print("dang xoa")
    wb4.save('TC_RF.xlsx')
    print("done")
    k = 1
