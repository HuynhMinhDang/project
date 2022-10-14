from email import message
from msilib.schema import RadioButton
from turtle import textinput
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Color, numbers
from openpyxl.worksheet.datavalidation import DataValidation
import os
from struct import pack
import tkinter as tk
from PIL import Image, ImageTk
from tkinter.filedialog import askopenfile
from tkinter import CENTER, filedialog
from tkinter import HORIZONTAL, messagebox
from tkinter import ttk
from tkinter import IntVar
import tkinter.ttk
import time
import string
from tkinter.filedialog import askopenfile

from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re

import pandas as pd
from pandas import read_html
import html5lib

url = 'DIDCheck_DPT.html'



        # print(result_of_string)

# ----------------------------------------------------------------
# excel
try:
    # os.system('TASKKILL /F /IM EXCEL.exe')
    os.remove("Report_review.xlsx")
    wb_report_review = Workbook()
    ws_report_review = wb_report_review.active

    ws_report_review.title = "Report_review"
    wb_report_review.save("Report_review.xlsx")
    os.close("Report_review.xlsx")
except:
    try:
        wb_report_review = Workbook()
        ws_report_review = wb_report_review.active
        ws_report_review.title = "Report_review"
        wb_report_review.save("Report_review.xlsx")
        # os.system('TASKKILL /F /IM EXCEL.exe')
    except OSError:
        print('Failed creating the file')
    else:
        print('File created')


# ------------------------------------------------------------------------------------------------------------------------
def create_DPT_can_load_file():
    try:
        with open("DPT_can_load.xlsx", "r") as file:
            # Print the success message
            print("File is already haved")
        # fd = os.open("RFvalue.xlsx", os.O_RDWR)
        # os.close(fd)
    except OSError:

        wb_convert_DPT = load_workbook('DIA_Questionnaire_LAS_CA_C281D_20220626.xlsx')

        ws_convert_DPT = wb_convert_DPT.active

        ws_convert_DPT = wb_convert_DPT['7-1_DIDList']

        sheet_convert_DPT = wb_convert_DPT.worksheets[0]

        wb_convert_DPT.save("DPT_can_load.xlsx")
        wb_convert_DPT.close()

        # messagebox.showinfo("COMPLETE", "File RFvalue.xlsx has been created in the same folder tool successfully, Please fill all value")
        # print('tao thanh cong')

def excute_excell_file():
    wb_read_DPT = load_workbook('DPT_can_load.xlsx')
    wb_add_dpt_value_to_report_review = load_workbook('Report_review.xlsx')
    
    ws_read_DPT = wb_read_DPT.active
    ws_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review.active

    ws_read_DPT = wb_read_DPT['7-1_DIDList']
    ws_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review['Report_review']

    sheet_read_DPT = wb_read_DPT.worksheets[12]
    sheet_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review.worksheets[0]
    
    row_count_read_DPT = sheet_read_DPT.max_row    
    row_count_add_dpt_value_to_report_review = sheet_add_dpt_value_to_report_review.max_row
    
    k = 4
    o = 4
    j = 5
    count_length_byte = 0
    count_Value_DPT = 0
    length_byte = []
    Value_DPT = []
    load_second = 0
    # print(row_read_DPT)
    
    while k < row_count_read_DPT:
        for row in range(o, j):
            for col in range(1, 2):
                char = get_column_letter(col)
                row_list_DID = ws_read_DPT[char + str(row)].value
                # print(row_list_DID)
                
        for row in range(o, j):
            for col in range(3, 4):
                char = get_column_letter(col)
                row_list_length_byte = ws_read_DPT[char + str(row)].value
                # print(row_list_length_byte)

        for row in range(o, j):
            for col in range(21, 22):
                char = get_column_letter(col)
                row_list_value = ws_read_DPT[char + str(row)].value
                # print(row_list_value)
        
        if str(row_list_length_byte) != 'None':
            length_byte = length_byte + [row_list_length_byte]
            count_length_byte +=  1
        
        
        
        if str(row_list_value) != 'None':
            Value_DPT = Value_DPT + [row_list_value]
            count_Value_DPT +=  1
        
        o += 1
        j += 1
        k += 1
    
    

    

    for i in range(count_length_byte):
        length_byte = length_byte + [length_byte[i]]

    # count_length_byte = int(count_length_byte) * 2

    # for i in range(count_length_byte):
    #     print(length_byte[i], end = ' ')
    #     # print(Value_DPT[i])

    for i in range(count_Value_DPT):
        Value_DPT = Value_DPT + [Value_DPT[i]]

    # count_Value_DPT = int(count_Value_DPT) * 2

    # for i in range(count_Value_DPT):
    #     print(Value_DPT[i], end = ' ')
        # print(Value_DPT[i])
    
    # print(length_byte[1])





def review_report(ws_report_review, wb_report_review):

    wb_report_review = load_workbook('Report_review.xlsx')

    ws_report_review = wb_report_review.active
    ws_report_review = wb_report_review['Report_review']

    sheet_report_review = wb_report_review.worksheets[0]
    row_count_report_review = sheet_report_review.max_row


    wb_read_DPT = load_workbook('DPT_can_load.xlsx')
    wb_add_dpt_value_to_report_review = load_workbook('Report_review.xlsx')
    
    ws_read_DPT = wb_read_DPT.active
    ws_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review.active

    ws_read_DPT = wb_read_DPT['7-1_DIDList']
    ws_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review['Report_review']

    sheet_read_DPT = wb_read_DPT.worksheets[12]
    sheet_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review.worksheets[0]
    
    row_count_read_DPT = sheet_read_DPT.max_row    
    row_count_add_dpt_value_to_report_review = sheet_add_dpt_value_to_report_review.max_row
    
    k = 4
    o = 4
    j = 5
    count_length_byte = 0
    count_Value_DPT = 0
    length_byte = []
    Value_DPT = []
    load_second = 0
    # print(row_read_DPT)
    
    while k < row_count_read_DPT:
        for row in range(o, j):
            for col in range(1, 2):
                char = get_column_letter(col)
                row_list_DID = ws_read_DPT[char + str(row)].value
                # print(row_list_DID)
                
        for row in range(o, j):
            for col in range(3, 4):
                char = get_column_letter(col)
                row_list_length_byte = ws_read_DPT[char + str(row)].value
                # print(row_list_length_byte)

        for row in range(o, j):
            for col in range(21, 22):
                char = get_column_letter(col)
                row_list_value = ws_read_DPT[char + str(row)].value
                # print(row_list_value)
        
        if str(row_list_length_byte) != 'None':
            length_byte = length_byte + [row_list_length_byte]
            count_length_byte +=  1
        
        
        
        if str(row_list_value) != 'None':
            Value_DPT = Value_DPT + [row_list_value]
            count_Value_DPT +=  1
        
        o += 1
        j += 1
        k += 1
    
    

    

    for i in range(count_length_byte):
        length_byte = length_byte + [length_byte[i]]

    # count_length_byte = int(count_length_byte) * 2

    # for i in range(count_length_byte):
    #     print(length_byte[i], end = ' ')
    #     # print(Value_DPT[i])

    for i in range(count_Value_DPT):
        Value_DPT = Value_DPT + [Value_DPT[i]]

    # count_Value_DPT = int(count_Value_DPT) * 2

    # for i in range(count_Value_DPT):
    #     print(Value_DPT[i], end = ' ')
        # print(Value_DPT[i])


    h = 0
    id = 1
    with open(url, 'r') as f:

        ws_report_review.append(['ID_'+str(id) , 'DID','Length Byte RECEIVED','ASCII RECEIVED Value', 'HEX RECEIVED Value','HEX EXPECTED Value', 'ASCII EXPECTED Value','Length Byte EXPECTED', 'DPT Value', 'Length Byte DPT', 'DPT'])
            
        result_of_string = f.read()
        
        result_of_string = result_of_string.partition("EXPECTED: 62")[2]

        while len(result_of_string) > 0:
            id += 1
            length_of_EXPECTED = result_of_string.find('<')
            # print(length_of_EXPECTED)
            

            result_of_EXPECTED = result_of_string
            result_of_EXPECTED = result_of_EXPECTED.split('<', 1)[0]

            value_of_EXPECTED = result_of_EXPECTED[4:length_of_EXPECTED]
            DID_of_EXPECTED = result_of_EXPECTED[0:4]

            # print(value_of_EXPECTED)

            try:
                bytes_object_of_EXPECTED = bytes.fromhex(value_of_EXPECTED)
                
            except:
                missing_lenght_byte_of_EXPECTED = value_of_EXPECTED.partition(".{")[2]
                missing_lenght_byte_of_EXPECTED = missing_lenght_byte_of_EXPECTED[:-1]
                # print(missing_lenght_byte_of_EXPECTED)

                # missing_lenght_byte_of_EXPECTED = int(missing_lenght_byte_of_EXPECTED) // 2


                length_of_EXPECTED = int(length_of_EXPECTED) + int(missing_lenght_byte_of_EXPECTED) - 4
                # print(length_of_EXPECTED)

                value_of_EXPECTED = value_of_EXPECTED.split('.', 1)[0]
                # print(value_of_EXPECTED)

                bytes_object_of_EXPECTED = bytes.fromhex(value_of_EXPECTED)
                # print(bytes_object_of_EXPECTED)

                ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")
                # print(ascii_string_of_EXPECTED)
                # Convert to bytes object
            try:
                ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")

            # Convert to ASCII representation
            except:
                # ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")
                # ascii_string_of_EXPECTED = ILLEGAL_CHARACTERS_RE.sub(r'', ascii_string_of_EXPECTED)
                print('khong')
            if str(value_of_EXPECTED) != '':
                length_of_EXPECTED = (length_of_EXPECTED - 4) // 2
                # print('value khong trong '+ str(length_of_EXPECTED))
            else:
                length_of_EXPECTED = (length_of_EXPECTED // 2) - 2
                # print('value trong')

            # print('ASCII Value EXPECTED: ' + ascii_string_of_EXPECTED)
            # print('HEX Value EXPECTED: ' + value_of_EXPECTED)

            result_of_string = result_of_string.partition("RECEIVED: 62")[2] 
            # print(result_of_string)

            ### This is the part that measures the length before '<'
            length_of_RECEIVED = result_of_string.find('<')
            # print(length)
            

            result_of_RECEIVED = result_of_string
            result_of_RECEIVED = result_of_RECEIVED.split('<', 1)[0]

            value_of_RECEIVED = result_of_RECEIVED[4:length_of_RECEIVED]
            DID_of_RECEIVED = result_of_RECEIVED[0:4]

            bytes_object_of_RECEIVED = bytes.fromhex(value_of_RECEIVED)
                # Convert to bytes object

            length_of_RECEIVED = (length_of_RECEIVED - 4) // 2

            try:
                ascii_string_of_RECEIVED = bytes_object_of_RECEIVED.decode("ASCII")
                # print(ascii_string_of_RECEIVED)
                ascii_string_of_RECEIVED = ILLEGAL_CHARACTERS_RE.sub(r'', ascii_string_of_RECEIVED)
                # print(ascii_string_of_RECEIVED)
                
            # Convert to ASCII representation
            except:
                ascii_string_of_RECEIVED = ILLEGAL_CHARACTERS_RE.sub(r'', ascii_string_of_RECEIVED)
                # print(ascii_string_of_RECEIVED)


            result_of_string = result_of_string.partition("EXPECTED: 62")[2]

            if str(ascii_string_of_EXPECTED) == str(ascii_string_of_RECEIVED) and str(value_of_EXPECTED) == str(value_of_RECEIVED) or str(length_of_EXPECTED) == str(length_of_RECEIVED):
                result = 'PASS'
            else:
                result = 'FAIL'
            
            try:
                ws_report_review.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(length_of_RECEIVED),str(ascii_string_of_RECEIVED), str(value_of_RECEIVED),str(value_of_EXPECTED), str(ascii_string_of_EXPECTED),str(length_of_EXPECTED),str(Value_DPT[h]) , str(length_byte[h]), 'DPT',str(result)])
            except:
                ws_report_review.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(length_of_RECEIVED),'', str(value_of_RECEIVED),str(value_of_EXPECTED), '',str(length_of_EXPECTED),str(Value_DPT[h]) , str(length_byte[h]), 'DPT',str(result)])
            
            h += 1
            
    
    with open('Report_review.txt', 'w') as f:
        f.write(result_of_string)

    wb_report_review.save("Report_review.xlsx")
    wb_report_review.close()
    

create_DPT_can_load_file()
review_report(ws_report_review, wb_report_review)
# excute_excell_file()

