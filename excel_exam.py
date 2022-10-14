from typing_extensions import final
import xlrd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

# n=0
# wb = openpyxl.load_workbook('RFvalue2.xlsx')
# sheets = wb.sheetnames
# sheet = wb.worksheets[0]
# ws = wb[sheets[n]]
# row_count = sheet.max_row
# print(row_count)
# wb = Workbook()
# ws = wb .active
# ws.title = "RFvalue_baseSW"
# print(wb.sheetnames)
# ws['A2'].value = "Test"
i = 0
# k = ''
m = 1
o = 2
j = 3
id = 2
# row = 2
hexvalue = ""
# char = 'D'
# while i < row_count:
#     for row  in range(o, j):
#         for col in range(4,5):
#             char = get_column_letter(col)
#             print(ws[char + str(row)].value)
#     for k in ws[char + str(row)].value:
#         hexvalue += hex(ord(k))[2:]
#     print(hexvalue)
#     o += 1
#     j += 1
#     i += 1
#     row += 1
    
# print(ws[char + str(row)].value)
# for k in str(ws[char + str(row)].value):
#     hexvalue += hex(ord(k))[2:]
# print(hexvalue)
# for row  in range(o, j):
#     for col in range(4,5):
#         char = get_column_letter(col)
#         print(char)
#         print(row)
#         print(ws[char + str(row)].value)
#         for k in ws[char + str(row)].value:
#             hexvalue += hex(ord(k))[2:]
#     print(hexvalue)

# while m < row_count:
#     for row in range(o, j):
#         for col in range(1, 2):
#             char = get_column_letter(col)
#             row_list_DID = ws[char + str(row)].value
#             # print(ws[char + str(row)].value)
#     for row in range(o, j):
#         for col in range(2, 3):
#             char = get_column_letter(col)
#             row_list_name = ws[char + str(row)].value
#             # print(ws[char + str(row)].value)
#     for row in range(o, j):
#         for col in range(4, 5):
#             char = get_column_letter(col)
#             row_list_values = ws[char + str(row)].value
#             # print(ws[char + str(row)].value)
#     id += 1
#     for i in row_list_values:
#         hexvalue += hex(ord(i))[2:]
#     print(hexvalue)
#     # ws.append(['ID_'+str(id),  '1.1.1.2.' + str(number) + ' ' + str(row_list_DID) + ' ' + str(row_list_name), 'To check value of the DID ' + str(row_list_DID), '1) Send service 0x22 to the camera for the DID ' +str(row_list_DID) + ' using physical addressing', '1) -', '1) RequestResponse(' + str(row_list_DID) + ','+str(hexvalue) + ', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
#     # number += 1
#     o += 1
#     j += 1
#     m += 1
        

# for row  in range(o, j):
#     for col in range(4,5):
#         char = get_column_letter(col)
#         print(ws[char + str(row)].value)
# for i in ws[char + str(row)].value:
#     hexvalue += hex(ord(i))[2:]
# print(hexvalue)
# o+=1
# j+=1
# for row in range(o, j):
#     for col in range(4, 5):
        # char = get_column_letter(col)
        # print(ws[char + str(row)].value)
# wb = load_workbook('RFvalue2.xlsx')
# ws = wb .active
# ws = wb['RFvalue_latestSW']
# sheet = wb.worksheets[1]
# row_count2 = sheet.max_row
# row_count = sheet.max_row


# char = 'D'
# row2 = 2
# print(ws[char + str(row2)].value)
# hexvalue =""
# for i in ws[char + str(row2)].value:
#     hexvalue += hex(ord(i))[2:]
# print(hexvalue)
# ws.title = "Data"

# # ws.append(['DiD',  'Value', 'Test', '1)\n2)\n3)'])

# wb.save('Test.xlsx')

# wb.save('TC_RF.xlsx')
# time.sleep(2)
# wb2 = load_workbook('TC_RF.xlsx')
# # ws2 = wb2 .active
# ws2 = wb2['TC_RF']
# dlrow = 1
# range = row_count + 1
# # print(k)
# # print(dbrow)
# # if k >= row_count:
# z = 36
# while z < 72:
#     ws2.delete_rows(1)
#     dlrow += 1
#     z += 1
#     print(dlrow)
#     print("dang xoa")
# # ws2.move_range("A"+str(range)+":"+"J"+str(dbrow), rows=0, cols=0)
# # print(k)
# ws2.title = "TC_RF"


# wb2.save('TC_RF.xlsx')
# print("done")

# wb = load_workbook('C:/Users/MDBASKETBALL/OneDrive/Máy tính/code_cong_ty/HuynhMinhDang08022001_dang/RFvalue2.xlsx')
# os.system('TASKKILL /F /IM C:/Users/MDBASKETBALL/OneDrive/Máy tính/code_cong_ty/HuynhMinhDang08022001_dang/RFvalue2.xlsx')
# os.close('C:/Users/MDBASKETBALL/OneDrive/Máy tính/code_cong_ty/HuynhMinhDang08022001_dang/RFvalue2.xlsx')
# wb = load_workbook(
#     r'C:\Users\MDBASKETBALL\OneDrive\Máy tính\code_cong_ty\HuynhMinhDang08022001_dang\RFvalue2.xlsx')

# ws = wb.active
# ws = wb['RFvalue_baseSW']
# sheet = wb.worksheets[0]
# row_count = sheet.max_row
# for row in range(1, 2):
#     for col in range(7, 8):
#         char = get_column_letter(col)
#         baseSW = ws[char + str(row)].value
# print(baseSW)

# input_str = "DivasDwivedi"

final_str2 = ""
final_str3 = ""

def open_file():
    input_str = "<_io.BufferedReader name='C:/Users/MDBASKETBALL/OneDrive/Máy tính/code_cong_ty/HuynhMinhDang08022001_dang/RFvalue2.xlsx'>"

    # Printing original string
    print("Original string: " + input_str)

    result_str = ""
    final_str = ""
    for i in range(0, len(input_str)):
        if i >= 26:
            result_str = result_str + input_str[i]
    new_str = result_str[::-1]
    for i in range(0, len(new_str)):
        if i >= 2:
            final_str = final_str + new_str[i]
    final_str2 = final_str[::-1]

    # Printing string after removal
    print("String after removal of i'th character : " + final_str2)
    return final_str2


locate_str = "C:/Users/MDBASKETBALL/OneDrive/Máy tính/code_cong_ty"

final_str3 = open_file()
wb = load_workbook(final_str3)
ws = wb.active
ws = wb['RFvalue_baseSW']
sheet = wb.worksheets[0]
row_count = sheet.max_row
for row in range(1, 2):
    for col in range(7, 8):
        char = get_column_letter(col)
        baseSW = ws[char + str(row)].value
print(baseSW)
# wb.save(locate_str +'/'+ "TC_RF.xlsx")
