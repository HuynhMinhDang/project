from email import message
from turtle import textinput
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Color, numbers
from openpyxl.worksheet.datavalidation import DataValidation
import os
from struct import pack
import tkinter as tk
from PIL import Image, ImageTk
from tkinter.filedialog import askopenfile
from tkinter import CENTER, filedialog
from tkinter import HORIZONTAL, messagebox
from tkinter import ttk
import tkinter.ttk
import time
import string
from tkinter.filedialog import askopenfile


try:
    # os.system('TASKKILL /F /IM EXCEL.exe')
    os.remove("TC_RF.xlsx")
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "TC_RF"
    wb2.save("TC_RF.xlsx")
    os.close("TC_RF.xlsx")
except:
    try:
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "TC_RF"
        wb2.save("TC_RF.xlsx")
        # os.system('TASKKILL /F /IM EXCEL.exe')
    except OSError:
        print('Failed creating the file')
    else:
        print('File created')


def create_value_file():
    try:
        with open("RFvalue.xlsx", "r") as file:
            # Print the success message
            print("File is already haved")
        # fd = os.open("RFvalue.xlsx", os.O_RDWR)
        # os.close(fd)
    except OSError:
        z = 0
        x = 0
        wb5 = Workbook()
        ws5 = wb5.active
        sheet = wb5.worksheets[0]
        noneFill = PatternFill(start_color='00FFFFFF',
                                end_color='00FFFFFF',
                                fill_type='solid'
                                )
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
        font_text = Font(name="Calibri", size=14, color='00FFFFFF', bold=True)
        font_text2 = Font(name="Calibri", size=11, color='000000', bold=False, italic = True)
        alignment = Alignment(horizontal='center', vertical='center')

        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['F'].width = 40
        sheet.column_dimensions['G'].width = 40
        sheet.column_dimensions['H'].width = 35
        sheet.column_dimensions['A'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['B'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['C'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['D'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['E'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['F'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['G'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['A'].alignment = alignment
        sheet.column_dimensions['B'].alignment = alignment
        sheet.column_dimensions['C'].alignment = alignment
        sheet.column_dimensions['D'].alignment = alignment
        sheet.column_dimensions['E'].alignment = alignment
        sheet.column_dimensions['F'].alignment = alignment
        sheet.column_dimensions['G'].alignment = alignment


# tao ra sheet base sw

        ws5.title = "RFvalue_baseSW"
        # ws5.append(['DID', 'Description', 'Length (Byte)','ASCII Value', 'HEX_Value', 'Type'])
        ws5.append(['DID', 'Description','ASCII Value', 'HEX_Value', 'Type'])

        ws5['F1'] = 'BaseSW Name'
        cell_header = ws5.cell(1, 6)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'Ticket BaseSW'])
        ws5['F2'] = 'Variant Name baseSW'
        cell_header = ws5.cell(2, 6)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'Variant BaseSW'])
        ws5['F3'] = 'Variant BaseSW'
        cell_header = ws5.cell(3, 6)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'DID check variant BaseSW'])
        ws5['F4'] = 'DID check variant BaseSW'
        cell_header = ws5.cell(4, 6)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['F5'] = 'Security Unlock Level'
        cell_header = ws5.cell(5, 6)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['H5'] = '0: No level; 1: level1; 2: level2; 3: level3'
        cell_header = ws5.cell(5, 8)
        cell_header.border = border
        cell_header.font = font_text2



        for col in range(1, 7):
            cell_header = ws5.cell(1, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='000066CC', end_color='000066CC', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
            cell_header.alignment = alignment

        data_validation_data = '"APPL, FBL, RBEOL"'
        for row in range(2, 30):
            data_validation = DataValidation(type='list', formula1 = data_validation_data)
            ws5.add_data_validation(data_validation)
            data_validation.add(ws5['E'+str(row)])



        # cell_header = ws5.cell(1, 5)
        # cell_header.fill = noneFill
        # cell_header.border = border

# ket thuc sheet baseSW

# tao ra sheet latest SW

        wb5.create_sheet("RFvalue_latestSW")
        ws5 = wb5['RFvalue_latestSW']
        sheet2 = wb5.worksheets[1]
        sheet2.column_dimensions['B'].width = 50
        sheet2.column_dimensions['C'].width = 20
        sheet2.column_dimensions['D'].width = 30
        sheet2.column_dimensions['E'].width = 20
        sheet2.column_dimensions['F'].width = 50
        sheet2.column_dimensions['G'].width = 40
        sheet2.column_dimensions['H'].width = 40
        sheet2.column_dimensions['I'].width = 35
        sheet2.column_dimensions['A'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['B'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['C'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['D'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['E'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['F'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['G'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['H'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['A'].alignment = alignment
        sheet2.column_dimensions['B'].alignment = alignment
        sheet2.column_dimensions['C'].alignment = alignment
        sheet2.column_dimensions['D'].alignment = alignment
        sheet2.column_dimensions['E'].alignment = alignment
        sheet2.column_dimensions['F'].alignment = alignment
        sheet2.column_dimensions['G'].alignment = alignment
        sheet2.column_dimensions['H'].alignment = alignment
        # column_count2 = sheet2.max_column
        # ws5.append(['DID', 'Description', 'Length (Byte)','ASCII Value', 'HEX_Value', 'Type'])
        ws5.append(['DID', 'Description','ASCII Value', 'HEX_Value', 'Type','DummySW value(hex)'])

        ws5['G1'] = 'LatestSW Name'
        cell_header = ws5.cell(1, 7)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'Ticket LatestSW'])
        ws5['G2'] = 'Variant Name LatestSW'
        cell_header = ws5.cell(2, 7)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'Variant LatestSW'])
        ws5['G3'] = 'Variant LatestSW'
        cell_header = ws5.cell(3, 7)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'DID check variant LatestSW'])
        ws5['G4'] = 'DID check variant LatestSW'
        cell_header = ws5.cell(4, 7)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['G5'] = 'Security Unlock Level'
        cell_header = ws5.cell(5, 7)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['I5'] = '0: No level; 1: level1; 2: level2; 3: level3'
        cell_header = ws5.cell(5, 9)
        cell_header.border = border
        cell_header.font = font_text2



        for col in range(1, 8):
            cell_header = ws5.cell(1, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='000066CC', end_color='000066CC', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
            cell_header.alignment = alignment
            
        data_validation_data = '"APPL, FBL, RBEOL"'
        
        for row in range(2, 30):
            data_validation = DataValidation(type='list', formula1 = data_validation_data)
            ws5.add_data_validation(data_validation)
            data_validation.add(ws5['E'+str(row)])
        


        wb5.create_sheet("General Information")
        ws5 = wb5['General Information']
        sheet4 = wb5.worksheets[2]
        sheet4.column_dimensions['A'].width = 50
        sheet4.column_dimensions['B'].width = 50
        sheet4.column_dimensions['C'].width = 50
        
        sheet4.column_dimensions['A'].number_format = numbers.FORMAT_TEXT
        sheet4.column_dimensions['B'].number_format = numbers.FORMAT_TEXT
        sheet4.column_dimensions['C'].number_format = numbers.FORMAT_TEXT
        
        sheet4.column_dimensions['A'].alignment = alignment
        sheet4.column_dimensions['B'].alignment = alignment
        sheet4.column_dimensions['C'].alignment = alignment

        ws5['A1'] = 'Ticket ID'
        cell_header = ws5.cell(1, 1)
        cell_header.fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'Ticket LatestSW'])
        ws5['A2'] = 'Programming Counter DID'
        cell_header = ws5.cell(2, 1)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'Variant LatestSW'])
        ws5['A3'] = 'Programming Attempt Counter DID'
        cell_header = ws5.cell(3, 1)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'DID check variant LatestSW'])
        ws5['A4'] = 'Step counter of PC DID'
        cell_header = ws5.cell(4, 1)
        cell_header.fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['A5'] = 'Step counter of PAC DID'
        cell_header = ws5.cell(5, 1)
        cell_header.fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment

        ws5['C4'] = '*PC-Programing_counter'
        cell_header = ws5.cell(4, 3)
        cell_header.font = font_text2
    
        ws5['C5'] = '*PAC-Programing_Attempt_Counter'
        cell_header = ws5.cell(5, 3)
        cell_header.font = font_text2

        
        for col in range(1, 2):
            cell_header = ws5.cell(1, col)
            # used hex code for red color
            cell_header.fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
            cell_header.alignment = alignment


# tao ra sheet general information

# ket thuc sheet general information

        # cell_header = ws5.cell(1, 5)
        # cell_header.fill = noneFill
        # cell_header.border = border

        wb5.save("RFvalue.xlsx")
        wb5.close()
        # os.system('TASKKILL /F /IM EXCEL.exe')
        messagebox.showinfo("COMPLETE", "File RFvalue.xlsx has been created in the same folder tool successfully, Please fill all value")
        print('tao thanh cong')


        
        # Mo mot file
        # path = "/py/"
        # dirs = os.listdir(path)

        # # Lenh de in tat ca file va thu muc
        # for file in dirs:
        #     print(file)


check_DID = 0


def DID_baseSW(ws2, wb, id, number1, number2, number3, number4, direct):
    if direct == '':
        wb = load_workbook('RFvalue.xlsx')
    else:
        wb = load_workbook(str(direct))
    ws = wb.active
    ws = wb['RFvalue_baseSW']
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    for row in range(1, 2):
        for col in range(7, 8):
            char = get_column_letter(col)
            baseSW = ws[char + str(row)].value

    # for row in range(2, 3):
    #     for col in range(6, 7):
    #         char = get_column_letter(col)
    #         DID_check_active_baseSW = ws[char + str(row)].value
    #         DID_check_active_baseSW_lowercase = str(DID_check_active_baseSW).lower()
    #         print(DID_check_active_baseSW_lowercase)

    i = 0
    o = 2
    j = 3

    k = 1
    number4 += 1

    i = 0
    count_string_number = 0
    hexvalue_baseSW = ""
    while k < row_count:
        for row in range(o, j):
            for col in range(1, 2):
                char = get_column_letter(col)
                row_list_DID_baseSW = ws[char + str(row)].value
                row_list_DID_baseSW_lowercase = str(row_list_DID_baseSW).lower()

        for row in range(o, j):
            for col in range(2, 3):
                char = get_column_letter(col)
                row_list_name_baseSW = ws[char + str(row)].value
                if str(row_list_DID_baseSW) == "None":
                    messagebox.showerror("ERROR", "No DID or DID invalid, Please add DID")
                    break
                # print(ws[char + str(row)].value)
        # for row in range(o, j):
        #     for col in range(3, 4):
        #         char = get_column_letter(col)
        #         row_list_length_byte_baseSW = ws[char + str(row)].value

        for row in range(o, j):
            for col in range(3, 4):
                char = get_column_letter(col)
                row_list_ASCII_values_baseSW = ws[char + str(row)].value
        
        for row in range(o, j):
            for col in range(4, 5):
                char = get_column_letter(col)
                row_list_hex_values_baseSW = ws[char + str(row)].value

        for row in range(o, j):
            for col in range(5, 6):
                char = get_column_letter(col)
                row_list_type_baseSW = ws[char + str(row)].value
                print(row_list_type_baseSW)

        id += 1
        # check lenghth byte
        count_hexvalue_baseSW = 0
        hexvalue_baseSW = ""
        length_byte = 0
        if str(row_list_ASCII_values_baseSW) == 'None'  and str(row_list_hex_values_baseSW) == 'None':
            # hexvalue_baseSW = str(".{" + str(row_list_length_byte_baseSW) + "}")
            # length_byte = int(row_list_length_byte_baseSW) * 2
            # print(length_byte)
            # hexvalue_baseSW = str(".{" + str(length_byte) + "}")
            hexvalue_baseSW = str(".*")
            # print('dung')
            # print(hexvalue_baseSW)
        else:
            if str(row_list_hex_values_baseSW) == "None":
                # change ascii sang hex value
                for i in str(row_list_ASCII_values_baseSW):
                    hexvalue_baseSW += hex(ord(i))[2:]

            # danh cho co length byter thi su dung
                # print(hexvalue_baseSW)
                # count_hexvalue_baseSW = len(hexvalue_baseSW)
                # count_hexvalue_baseSW = int(count_hexvalue_baseSW) // 2
                # # print(count_hexvalue_baseSW)
                # # print(type(row_list_length_byte_baseSW))
                # if str(count_hexvalue_baseSW) < row_list_length_byte_baseSW:
                #     # print("Day la do dai byte",row_list_length_byte_baseSW)
                #     length_byte = (int(row_list_length_byte_baseSW) -
                #                 int(count_hexvalue_baseSW)) * 2
                #     # print("byte bi thieu", length_byte)
                #     hexvalue_baseSW = hexvalue_baseSW.lower()
                #     hexvalue_baseSW = str(hexvalue_baseSW + ".{" + str(length_byte) + "}")
# --------------------------------------------------------------------------------------------------------------

                hexvalue_baseSW = hexvalue_baseSW.lower()
                hexvalue_baseSW = str(".*" + hexvalue_baseSW + ".*")
            else:

            # danh cho co length byter thi su dung
                # count_hexvalue_baseSW = len(row_list_hex_values_baseSW)
                # count_hexvalue_baseSW = int(count_hexvalue_baseSW) // 2
                # if str(count_hexvalue_baseSW) < row_list_length_byte_baseSW:
                #     length_byte = (int(row_list_length_byte_baseSW) - int(count_hexvalue_baseSW)) * 2
                #     hexvalue_baseSW = row_list_hex_values_baseSW.lower()
                #     hexvalue_baseSW = str(hexvalue_baseSW + ".{" + str(length_byte) + "}")
# ------------------------------------------------------------------------------------------------------------------

                hexvalue_baseSW = row_list_hex_values_baseSW.lower()
                hexvalue_baseSW = str(".*" + hexvalue_baseSW + ".*")
            
            # print("hoan thanh",hexvalue_baseSW)
        if str(row_list_name_baseSW) == "None":
            row_list_name_baseSW = ""
            
        # if str(row_list_name_baseSW) != "Supplier Software number":

        if str(row_list_type_baseSW) == "APPL":
            
            ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW)+' in Application', '1) Send service 0x22 to the camera for the DID ' +
                        str(row_list_DID_baseSW) + ' using physical addressing', '1) -', '1) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
        
        if str(row_list_type_baseSW) == "FBL":
            # ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW) + ' in Programming', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Wait 3s\n4) Check active session should be Extended\n4) Wait 3s\n5) Change to Programming session with Service 0x10 02\n6) Wait 3s\n7) Check active session should be Programming\n8) Wait 3s\n9) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + ' using physical addressing\n10) Wait 3s\n11) Change to Default session with Service 0x10 01\n12) Wait 3s\n13) Check active session should be Default\n14) Wait 3s\n15) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + 'using physical addressing', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -\n13) -\n14) -\n15) -', '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(3000)\n5) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}3, Regexp)\n6) wait(1000)\n7) RequestResponse(1002, 5002.*, Regexp)\n8) wait(3000)\n9) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}2, Regexp)\n10) wait(3000)\n11) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)\n12) RequestResponse(1001, 5001.*, Regexp)\n13) wait(3000)\n14) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}2, Regexp)\n15) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
            ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW) + ' in Programming', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Wait 1s\n4) Change to Programming session with Service 0x10 02\n5) Wait 5s\n6) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + ' using physical addressing\n7) Wait 3s\n8) Change to Default session with Service 0x10 01\n9) Wait 1s\n10) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + 'using physical addressing', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -', '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(1002, 5002.*, Regexp)\n5) wait(5000)\n6) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)\n7) wait(3000)\n8) RequestResponse(1001, 5001.*, Regexp)\n9) wait(1000)\n10) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

        if str(row_list_type_baseSW) == "RBEOL":
            ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW) + ' in RBEOL', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID ' +
                str(row_list_DID_baseSW) + ' using physical addressing\n5) Reset ECU\n6) Wait 3s\n7) Send 1001\n8) Wait 3s', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)\n5) RequestResponse(1101, 5101, Equal)\n6) wait(3000)\n7) RequestResponse(1001, 5001.*, Regexp)\n8) wait(3000)', 'Automated Testcase', 'implemented', baseSW, ''])
            
        number4 += 1
        o += 1
        j += 1
        k += 1

    return id


def DID_latestSW(ws3, wb3, id, number1, number2, number3, number4, direct,dummy):
    if direct == '':
        wb3 = load_workbook('RFvalue.xlsx')
    else:
        wb3 = load_workbook(str(direct))

    ws3 = wb3 .active
    ws3 = wb3['RFvalue_latestSW']
    sheet2 = wb3.worksheets[1]
    row_count2 = sheet2.max_row
    for row in range(1, 2):
        for col in range(8, 9):
            char = get_column_letter(col)
            latestSW = ws3[char + str(row)].value

    i = 0
    o = 2
    j = 3
    k = 1
    # c = 0
    status_dummy = 0
    number4 += 1

    i = 0

    while k < row_count2:
        for row in range(o, j):
            for col in range(1, 2):
                char = get_column_letter(col)
                row_list_DID_latestSW = ws3[char + str(row)].value
                row_list_DID_latestSW_lowercase = str(
                    row_list_DID_latestSW).lower()
                if str(row_list_DID_latestSW) == "None":
                    return id
                # print(ws3[char + str(row)].value)
        for row in range(o, j):
            for col in range(2, 3):
                char = get_column_letter(col)
                row_list_name_latestSW = ws3[char + str(row)].value
                if str(row_list_DID_latestSW) == "None":
                    return id
                # print(ws3[char + str(row)].value)
        # for row in range(o, j):
        #     for col in range(3, 4):
        #         char = get_column_letter(col)
        #         row_list_length_byte_latestSW = ws3[char + str(row)].value
        #         if str(row_list_DID_latestSW) == "None":
        #             return id
                # print(ws[char + str(row)].value)
                # print(row_list_length_byte_latestSW)
        for row in range(o, j):
            for col in range(3, 4):
                char = get_column_letter(col)
                row_list_ASCII_values_latestSW = ws3[char + str(row)].value
                if str(row_list_DID_latestSW) == "None":
                    return id
                # print(ws3[char + str(row)].value)
        
        for row in range(o, j):
            for col in range(4, 5):
                char = get_column_letter(col)
                row_list_hex_values_latestSW = ws3[char + str(row)].value

        for row in range(o, j):
            for col in range(5, 6):
                char = get_column_letter(col)
                row_list_type_latestSW = ws3[char + str(row)].value
                print("latest" + str(row_list_type_latestSW))

        for row in range(o, j):
            for col in range(6, 7):
                char = get_column_letter(col)
                row_list_hex_values_DummySW = ws3[char + str(row)].value
                print("latest" + str(row_list_hex_values_DummySW))
        
        if str(row_list_DID_latestSW) != "None" :
            id += 1
            # check lenghth byte
            hexvalue_latestSW = ""
            
            # length_byte = ""
            if dummy == "dummy":
                if str(row_list_name_latestSW) == "Supplier Software number":
                    # length_byte = int(row_list_length_byte_latestSW) * 2
                    # print(length_byte)
                    # hexvalue_latestSW = str(".{" + str(length_byte) + "}")
                    hexvalue_latestSW = str(".*")
                    # print('dung')
                    print(hexvalue_latestSW)
                    dummy = ""
                    # c = 1
                    status_dummy  = 1
                    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_latestSW) + ' ' + str(row_list_name_latestSW), 'To check value of the DID ' + str(row_list_DID_latestSW), '1) Send service 0x22 to the camera for the DID ' +
                        str(row_list_DID_latestSW) + ' using physical addressing', '1) -', '1) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])                                    
                    
            else:
                if status_dummy == 0:
                    if str(row_list_ASCII_values_latestSW) == 'None' and str(row_list_hex_values_latestSW) == 'None':
                        # length_byte = int(row_list_length_byte_latestSW) * 2
                        # print(length_byte)
                        # hexvalue_latestSW = str(".{" + str(length_byte) + "}")
                        hexvalue_latestSW = str(".*")
                        # print('dung')
                        # print(hexvalue_latestSW)
                        # c = 1
                    else:
                        if str(row_list_hex_values_latestSW) == "None":
                            # change ascii sang hex value
                            for i in str(row_list_ASCII_values_latestSW):
                                hexvalue_latestSW += hex(ord(i))[2:]

                        # danh cho co length byter thi su dung
                            # # print(hexvalue_latestSW)
                            # count_hexvalue_latestSW = len(hexvalue_latestSW)
                            # count_hexvalue_latestSW = int(count_hexvalue_latestSW) // 2
                            # # print(count_hexvalue_latestSW)
                            # # print(type(row_list_length_byte_latestSW))
                            # if str(count_hexvalue_latestSW) < row_list_length_byte_latestSW:
                            #     # print("Day la do dai byte",row_list_length_byte_latestSW)
                            #     length_byte = (int(row_list_length_byte_latestSW) -
                            #                 int(count_hexvalue_latestSW)) * 2
                            #     # print("byte bi thieu", length_byte)
                            #     hexvalue_latestSW = hexvalue_latestSW.lower()
                            #     hexvalue_latestSW = str(hexvalue_latestSW + ".{" + str(length_byte) + "}")
                            #     c = 1
                            # else:
                            #     c = 0
                            hexvalue_latestSW = hexvalue_latestSW.lower()
                            hexvalue_latestSW = str(".*" + hexvalue_latestSW + ".*")
                        else:
                        # danh cho co length byter thi su dung
                            # count_hexvalue_latestSW = len(row_list_hex_values_latestSW)
                            # count_hexvalue_latestSW = int(count_hexvalue_latestSW) // 2
                            # if str(count_hexvalue_latestSW) < row_list_length_byte_latestSW:
                            #     length_byte = (int(row_list_length_byte_latestSW) - int(count_hexvalue_latestSW)) * 2
                            #     hexvalue_latestSW = row_list_hex_values_latestSW.lower()
                            #     hexvalue_latestSW = str(hexvalue_latestSW + ".{" + str(length_byte) + "}")
                            #     c = 1
                            hexvalue_latestSW = row_list_hex_values_latestSW.lower()
                            hexvalue_latestSW = str(".*" + hexvalue_latestSW + ".*")
                if  status_dummy  == 1:
                    if str(row_list_hex_values_DummySW) == 'None':
                        if str(row_list_ASCII_values_latestSW) == 'None' and str(row_list_hex_values_latestSW) == 'None':
                            
                            hexvalue_latestSW = str(".*")
                            
                        else:
                            if str(row_list_hex_values_latestSW) == "None":
                                # change ascii sang hex value
                                for i in str(row_list_ASCII_values_latestSW):
                                    hexvalue_latestSW += hex(ord(i))[2:]
                            
                                hexvalue_latestSW = hexvalue_latestSW.lower()
                                hexvalue_latestSW = str(".*" + hexvalue_latestSW + ".*")
                            else:
                            
                                hexvalue_latestSW = row_list_hex_values_latestSW.lower()
                                hexvalue_latestSW = str(".*" + hexvalue_latestSW + ".*")
                    else:

                        hexvalue_latestSW = row_list_hex_values_DummySW.lower()
                        hexvalue_latestSW = str(".*" + hexvalue_latestSW + ".*")                                            
                    # print(hexvalue_latestSW)
            
            # if str(row_list_name_latestSW) == "None":
            #     row_list_name_latestSW = ""
            
            # if c == 1:
            #     ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_latestSW) + ' ' + str(row_list_name_latestSW), 'To check value of the DID ' + str(row_list_DID_latestSW), '1) Send service 0x22 to the camera for the DID ' +
            #                 str(row_list_DID_latestSW) + ' using physical addressing', '1) -', '1) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
            # if c == 0:
            #     ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_latestSW) + ' ' + str(row_list_name_latestSW), 'To check value of the DID ' + str(row_list_DID_latestSW), '1) Send service 0x22 to the camera for the DID ' +
            #                 str(row_list_DID_latestSW) + ' using physical addressing', '1) -', '1) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Equal)', 'Automated Testcase', 'implemented', latestSW, ''])
            
                if str(row_list_name_latestSW) == "None":
                    row_list_name_latestSW = ""
            
                # if str(row_list_name_baseSW) != "Supplier Software number":

                if str(row_list_type_latestSW) == "APPL":

                    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_latestSW) + ' ' + str(row_list_name_latestSW), 'To check value of the DID ' + str(row_list_DID_latestSW), '1) Send service 0x22 to the camera for the DID ' +
                        str(row_list_DID_latestSW) + ' using physical addressing', '1) -', '1) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

                if str(row_list_type_latestSW) == "FBL":
                    # ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW) + ' in Programming', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Wait 3s\n4) Check active session should be Extended\n4) Wait 3s\n5) Change to Programming session with Service 0x10 02\n6) Wait 3s\n7) Check active session should be Programming\n8) Wait 3s\n9) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + ' using physical addressing\n10) Wait 3s\n11) Change to Default session with Service 0x10 01\n12) Wait 3s\n13) Check active session should be Default\n14) Wait 3s\n15) Send service 0x22 to the camera for the DID ' +str(row_list_DID_baseSW) + 'using physical addressing', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -\n13) -\n14) -\n15) -', '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(3000)\n5) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}3, Regexp)\n6) wait(1000)\n7) RequestResponse(1002, 5002.*, Regexp)\n8) wait(3000)\n9) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}2, Regexp)\n10) wait(3000)\n11) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)\n12) RequestResponse(1001, 5001.*, Regexp)\n13) wait(3000)\n14) RequestResponse(' + '22' + str(DID_check_active_baseSW_lowercase) + ','+'62'+str(DID_check_active_baseSW_lowercase) + '.{'+'1}2, Regexp)\n15) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
                    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_latestSW) + ' ' + str(row_list_name_latestSW), 'To check value of the DID ' + str(row_list_DID_latestSW) + ' in Programming', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Wait 1s\n4) Change to Programming session with Service 0x10 02\n5) Wait 5s\n6) Send service 0x22 to the camera for the DID ' +str(row_list_DID_latestSW) + ' using physical addressing\n7) Wait 3s\n8) Change to Default session with Service 0x10 01\n9) Wait 1s\n10) Send service 0x22 to the camera for the DID ' +str(row_list_DID_latestSW) + 'using physical addressing', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -', '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(1002, 5002.*, Regexp)\n5) wait(5000)\n6) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Regexp)\n7) wait(3000)\n8) RequestResponse(1001, 5001.*, Regexp)\n9) wait(1000)\n10) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

                if str(row_list_type_latestSW) == "RBEOL":
                    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_latestSW) + ' ' + str(row_list_name_latestSW), 'To check value of the DID ' + str(row_list_DID_latestSW) + ' in RBEOL', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID ' +
                        str(row_list_DID_latestSW) + ' using physical addressing\n5) Reset ECU\n6) Wait 3s\n7) Send 1001\n8) Wait 3s', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + str(hexvalue_latestSW) + ', Regexp)\n5) RequestResponse(1101, 5101, Equal)\n6) wait(3000)\n7) RequestResponse(1001, 5001.*, Regexp)\n8) wait(3000)', 'Automated Testcase', 'implemented', latestSW, ''])
            
            number4 += 1
            o += 1
            j += 1
            k += 1
        else:
            return id
    return id


def variant_base_sw(id, number1, number2, number3, number4, tasks):
    direct = Input_path_text.get()
    if direct == '':
        wb = load_workbook('RFvalue.xlsx')
    else:
        wb = load_workbook(direct)
    ws = wb.active
    ws = wb['RFvalue_baseSW']

    for row in range(1, 2):
        for col in range(7, 8):
            char = get_column_letter(col)
            baseSW = ws[char + str(row)].value
            # print(baseSW)

    for row in range(3, 4):
        for col in range(7, 8):
            char = get_column_letter(col)
            row_Variant_BaseSW = ws[char + str(row)].value
            row_Variant_BaseSW_lowercase = str(row_Variant_BaseSW).lower()
            # print(row_Variant_BaseSW)
    for row in range(4, 5):
        for col in range(7, 8):
            char = get_column_letter(col)
            row_DID_check_variant_BaseSW = ws[char + str(row)].value
            row_DID_check_variant_BaseSW_lowercase = str(
                row_DID_check_variant_BaseSW).lower()
            # print(row_DID_check_variant_BaseSW)
    
    for row in range(5, 6):
        for col in range(7, 8):
            char = get_column_letter(col)
            row_Security_level_BaseSW = ws[char + str(row)].value
            if str(row_Security_level_BaseSW) == "None":
                messagebox.showerror("ERROR", "No Value SECURITY LEVEL, Please add SECURITY LEVEL")
                break            
            # row_DID_check_variant_BaseSW_lowercase = str(
            #     row_DID_check_variant_BaseSW).lower()
    
    if str(row_Security_level_BaseSW) == '0':
        if str(row_Variant_BaseSW) != "None" and str(row_DID_check_variant_BaseSW) != "None":
            ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) wait\n4) Select variant\n5) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(2e' + str(row_Variant_BaseSW_lowercase) + ', 6e' + str(row_DID_check_variant_BaseSW_lowercase) + ', Equal)\n5) RequestResponse(22' + str(row_DID_check_variant_BaseSW_lowercase) + ', 62' + str(row_Variant_BaseSW_lowercase) + ', Equal)', 'Automated Testcase', 'implemented', baseSW, ''])
        else:
            ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check_variant (Variant is default)', 'To check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) wait\n4) Check variant', '1) -\n2) -\n3) -\n4) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(22' + str(row_DID_check_variant_BaseSW_lowercase) + ', 62.*, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    if str(row_Security_level_BaseSW) == '1':
        if str(row_Variant_BaseSW) != "None" and str(row_DID_check_variant_BaseSW) != "None":
            ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) wait\n7) Select variant\n8) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel'+str(row_Security_level_BaseSW)+'(1;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel'+str(row_Security_level_BaseSW)+'(0;0))\n6) wait(1000)\n7) RequestResponse(2e' + str(row_Variant_BaseSW_lowercase) + ', 6e' + str(row_DID_check_variant_BaseSW_lowercase) + ', Equal)\n8) RequestResponse(22' + str(row_DID_check_variant_BaseSW_lowercase) + ', 62' + str(row_Variant_BaseSW_lowercase) + ', Equal)', 'Automated Testcase', 'implemented', baseSW, ''])
        else:
            ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check_variant (Variant is default)', 'To check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) wait\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel'+str(row_Security_level_BaseSW)+'(1;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel'+str(row_Security_level_BaseSW)+'(0;0))\n6) wait(1000)\n7) RequestResponse(22' + str(row_DID_check_variant_BaseSW_lowercase) + ', 62.*, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    return id


def variant_latest_sw(id, number1, number2, number3, number4, tasks):
    direct = Input_path_text.get()
    if direct == '':
        wb3 = load_workbook('RFvalue.xlsx')
    else:
        wb3 = load_workbook(str(direct))
    ws3 = wb3.active
    ws3 = wb3['RFvalue_latestSW']
    sheet2 = wb3.worksheets[1]
    row_count = sheet2.max_row
    for row in range(1, 2):
        for col in range(8, 9):
            char = get_column_letter(col)
            latestSW = ws3[char + str(row)].value
            # print(latestSW)

    for row in range(3, 4):
        for col in range(8, 9):
            char = get_column_letter(col)
            row_Variant_LatestSW = ws3[char + str(row)].value
            row_Variant_LatestSW_lowercase = str(row_Variant_LatestSW).lower()
            # print(row_Variant_LatestSW)
    for row in range(4, 5):
        for col in range(8, 9):
            char = get_column_letter(col)
            row_DID_check_variant_LatestSW = ws3[char + str(row)].value
            row_DID_check_variant_LatestSW_lowercase = str(
                row_DID_check_variant_LatestSW).lower()
            # print(row_DID_check_variant_LatestSW)

    for row in range(5, 6):
        for col in range(8, 9):
            char = get_column_letter(col)
            row_Security_level_latestSW = ws3[char + str(row)].value
            # row_DID_check_variant_BaseSW_lowercase = str(
            #     row_DID_check_variant_BaseSW).lower()
    
    if str(row_Security_level_latestSW) == '0':
        if str(row_Variant_LatestSW) != "None" and str(row_DID_check_variant_LatestSW) != "None":
            ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) wait\n4) Select variant\n5) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(2e' + str(row_Variant_LatestSW_lowercase) + ', 6e' + str(row_DID_check_variant_LatestSW_lowercase) + ', Equal)\n5) RequestResponse(22' + str(row_DID_check_variant_LatestSW_lowercase) + ', 62' + str(row_Variant_LatestSW_lowercase) + ', Equal)', 'Automated Testcase', 'implemented', latestSW, ''])
        else:
            ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check_variant (Variant is default)', 'To check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) wait\n4) Check variant', '1) -\n2) -\n3) -\n4) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) wait(1000)\n4) RequestResponse(22' + str(row_DID_check_variant_LatestSW_lowercase) + ', 62.*, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    if str(row_Security_level_latestSW) != '0':
        if str(row_Variant_LatestSW) != "None" and str(row_DID_check_variant_LatestSW) != "None":
            ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) wait\n7) Select variant\n8) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel'+str(row_Security_level_latestSW)+'(1;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel'+str(row_Security_level_latestSW)+'(0;0))\n6) wait(1000)\n7) RequestResponse(2e' + str(row_Variant_LatestSW_lowercase) + ', 6e' + str(row_DID_check_variant_LatestSW_lowercase) + ', Equal)\n8) RequestResponse(22' + str(row_DID_check_variant_LatestSW_lowercase) + ', 62' + str(row_Variant_LatestSW_lowercase) + ', Equal)', 'Automated Testcase', 'implemented', latestSW, ''])
        else:
            ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check_variant (Variant is default)', 'To check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) wait\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                        '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel'+str(row_Security_level_latestSW)+'(1;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel'+str(row_Security_level_latestSW)+'(0;0))\n6) wait(1000)\n7) RequestResponse(22' + str(row_DID_check_variant_LatestSW_lowercase) + ', 62.*, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    return id
# -------------------------------------------------------------------------------------------------------------
# def DID_RBEOL(id, number1, number2, number3, number4, tasks):
    #  if str(row_list_type_baseSW) == "RBEOL":
            # ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW)+' in Application', '1) Send service 0x22 to the camera for the DID ' +
            #             str(row_list_DID_baseSW) + ' using physical addressing', '1) -', '1) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + str(hexvalue_baseSW) + ', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    # ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259 using physical addressing\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -',
    #             '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)', 'Automated Testcase', 'implemented', baseSW, ''])

# -------------------------------------------------------------------------------------------------------------

# SOURCE

def run_program():
    run_btn_text.set("Loading...")

    direct = Input_path_text.get()
    locate_save = Output_path_text.get()


    if direct == '':
        wb = load_workbook('RFvalue.xlsx')
        wb3 = load_workbook('RFvalue.xlsx')
        wb4 = load_workbook('RFvalue.xlsx')
        wb6 = load_workbook('TC_RF.xlsx')
    else:
        wb = load_workbook(direct)
        wb3 = load_workbook(direct)
        wb4 = load_workbook(direct)
        wb6 = load_workbook(locate_save + '/'+'TC_RF.xlsx')
    # wb3 = load_workbook(direct)
    ws = wb.active
    ws3 = wb3.active
    ws4 = wb4.active
    ws = wb['RFvalue_baseSW']
    ws3 = wb3['RFvalue_latestSW']
    ws4 = wb4['General Information'] 
    sheet2 = wb.worksheets[1]
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    row_count2 = sheet2.max_row
    tasks = row_count2 + row_count + 160

    # # try:
    #     # os.system('TASKKILL /F /IM EXCEL.exe')
    # if locate_save == '':
    #     # wb6 = load_workbook('TC_RF.xlsx')
    #     os.remove("TC_RF.xlsx")
    # else:
    #     # wb6 = load_workbook(locate_save + '/'+'TC_RF.xlsx')
    #     os.remove(locate_save + '/'+'TC_RF.xlsx')
    
    
    # wb2 = Workbook()
    # ws2 = wb2.active
    # ws2.title = "TC_RF"
    # wb2.save("TC_RF.xlsx")
    # os.close("TC_RF.xlsx")
    # except:
    #     try:
    #         wb2 = Workbook()
    #         ws2 = wb2.active
    #         ws2.title = "TC_RF"
    #         wb2.save("TC_RF.xlsx")
    #         # os.system('TASKKILL /F /IM EXCEL.exe')
    #     except OSError:
    #         print('Failed creating the file')
    #     else:
    #         print('File created') 

    # if locate_save == '':
    #     wb6 = load_workbook('TC_RF.xlsx')
    # else:
    #     wb6 = load_workbook(locate_save + '/'+'TC_RF.xlsx')

    ws6 = wb6.active
    ws6 = wb6['TC_RF']
    sheet3 = wb6.worksheets[0]
    row_count3 = sheet3.max_row

    print(row_count3)
    print(tasks)

    if row_count3 >= tasks:
        # ws6 = wb6.active
        # ws7 = wb7['TC_RF']
        
        # print(row_count4)
        n = 0
        while n < row_count3:
            ws2.delete_rows(1)
            # dlrow += 1
            n += 1
            # print(dlrow)
            print("dang xoa")
        # wb7.save('TC_RF.xlsx')
        # wb7 .close()
        print("done")
        n = 0
        # row_count4 = 0

    # if locate_save == '':
    #     wb6.save('TC_RF.xlsx')
    # else:
    #     wb6.save(locate_save + '/'+'TC_RF.xlsx')
    # wb6.close()

    for row in range(1, 2):
        for col in range(7, 8):
            char = get_column_letter(col)
            baseSW = ws[char + str(row)].value

    for row in range(1, 2):
        for col in range(2, 3):
            char = get_column_letter(col)
            ticket_baseSW = ws4[char + str(row)].value
            ticket_latestSW = ws4[char + str(row)].value
            # print(ticket_baseSW)

    for row in range(1, 2):
        for col in range(8, 9):
            char = get_column_letter(col)
            latestSW = ws3[char + str(row)].value

    for row in range(2, 3):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_counter_DID = ws4[char + str(row)].value


    for row in range(3, 4):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_Attempt_counter_DID = ws4[char + str(row)].value

    for row in range(4, 5):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_counter_step = ws4[char + str(row)].value


    for row in range(5, 6):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_Attempt_counter_step = ws4[char + str(row)].value

    # script begin
    id = 2
    number1 = 1
    number2 = 1
    number3 = 1
    number4 = 1
    border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(
        border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
    font_text_header = Font(name="Calibri", size=13,
                            color='00FFFFFF', bold=True)
    font_text = Font(name="Calibri", size=11, color='00000000', bold=False)
    alignment = Alignment(horizontal='center', vertical='center')
    ws2.append(['ID', 'XXX Component',  'Test Description', 'Test Steps',  'Test Response','Teststep keywords', 'ObjectType', 'TestStatus', 'Project', 'TestResult'])
    
    for col in range(1, 11):
        cell_header = ws2.cell(1, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text_header
        # cell_header.alignment = alignment
    ws2.append(['ID_'+str(id),  '1 REFFLASH', '','', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(2, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        # cell_header.alignment = alignment
    # ------------------------------------------------------------------------------------------------------
    # BEGIN TEST CASE 1
    # TEST CASE 1 base SW to latestSW M3
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) +' Base SW to Latest SW M3', '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH BASE_SW VIA UART script
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) +' Flash Base SW via UART', '', '', '', '', 'Test group', '', '', ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])

    for row in range(id - 2, id + 1):
        for col in range(1, 11):
            cell_header = ws2.cell(row, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART flash ' + str(baseSW), 'Detail information is mentioned in the ticket: ' +str(ticket_baseSW), '1) Flash Base Software '+ str(baseSW) +' via UART successful', '1) -', "1) TesterConfirm('Do you flash base software "+str(baseSW)+" via UART ?')", 'Automated Testcase', 'implemented', str(baseSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification','', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_base_sw(id, number1, number2, number3, number4, tasks)

    # id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    # number4 += 1
    direct = Input_path_text.get()
    id = DID_baseSW(ws2, wb, id, number1, number2, number3, number4, direct)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])
    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    # print(id)

    PC_step = 0
    PAC_step = 0
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL','', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # ws2.title = "TC_RF"
    # wb2.save('TC_RF.xlsx')

    # Step2 FLASH LATEST_SW M3 VIA Xflash TOOLS

    ws3 = wb .active
    ws3 = wb3['RFvalue_latestSW']

    sheet = wb.worksheets[1]
    row_count = sheet.max_row

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Latest SW M3 via X-Flash 1st',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via IFLASH successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via IFLASH ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    dummy  = ""
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4, direct, dummy)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = Programming_counter_step
    PAC_step = Programming_counter_step
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL','', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # ws2.title = "TC_RF"
    # wb2.save('TC_RF.xlsx')

    # END TEST CASE 1

    # ------------------------------------------------------------------------------------------------------
    # # BEGIN TEST CASE 2
    # # TEST CASE 2 base SW to latestSW M5
    number1 += 1
    number2 = 1
    number3 = 1
    number4 = 1
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) +
                ' Base SW to Latest SW M5', '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH BASE_SW VIA UART script
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Flash Base SW via UART',
                '', '', '', '', 'Test group', '', '', ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Flash SW', '',
                '', '', '', 'Test group', '', '', ''])

    for row in range(id - 2, id + 1):
        for col in range(1, 11):
            cell_header = ws2.cell(row, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART flash ' + str(baseSW), 'Detail information is mentioned in the ticket: ' +str(ticket_baseSW), '1) Flash Base Software '+ str(baseSW) +' via UART successful', '1) -', "1) TesterConfirm(''Do you flash base software "+str(baseSW)+" via UART ?')", 'Automated Testcase', 'implemented', str(baseSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_base_sw(id, number1, number2, number3, number4, tasks)

    # id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    # number4 += 1
    direct = Input_path_text.get()
    id = DID_baseSW(ws2, wb, id, number1, number2, number3, number4, direct)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = 0
    PAC_step = 0
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # # ws2.title = "TC_RF"
    # # wb2.save('TC_RF.xlsx')

    # # Step2 FLASH LATEST_SW M5 1st VIA Xflash TOOLS

    # # # Reflash latest SW M5 via xflash tool

    ws3 = wb .active
    ws3 = wb3['RFvalue_latestSW']

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Latest SW M5 via X-Flash 1st','', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via IFLASH successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via IFLASH ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    # number4 += 1
    direct = Input_path_text.get()
    dummy = ""
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4, direct, dummy)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = Programming_counter_step
    PAC_step = Programming_Attempt_counter_step
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text


    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*5,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # # Step2 FLASH LATEST_SW M3 2nd VIA Xflash TOOLS

    # # # Reflash latest SW M3 via xflash tool

    ws3 = wb .active
    ws3 = wb3['RFvalue_latestSW']

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Latest SW M3 via X-Flash 2nd',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via IFLASH successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via IFLASH ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    dummy = ""
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4, direct, dummy)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = int(Programming_counter_step) + int(Programming_counter_step)
    PAC_step = int(Programming_Attempt_counter_step) + int(Programming_Attempt_counter_step)
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # # ws2.title = "TC_RF"
    # # wb2.save('TC_RF.xlsx')

    # # END TEST CASE 2

    # # ------------------------------------------------------------------------------------------------------

    # # BEGIN TEST CASE 3
    # # TEST CASE 3 latest SW to DummySW M3
    ws3 = wb .active
    ws3 = wb3['RFvalue_latestSW']

    number1 += 1
    number2 = 1
    number3 = 1
    number4 = 1
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + ' latest SW to Dummy SW M3',
                '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH latest_SW VIA UART script

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Flash latest SW via UART',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Flash SW', '',
                '', '', '', 'Test group', '', '', ''])

    for row in range(id - 2, id + 1):
        for col in range(1, 11):
            cell_header = ws2.cell(row, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART Flash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via UART successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via UART ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    dummy = ""
    id = DID_latestSW(ws2, wb, id, number1, number2, number3, number4, direct,dummy)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = 0
    PAC_step = 0
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # # ws2.title = "TC_RF"
    # # wb2.save('TC_RF.xlsx')

    # # Step2 FLASH DUMMY_SW M3 VIA Xflash TOOLS

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Dummy SW M3 via X-Flash 1st',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via IFLASH successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via IFLASH ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    dummy = "dummy"
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4, direct, dummy)
    # print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = Programming_counter_step
    PAC_step = Programming_Attempt_counter_step
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # # ws2.title = "TC_RF"
    # # wb2.save('TC_RF.xlsx')

    # # END TEST CASE 3

    # # ------------------------------------------------------------------------------------------------------
    # # BEGIN TEST CASE 4
    # # TEST CASE 4 latest SW to DummySW M5
    number1 += 1
    number2 = 1
    number3 = 1
    number4 = 1
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + ' latest SW to Dummy SW M5',
                '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH latest_SW VIA UART script

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Flash latest SW via UART',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Flash SW', '',
                '', '', '', 'Test group', '', '', ''])

    for row in range(id - 2, id + 1):
        for col in range(1, 11):
            cell_header = ws2.cell(row, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART Flash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via UART successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via UART ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    # number4 += 1
    direct = Input_path_text.get()
    dummy = ""
    id = DID_latestSW(ws2, wb, id, number1, number2, number3, number4, direct,dummy)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = 0
    PAC_step = 0
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # ws2.title = "TC_RF"
    # wb2.save('TC_RF.xlsx')

# --------------------------------------------------------------------------------------------------------
    # Step2 FLASH DUMMY_SW M5 1st VIA Xflash TOOLS
    # # Reflash Dummy SW M5 via xflash tool

    ws3 = wb .active
    ws3 = wb3['RFvalue_latestSW']

    # number = df.shape[0]
    # print(number)
    sheet = wb.worksheets[1]
    row_count = sheet.max_row

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Dummy SW M5 via X-Flash 1st',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via IFLASH successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via IFLASH ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    # DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
    dummy = "dummy"
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4, direct, dummy)
    # print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = Programming_counter_step
    PAC_step = Programming_Attempt_counter_step
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*5,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])

    # Step2 FLASH DUMMY_SW M3 2nd VIA Xflash TOOLS

    # # Reflash Dummy SW M3 via xflash tool

    ws3 = wb .active
    ws3 = wb3['RFvalue_latestSW']

    # number = df.shape[0]
    # print(number)
    sheet = wb.worksheets[1]
    row_count = sheet.max_row

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Dummy SW M3 via X-Flash 2nd',
                '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash ' + str(latestSW), 'Detail information is mentioned in the ticket: ' +str(ticket_latestSW), '1) Flash Latest Software '+ str(latestSW) +' via IFLASH successful', '1) -', "1) TesterConfirm('Do you flash latest software "+str(latestSW)+" via IFLASH ?')", 'Automated Testcase', 'implemented', str(latestSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    # DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
    dummy = "dummy"
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4, direct, dummy)
    # print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    PC_step = int(Programming_counter_step) + int(Programming_counter_step)
    PAC_step = int(Programming_Attempt_counter_step) + int(Programming_Attempt_counter_step)
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(Programming_counter_DID) + '_ProgrammingCounter', 'To check value of the DID '+ str(Programming_counter_DID), '1) Send service 0x22 to the camera for the DID '+ str(Programming_counter_DID) + ' using physical addressing',
                '1) -', '1) RequestResponse(22' + str(Programming_counter_DID) + ', 62'+ str(Programming_counter_DID)+'.*'+str(PC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' '+ str(Programming_Attempt_counter_DID) + '_ProgrammingAttemptCounter', 'To check value of the DID '+ str(Programming_Attempt_counter_DID),
                '1) Send service 0x22 to the camera for the DID '+ str(Programming_Attempt_counter_DID) + ' using physical addressing', '1) -', '1) RequestResponse(22' + str(Programming_Attempt_counter_DID) + ', 62'  + str(Programming_Attempt_counter_DID) + '.*'+str(PAC_step)+', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    # Check  Check DID in RBEOL
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check DID in RBEOL', 'To check value of the DID F1E0, F1DD, 4255, 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0\n5) Send service 0x22 to the camera for the DID F1DD\n6) Send service 0x22 to the camera for the DID 4255\n7) Send service 0x22 to the camera for the DID 4259\n8) Reset ECU\n9) Wait 3s\n10) Send 1001\n11) Wait 3s\n12) Reset ECU', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -\n8) -\n9) -\n10) -\n11) -\n12) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)\n5) RequestResponse(22f1dd,62f1dd.*,Regexp)\n6) RequestResponse(224255,624255.*,Regexp)\n7) RequestResponse(224259,624259.*,Regexp)\n8) RequestResponse(1101, 5101, Equal)\n9) wait(3000)\n10) RequestResponse(1001, 5001.*, Regexp)\n11) wait(3000)\n12) RequestResponse(1101, 5101, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])
    
    for row in range(id - id + 1, id + 1):
        for col in range(1, 11):
            cell_header = ws2.cell(row, col)
            # used hex code for red color
            cell_header.border = border
    # END TEST CASE 4

    locate_save = Output_path_text.get()
    ws2.title = "TC_RF"
    print(locate_save)
    if locate_save == '':
        # print('Ok')
        wb2.save('TC_RF.xlsx')
    else:
        wb2.save(locate_save + '/'+'TC_RF.xlsx')

    if locate_save == '':
        wb7 = load_workbook('TC_RF.xlsx')
    else:
        wb7 = load_workbook(locate_save + '/'+'TC_RF.xlsx')

    ws7 = wb7['TC_RF']
    sheet4 = wb7.worksheets[0]
    row_count4 = sheet4.max_row

    # sheet4 = wb6.worksheets[0]
    sheet4.column_dimensions['A'].width = 30
    sheet4.column_dimensions['B'].width = 30
    sheet4.column_dimensions['C'].width = 30
    sheet4.column_dimensions['D'].width = 30
    sheet4.column_dimensions['E'].width = 30
    sheet4.column_dimensions['F'].width = 30
    sheet4.column_dimensions['G'].width = 30
    sheet4.column_dimensions['H'].width = 30
    sheet4.column_dimensions['I'].width = 30
    sheet4.column_dimensions['J'].width = 30
    print(row_count4)
    print(tasks)


    if locate_save == '':
        wb7.save('TC_RF.xlsx')
    else:
        wb7.save(locate_save + '/'+'TC_RF.xlsx')
    
    # wb6.close()
    # print(id)
    # print(percent)
    # print(tasks)
    # print(row_count)
    # print(row_count2)
    run_btn_text.set("RUN")
    tkinter.messagebox.showinfo("GREAT!", "Test case RFlash tool created successfully")


def start_program():
    direct = Input_path_text.get()
    locate_save = Output_path_text.get()
    if direct == '':
        wb = load_workbook('RFvalue.xlsx')
        wb3 = load_workbook('RFvalue.xlsx')
        wb4 = load_workbook('RFvalue.xlsx')
    else:
        wb = load_workbook(direct)
        wb3 = load_workbook(direct)
        wb4 = load_workbook(direct)
        
    
    ws = wb.active
    ws = wb['RFvalue_baseSW']
    ws3 = wb3.active
    ws3 = wb3['RFvalue_latestSW']
    ws4 = wb4.active
    ws4 = wb4['General Information']
    sheet2 = wb.worksheets[1]
    sheet = wb.worksheets[0]
    row_count = sheet.max_row

    for row in range(2, 3):
        for col in range(1, 2):
            char = get_column_letter(col)
            row_list_DID_baseSW = ws[char + str(row)].value
            row_list_DID_latestSW = ws3[char + str(row)].value

    for row in range(2, 3):
        for col in range(1, 2):
            char = get_column_letter(col)
            row_list_DID_baseSW = ws[char + str(row)].value
            row_list_DID_latestSW = ws3[char + str(row)].value
    
    for row in range(1, 2):
        for col in range(7, 8):
            char = get_column_letter(col)
            baseSW = ws[char + str(row)].value
    for row in range(1, 2):
        for col in range(8, 9):
            char = get_column_letter(col)
            latestSW = ws3[char + str(row)].value

    for row in range(1, 2):
        for col in range(2, 3):
            char = get_column_letter(col)
            ticket_baseSW = ws4[char + str(row)].value
            ticket_latestSW = ws4[char + str(row)].value

    for row in range(2, 3):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_counter_DID = ws4[char + str(row)].value


    for row in range(3, 4):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_Attempt_counter_DID = ws4[char + str(row)].value

    for row in range(4, 5):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_counter_step = ws4[char + str(row)].value


    for row in range(5, 6):
        for col in range(2, 3):
            char = get_column_letter(col)
            Programming_Attempt_counter_step = ws4[char + str(row)].value




    if str(row_list_DID_baseSW) == "None":
        messagebox.showerror(
            "ERROR", "No DID in row 1 at sheet RFvalue_baseSW, Please add DID")

    if str(row_list_DID_latestSW) == "None":
        messagebox.showerror(
            "ERROR", "No DID in row 1 at sheet RFvalue_latestSW, Please add DID")

    if str(baseSW) == "None":
        messagebox.showerror("ERROR", "BaseSW Name is invalid")

    if str(ticket_baseSW) == "None":
        messagebox.showerror("ERROR", "ticket_baseSW Name is invalid")

    if str(latestSW) == "None":
        messagebox.showerror("ERROR", "latestSW Name is invalid")

    if str(ticket_latestSW) == "None":
        messagebox.showerror("ERROR", "ticket_latestSW Name is invalid")

    if str(Programming_counter_DID) == "None":
        messagebox.showerror("ERROR", "Programming_counter_DID is invalid")

    if str(Programming_Attempt_counter_DID) == "None":
        messagebox.showerror("ERROR", "Programming_Attempt_counter_DID is invalid")

    if str(Programming_counter_step) == "None":
        messagebox.showerror("ERROR", "Programming_counter_step is invalid")

    if str(Programming_Attempt_counter_step) == "None":
        messagebox.showerror("ERROR", "Programming_Attempt_counter_step is invalid")


    if str(baseSW) != "None" and str(ticket_baseSW) != "None" and str(latestSW) != "None" and str(ticket_latestSW) != "None" and str(row_list_DID_baseSW) != "None" and str(row_list_DID_latestSW) != "None" and str(Programming_counter_DID) != "None" and str(Programming_Attempt_counter_DID) != "None" and str(Programming_counter_step) != "None" and str(Programming_Attempt_counter_step) != "None":
        print('run_program')
        # print(Programing_counter_text.get())
        if locate_save == '':
            try:
                # with open("TC_RF.xlsx", "r") as file:
                    # print("File was already have")
                fd = os.open("TC_RF.xlsx", os.O_RDWR)
                os.close(fd)
                
                run_program()
            except IOError:
                messagebox.showerror("ERROR","File TC_RF.xlsx has opened already. Please close the file in the same folder before push run button")
                # print("File has opened already. Please close the TC_RF.xlsx before push run button")
            
        else:
            try:
                # with open(locate_save + '/'+'TC_RF.xlsx', "r") as file:
                #     print("File has been remove")
                fd2 = os.open(locate_save + '/'+'TC_RF.xlsx', os.O_RDWR)
                os.close(fd2)
                
                run_program()
            except IOError:
                messagebox.showerror('ERROR','File TC_RF.xlsx has opened already. Please close the file at :'+ locate_save + '/'+'TC_RF.xlsx'+' before push run button')
                # print("File has opened already. Please close the TC_RF.xlsx before push run button")
        


# app front end
app = tk.Tk()

app.title('Basic Software Allied tool')
app.geometry('700x420')


create_value_file()


def open_file():
    browse_input_path_text.set("loading...")
    file_path = askopenfile(parent=app, mode='rb', title="Choose location take file", filetype=[
                            ("excel file", ".xlsx")])
    print("Original string: " + str(file_path))

    result_str = ""
    final_str = ""
    for i in range(0, len(str(file_path))):
        if i >= 26:
            result_str = result_str + str(file_path)[i]
    reverse_str = result_str[::-1]
    for i in range(0, len(reverse_str)):
        if i >= 2:
            final_str = final_str + reverse_str[i]
    complete_str = final_str[::-1]
    print(type(complete_str))
    print(complete_str)
    if file_path:
        Input_path_text.set(str(complete_str))
        browse_input_path_text.set("Browse")
    return complete_str


def save_file():
    # print("is this working??")
    browse_output_path_text.set("loading...")
    file_path2 = filedialog.askdirectory()
    print(file_path2)
    if file_path2:
        Output_path_text.set(str(file_path2))
        browse_output_path_text.set("Browse")
    return str(file_path2)


file_path = ""

# Part Base SW
frameall = tk.Frame(app)
frame1 = tk.Frame(frameall)
frame2 = tk.Frame(frameall)
    
    

Input_path_text = tk.StringVar()
Input_path_label = tk.Label(frame1, text='Input path', font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=0, column=0, sticky='w')
Input_path_entry = tk.Entry(frame1, textvariable=Input_path_text,
                            font='large_font', width=55).grid(row=1, column=0, sticky='w')

Output_path_text = tk.StringVar()
Output_path_label = tk.Label(frame1, text='Output path', font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=2, column=0, sticky='w')
Output_path_entry = tk.Entry(frame1, textvariable=Output_path_text,
                            font='large_font', width=55).grid(row=3, column=0, sticky='w')

# browse button open file
browse_input_path_text = tk.StringVar()
browse_btn_input_path = tk.Button(frame1, textvariable=browse_input_path_text, command=lambda: open_file(
), font="bold", width=7, height=1).grid(row=1, column=1, pady=5, padx=10)
browse_input_path_text.set("Browse")

# browse button save file
browse_output_path_text = tk.StringVar()
browse_btn_output_path = tk.Button(frame1, textvariable=browse_output_path_text, command=lambda: save_file(
), font="bold", width=7, height=1).grid(row=3, column=1, pady=5, padx=10)
browse_output_path_text.set("Browse")

# frame1.pack()

# Run program Buttons
run_btn_text = tk.StringVar()
run_btn = tk.Button(frame1, textvariable=run_btn_text, command=start_program,
                    font="bold", width=15).grid(row=4, column=0, columnspan=2, pady=20)
run_btn_text.set("RUN")

# frame1.pack()
# frame2.pack()

noneFill = tk.StringVar()
noneLabel = tk.Label(frame1, textvariable=noneFill, font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=5, column=0, columnspan=2)
# frame3.pack()

# frame1.pack()

# progress bar
bar = ttk.Progressbar(app, orient='horizontal', length=583, mode='determinate')

bar.place(relx=0.5, rely=0.6, anchor=CENTER)

# frame1.pack()

# frame4.pack()
frame1.pack()

# programing counter / attempt couter step
# step = ('1', '2', '3', '4', '5', '6', '7', '8', '9', '10')

# Programing_counter_text = tk.StringVar()
# Programing_counter_label = tk.Label(frame2, text='Programming_counter_step', font=(
#     'bold', 14), bg="#20bebe", fg="black").grid(row=0, column=0, padx=30, pady=10, sticky='w')
# Programing_counter = tk.Spinbox(frame2, textvariable=Programing_counter_text,
#                                 values=step, width=10, font=('helvetica', 15)).grid(row=1, column=0, padx=30)


# Programing_Attempt_counter_text = tk.StringVar()
# Programing_Attempt_counter_label = tk.Label(frame2, text='Programming_Attempt_counter_step', font=(
#     'bold', 14), bg="#20bebe", fg="black").grid(row=0, column=1, padx=30, pady=10, sticky='w')
# Programing_Attempt_counter = tk.Spinbox(frame2, textvariable=Programing_Attempt_counter_text,
#                                         values=step, width=10, font=('helvetica', 15)).grid(row=1, column=1, padx=30)
percent = tk.StringVar()
percentLabel = tk.Label(frame2, textvariable=percent, font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=5, column=0, columnspan=2)

noneFill = tk.StringVar()
noneLabel = tk.Label(frame2, textvariable=noneFill, font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=6, column=0, columnspan=2)
# frame2.pack()

noneFill = tk.StringVar()
noneLabel = tk.Label(frame2, textvariable=noneFill, font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=7, column=0, columnspan=2)

noneFill = tk.StringVar()
noneLabel = tk.Label(frame2, textvariable=noneFill, font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=8, column=0, columnspan=2)



frame2.pack()



frameall.place(relx=0.5, rely=0.5, anchor=CENTER)

# instruction

instruction = tk.Label(
    app, text="          Welcome to ReFlash tool create by dev Huynh Minh Dang", font=("helvetica", 14))
instruction_version = tk.Label(app, text="R1.3.1", font=("helvetica", 14))
instruction_version.pack(side="right", anchor='s')
instruction.pack(side="bottom", fill='both', anchor=CENTER)

frameall.configure(background="#20bebe")
frame1.configure(background="#20bebe")
frame2.configure(background="#20bebe")

app.configure(background="#20bebe")


# Start program
app.mainloop()
