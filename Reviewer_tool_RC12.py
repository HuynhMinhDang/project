from email import message
from msilib.schema import RadioButton
from turtle import textinput
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Color, numbers
from openpyxl.worksheet.datavalidation import DataValidation
import os
from struct import pack
import tkinter as tk
from PIL import Image, ImageTk
from tkinter.filedialog import askopenfile
from tkinter import CENTER, filedialog
from tkinter import HORIZONTAL, messagebox
from tkinter import ttk
from tkinter import IntVar
import tkinter.ttk
import time
import string
from tkinter.filedialog import askopenfile

from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re

import pandas as pd
from pandas import read_html
import html5lib


# ----------------------------------------------------------------
# excel
try:
    # os.system('TASKKILL /F /IM EXCEL.exe')
    os.remove("Report_review.xlsx")
    wb_report_review = Workbook()
    ws_report_review = wb_report_review.active

    ws_report_review.title = "DID"
    wb_report_review.create_sheet("Length Byte")
    wb_report_review.create_sheet("Hex Value")
    wb_report_review.create_sheet("ASCII Value")

    wb_report_review.save("Report_review.xlsx")


    os.close("Report_review.xlsx")
except:
    try:
        wb_report_review = Workbook()
        ws_report_review = wb_report_review.active

        ws_report_review.title = "DID"
        wb_report_review.create_sheet("Length Byte")
        wb_report_review.create_sheet("Hex Value")
        wb_report_review.create_sheet("ASCII Value")

        wb_report_review.save("Report_review.xlsx")
        # os.system('TASKKILL /F /IM EXCEL.exe')
    except OSError:
        print('Failed creating the file')
    else:
        print('File created')


# ------------------------------------------------------------------------------------------------------------------------

def create_DPT_can_load_file(direct_DPT,url):
    try:
        with open("DPT_can_load.xlsx", "r") as file:
            # Print the success message
            print("File is already haved")
        # fd = os.open("RFvalue.xlsx", os.O_RDWR)
        # os.close(fd)
    except OSError:
        direct_DPT = Input_DPT_path_text.get()
        if direct_DPT == '':
            messagebox.showerror("ERROR", "Locate file DPT fail, Please add locate of DPT file")
            print('tao thanh cong')
        else:
            wb_convert_DPT = load_workbook(str(direct_DPT))

        direct_report = Input_report_path_text.get()

        ws_convert_DPT = wb_convert_DPT.active

        ws_convert_DPT = wb_convert_DPT['7-1_DIDList']

        sheet_convert_DPT = wb_convert_DPT.worksheets[0]

        wb_convert_DPT.save("DPT_can_load.xlsx")
        wb_convert_DPT.close()
    return url

def DID_Compare(ws_report_review_DID_Compare, wb_report_review_DID_Compare,url):
    
    Sheet_DPT_Name = Input_Sheet_Name_text.get()

    Column_DID_DPT = Input_Column_DID_text.get()
    Column_DID = Column_DID_DPT[:-1]
    # print(Column_DID)
    # print(type(Column_DID))
    line_DID = Column_DID_DPT[-1:]

    # print(Column_DID + line_DID)
    

    Column_Length_Byte_DPT = Input_Column_Length_Byte_text.get()
    Column_Length_Byte = Column_Length_Byte_DPT[:-1]
    # print(Column_Length_Byte)
    # print(type(Column_Length_Byte))
    line_Length_Byte = Column_Length_Byte_DPT[-1:]

    Column_Value_DPT = Input_Column_Value_text.get()
    Column_Value = Column_Value_DPT[:-1]
    # print(Column_Value)
    # print(type(Column_Value))
    line_Value = Column_Value_DPT[-1:]
    


    wb_report_review_DID_Compare = load_workbook('Report_review.xlsx')
    wb_add_dpt_value_to_report_review = load_workbook('Report_review.xlsx')
    wb_read_DPT = load_workbook('DPT_can_load.xlsx')
    wb_report_review_DID_Compare_add_result = load_workbook('Report_review.xlsx')
    wb_report_review_DID_Compare_Clear = load_workbook('Report_review.xlsx')

    ws_report_review_DID_Compare = wb_report_review_DID_Compare.active
    ws_read_DPT = wb_read_DPT.active
    ws_report_review_DID_Compare_Clear = wb_report_review_DID_Compare.active
    ws_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review.active
    ws_report_review_DID_Compare_add_result = wb_report_review_DID_Compare.active

    ws_report_review_DID_Compare = wb_report_review_DID_Compare['DID']
    ws_report_review_DID_Compare_Clear = wb_report_review_DID_Compare['DID']
    ws_read_DPT = wb_read_DPT[Sheet_DPT_Name]
    ws_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review['DID']
    ws_report_review_DID_Compare_add_result = wb_report_review_DID_Compare_add_result['DID']
    
    count_sheet_check = 0
    x = 0
    while x < len(wb_read_DPT.sheetnames):
        if wb_read_DPT.sheetnames[x] == str(Sheet_DPT_Name):
            count_sheet_check = x
            print(count_sheet_check)
            x = len(wb_read_DPT.sheetnames)
        x += 1

    # print(wb_read_DPT.sheetnames)
    # print(len(wb_read_DPT.sheetnames))
    # print(wb_read_DPT.sheetnames[12])
    # x = 0
    # while x < len(wb_read_DPT.sheetnames)

    sheet_report_review_DID_Compare = wb_report_review_DID_Compare.worksheets[0]
    sheet_read_DPT = wb_read_DPT.worksheets[count_sheet_check]
    sheet_report_review_DID_Compare_Clear = wb_report_review_DID_Compare.worksheets[0]
    sheet_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review.worksheets[0]
    sheet_report_review_DID_Compare_add_result = wb_report_review_DID_Compare_add_result.worksheets[0]


    row_count_report_review = sheet_report_review_DID_Compare.max_row
    row_count_read_DPT = sheet_read_DPT.max_row
    row_count_report_review_Clear = sheet_report_review_DID_Compare_Clear.max_row    
    row_count_add_dpt_value_to_report_review = sheet_add_dpt_value_to_report_review.max_row

    # row_list_DID_In_DPT = ws_read_DPT[Column_DID + line_DID].value
    # print(row_list_DID_In_DPT)


    if row_count_report_review_Clear != '':
        b = 0
        while b < row_count_report_review_Clear:
            ws_report_review_DID_Compare.delete_rows(1)
            b += 1
            print("dang xoa")
        print("done")
        b = 0

    noneFill = PatternFill(start_color='00FFFFFF',
                            end_color='00FFFFFF',
                            fill_type='solid'
                            )
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    font_text = Font(name="Calibri", size=14, color='00FFFFFF', bold=True)
    font_text3 = Font(name="Calibri", size=11, color='00FFFFFF', bold=True)
    font_text2 = Font(name="Calibri", size=11, color='000000', bold=False, italic = True)
    alignment = Alignment(horizontal='center', vertical='center')

    sheet_report_review_DID_Compare.column_dimensions['F'].width = 20

    sheet_report_review_DID_Compare.column_dimensions['A'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_DID_Compare.column_dimensions['B'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_DID_Compare.column_dimensions['C'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_DID_Compare.column_dimensions['D'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_DID_Compare.column_dimensions['E'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_DID_Compare.column_dimensions['F'].number_format = numbers.FORMAT_TEXT

    sheet_report_review_DID_Compare.column_dimensions['A'].alignment = alignment
    sheet_report_review_DID_Compare.column_dimensions['B'].alignment = alignment
    sheet_report_review_DID_Compare.column_dimensions['C'].alignment = alignment
    sheet_report_review_DID_Compare.column_dimensions['D'].alignment = alignment
    sheet_report_review_DID_Compare.column_dimensions['E'].alignment = alignment
    sheet_report_review_DID_Compare.column_dimensions['F'].alignment = alignment


    k = 5
    o = 4
    j = 5
    count_length_byte = 0
    count_Value_DPT = 0
    count_DID_In_DPT = 0
    length_byte = []
    Value_DPT = []
    DID_In_DPT = []
    load_second = 0
    # print(row_read_DPT)
    
    while k < row_count_read_DPT:
        if str(Column_DID) == '':
            for row in range(o, j):
                for col in range(1, 2):
                    char = get_column_letter(col)
                    row_list_DID_In_DPT = ws_read_DPT[char + str(row)].value
                    row_list_DID_In_DPT = row_list_DID_In_DPT.lower()


                    # print(type(char))
                    # print(type(row))
                    # print(row_list_DID)
        else:
            row_list_DID_In_DPT = ws_read_DPT[Column_DID + str(line_DID)].value

            # print(row_list_DID_In_DPT)

        if str(Column_Length_Byte) == '':      
            for row in range(o, j):
                for col in range(3, 4):
                    char = get_column_letter(col)
                    # print(char)
                    row_list_length_byte = ws_read_DPT[char + str(row)].value
                    # print(row_list_length_byte)
        else:
            row_list_length_byte = ws_read_DPT[Column_Length_Byte + str(line_Length_Byte)].value
            # print(row_list_length_byte)
        
        if str(row_list_DID_In_DPT) != 'None' :
            DID_In_DPT = DID_In_DPT + [row_list_DID_In_DPT]
            count_DID_In_DPT +=  1

        if str(row_list_length_byte) != 'None':
            length_byte = length_byte + [row_list_length_byte]
            count_length_byte +=  1
        
        
        o += 1
        j += 1
        k += 1
        line_DID = int(line_DID) + 1
        line_Length_Byte = int(line_Length_Byte) + 1



    
    id = 1
    r = 2
    with open(url, 'r') as f:

        ws_report_review_DID_Compare.append(['ID' ,'RECEIVED','EXPECTED','DPT','RESULT'])
        
        for col in range(1, 7):
            cell_header = ws_report_review_DID_Compare.cell(1, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='000066CC', end_color='000066CC', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
            cell_header.alignment = alignment

        result_of_string = f.read()
        
        result_of_string = result_of_string.partition("EXPECTED: 62")[2]
        # print(len(result_of_string))
        while len(result_of_string) > 0:

            id += 1
            # print(len(result_of_string))
            percent.set(str((id//count_DID_In_DPT)*100)+"%")
            length_of_EXPECTED = result_of_string.find('<')
            # print(length_of_EXPECTED)
            

            result_of_EXPECTED = result_of_string
            result_of_EXPECTED = result_of_EXPECTED.split('<', 1)[0]

            value_of_EXPECTED = result_of_EXPECTED[4:length_of_EXPECTED]
            DID_of_EXPECTED = result_of_EXPECTED[0:4]
            DID_of_EXPECTED = DID_of_EXPECTED.upper()
            # print(value_of_EXPECTED)

            try:
                bytes_object_of_EXPECTED = bytes.fromhex(value_of_EXPECTED)
                
            except:
                missing_lenght_byte_of_EXPECTED = value_of_EXPECTED.partition(".{")[2]
                missing_lenght_byte_of_EXPECTED = missing_lenght_byte_of_EXPECTED[:-1]

                length_of_EXPECTED = int(length_of_EXPECTED) + int(missing_lenght_byte_of_EXPECTED) - 4

                value_of_EXPECTED = value_of_EXPECTED.split('.', 1)[0]

                bytes_object_of_EXPECTED = bytes.fromhex(value_of_EXPECTED)

                ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")

                # Convert to bytes object
            try:
                ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")

            # Convert to ASCII representation
            except:
                print('khong')
            if str(value_of_EXPECTED) != '':
                length_of_EXPECTED = (length_of_EXPECTED - 4) // 2
            else:
                length_of_EXPECTED = (length_of_EXPECTED // 2) - 2

            result_of_string = result_of_string.partition("RECEIVED: 62")[2] 
            # print(result_of_string)

            ### This is the part that measures the length before '<'
            length_of_RECEIVED = result_of_string.find('<')
            # print(length)
            

            result_of_RECEIVED = result_of_string
            result_of_RECEIVED = result_of_RECEIVED.split('<', 1)[0]

            value_of_RECEIVED = result_of_RECEIVED[4:length_of_RECEIVED]
            DID_of_RECEIVED = result_of_RECEIVED[0:4]
            DID_of_RECEIVED = DID_of_RECEIVED.upper()

            bytes_object_of_RECEIVED = bytes.fromhex(value_of_RECEIVED)
                # Convert to bytes object

            length_of_RECEIVED = (length_of_RECEIVED - 4) // 2

            try:
                ascii_string_of_RECEIVED = bytes_object_of_RECEIVED.decode("ASCII")
                # print(ascii_string_of_RECEIVED)
                ascii_string_of_RECEIVED = ILLEGAL_CHARACTERS_RE.sub(r'', ascii_string_of_RECEIVED)
                # print(ascii_string_of_RECEIVED)
                
            # Convert to ASCII representation
            except:
                ascii_string_of_RECEIVED = ILLEGAL_CHARACTERS_RE.sub(r'', ascii_string_of_RECEIVED)
                # print(ascii_string_of_RECEIVED)


            result_of_string = result_of_string.partition("EXPECTED: 62")[2]

            ws_report_review_DID_Compare.append(['ID_'+str(id) ,str(DID_of_RECEIVED),str(DID_of_EXPECTED),'' , '', '','',''])

            # try:
            #     ws_report_review_DID_Compare.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(length_of_RECEIVED),str(length_of_EXPECTED),'' , '', '','',''])
            # except:
            #     ws_report_review_DID_Compare.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(length_of_RECEIVED),str(length_of_EXPECTED),'' , '', '','',''])
            
            r += 1

    data_validation_data = '"PASS, FAIL"'
    
    for row in range(2, row_count_read_DPT):
        data_validation = DataValidation(type='list', formula1 = data_validation_data)
        ws_report_review_DID_Compare.add_data_validation(data_validation)
        data_validation.add(ws_report_review_DID_Compare['D'+str(row)])
    
    locate_save = Output_path_text.get()
    if locate_save == '':
            # print('Ok')
        wb_report_review_DID_Compare.save("Report_review.xlsx")
    else:
        wb_report_review_DID_Compare.save(locate_save + '/'+'Report_review.xlsx')
    
    wb_report_review_DID_Compare.close()

    l = 2
    h = 0
    
    wb_report_review_DID_Compare_add_dpt_value = load_workbook('Report_review.xlsx')
    ws_report_review_DID_Compare_add_dpt_value = wb_report_review_DID_Compare_add_dpt_value.active
    ws_report_review_DID_Compare_add_dpt_value = wb_report_review_DID_Compare_add_dpt_value['DID']

    sheet_report_review_add_dpt_value = wb_report_review_DID_Compare_add_dpt_value.worksheets[0]
    row_count_report_review = sheet_report_review_add_dpt_value.max_row

    while l < row_count_report_review:
        if len(DID_In_DPT[h]) < 7:
            if len(DID_In_DPT[h]) > 4:
                length_remove = len(DID_In_DPT[h]) - 4
                DID_In_DPT_new = DID_In_DPT[h][length_remove:]
                ws_report_review_DID_Compare_add_dpt_value['D'+ str(l)] = str(DID_In_DPT_new)
            else:
                ws_report_review_DID_Compare_add_dpt_value['D'+ str(l)] = str(DID_In_DPT[h])
            l += 1

        
        if len(DID_In_DPT[h]) < 7:
            if len(DID_In_DPT[h]) > 4:
                length_remove = len(DID_In_DPT[h]) - 4
                DID_In_DPT_new = DID_In_DPT[h][length_remove:]
                ws_report_review_DID_Compare_add_dpt_value['D'+ str(l)] = str(DID_In_DPT_new)
            else:
                ws_report_review_DID_Compare_add_dpt_value['D'+ str(l)] = str(DID_In_DPT[h])
            l += 1
            
        h += 1
        
        

    if locate_save == '':
        wb_report_review_DID_Compare_add_dpt_value.save("Report_review.xlsx")
    else:
        wb_report_review_DID_Compare_add_dpt_value.save(locate_save + '/'+'Report_review.xlsx')

    
    wb_report_review_DID_Compare_add_dpt_value.close()


    wb_report_review_DID_Compare_add_result = load_workbook('Report_review.xlsx')
    ws_report_review_DID_Compare_add_result = wb_report_review_DID_Compare_add_result.active
    ws_report_review_DID_Compare_add_result = wb_report_review_DID_Compare_add_result['DID']

    sheet_report_review_add_result = wb_report_review_DID_Compare_add_result.worksheets[0]
    row_count_report_review_add_result = sheet_report_review_add_result.max_row


    z = 2
    m = 2
    n = 3
    v = 2
    while z < row_count_report_review_add_result + 1:
        for row in range(m, n):
            for col in range(2, 3):
                char = get_column_letter(col)
                DID_of_RECEIVED_report = ws_report_review_DID_Compare_add_result[char + str(row)].value
                # print(DID_of_RECEIVED_report)
                

        for row in range(m, n):
            for col in range(3, 4):
                char = get_column_letter(col)
                DID_of_EXPECTED_report = ws_report_review_DID_Compare_add_result[char + str(row)].value
                # print(DID_of_EXPECTED)
            
        for row in range(m, n):
            for col in range(4, 5):
                char = get_column_letter(col)
                DID_of_DID_DPT_report = ws_report_review_DID_Compare_add_result[char + str(row)].value
                # DID_of_DID_DPT_report = DID_of_DID_DPT_report
                # row_list_DID_In_DPT_check = DID_of_DID_DPT_report
                # if len(row_list_DID_In_DPT_check) > 4 :
                #     length_check = len(row_list_DID_In_DPT_check) - 4
                #     row_list_DID_In_DPT = row_list_DID_In_DPT_check[-length_check:]
                # print(DID_of_DID_DPT_report)
        
        if str(DID_of_EXPECTED_report) == str(DID_of_RECEIVED_report) and str(DID_of_EXPECTED_report) == str(DID_of_DID_DPT_report):
            # print('pass DID ')
            result_DID_value = 'PASS'
        else:
            result_DID_value = 'FAIL'

        ws_report_review_DID_Compare_add_result['E'+ str(z)] = str(result_DID_value)

        if str(result_DID_value) == 'FAIL':
            for col in range(2, 6):
                cell_header_result_DID_value = ws_report_review_DID_Compare_add_result.cell(v, col)
            # used hex code for red color
                cell_header_result_DID_value.fill = PatternFill(
                    start_color='00FF0000', end_color='00FF0000', fill_type="solid")
                cell_header_result_DID_value.border = border
                cell_header_result_DID_value.font = font_text3
                cell_header_result_DID_value.alignment = alignment

        else:
            for col in range(5, 6):
                cell_header_result_DID_value = ws_report_review_DID_Compare_add_result.cell(v, col)
            # used hex code for red color
                cell_header_result_DID_value.fill = PatternFill(
                    start_color='0099CC00', end_color='0099CC00', fill_type="solid")
                cell_header_result_DID_value.border = border
                cell_header_result_DID_value.font = font_text3
                cell_header_result_DID_value.alignment = alignment

        n += 1
        m += 1
        v += 1
        z += 1

    if locate_save == '':
        wb_report_review_DID_Compare_add_result.save("Report_review.xlsx")
    else:
        wb_report_review_DID_Compare_add_result.save(locate_save + '/'+'Report_review.xlsx')

    
    wb_report_review_DID_Compare_add_result.close()

    # messagebox.showinfo("COMPLETE", "File Report_review.xlsx has been created in the same folder tool successfully")
    # print('tao thanh cong')


def Length_Byte(ws_report_review_Lenght_Byte, wb_report_review_Length_Byte,url):

    Sheet_DPT_Name = Input_Sheet_Name_text.get()

    Column_DID_DPT = Input_Column_DID_text.get()
    Column_DID = Column_DID_DPT[:-1]
    # print(Column_DID)
    # print(type(Column_DID))
    line_DID = Column_DID_DPT[-1:]

    # print(Column_DID + line_DID)
    

    Column_Length_Byte_DPT = Input_Column_Length_Byte_text.get()
    Column_Length_Byte = Column_Length_Byte_DPT[:-1]
    # print(Column_Length_Byte)
    # print(type(Column_Length_Byte))
    line_Length_Byte = Column_Length_Byte_DPT[-1:]

    Column_Value_DPT = Input_Column_Value_text.get()
    Column_Value = Column_Value_DPT[:-1]
    # print(Column_Value)
    # print(type(Column_Value))
    line_Value = Column_Value_DPT[-1:]
    


    wb_report_review_Length_Byte = load_workbook('Report_review.xlsx')
    wb_add_dpt_value_to_report_review = load_workbook('Report_review.xlsx')
    wb_read_DPT = load_workbook('DPT_can_load.xlsx')
    wb_report_review_Length_Byte_add_result = load_workbook('Report_review.xlsx')
    wb_report_review_Length_Byte_Clear = load_workbook('Report_review.xlsx')

    ws_report_review_Lenght_Byte = wb_report_review_Length_Byte.active
    ws_read_DPT = wb_read_DPT.active
    ws_report_review_Lenght_Byte_Clear = wb_report_review_Length_Byte.active
    ws_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review.active
    ws_report_review_Lenght_Byte_add_result = wb_report_review_Length_Byte.active

    ws_report_review_Lenght_Byte = wb_report_review_Length_Byte['Length Byte']
    ws_report_review_Lenght_Byte_Clear = wb_report_review_Length_Byte['Length Byte']
    ws_read_DPT = wb_read_DPT[Sheet_DPT_Name]
    ws_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review['Length Byte']
    ws_report_review_Lenght_Byte_add_result = wb_report_review_Length_Byte_add_result['Length Byte']
    
    count_sheet_check = 0
    x = 0
    while x < len(wb_read_DPT.sheetnames):
        if wb_read_DPT.sheetnames[x] == str(Sheet_DPT_Name):
            count_sheet_check = x
            print(count_sheet_check)
            x = len(wb_read_DPT.sheetnames)
        x += 1

    # print(wb_read_DPT.sheetnames)
    # print(len(wb_read_DPT.sheetnames))
    # print(wb_read_DPT.sheetnames[12])
    # x = 0
    # while x < len(wb_read_DPT.sheetnames)

    sheet_report_review_Lenght_Byte = wb_report_review_Length_Byte.worksheets[1]
    sheet_read_DPT = wb_read_DPT.worksheets[count_sheet_check]
    sheet_report_review_Lenght_Byte_Clear = wb_report_review_Length_Byte.worksheets[1]
    sheet_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review.worksheets[1]
    sheet_report_review_Lenght_Byte_add_result = wb_report_review_Length_Byte_add_result.worksheets[1]


    row_count_report_review = sheet_report_review_Lenght_Byte.max_row
    row_count_read_DPT = sheet_read_DPT.max_row
    row_count_report_review_Clear = sheet_report_review_Lenght_Byte_Clear.max_row    
    row_count_add_dpt_value_to_report_review = sheet_add_dpt_value_to_report_review.max_row

    # row_list_DID_In_DPT = ws_read_DPT[Column_DID + line_DID].value
    # print(row_list_DID_In_DPT)


    if row_count_report_review_Clear != '':
        b = 0
        while b < row_count_report_review_Clear:
            ws_report_review_Lenght_Byte.delete_rows(1)
            b += 1
            print("dang xoa")
        print("done")
        b = 0

    noneFill = PatternFill(start_color='00FFFFFF',
                            end_color='00FFFFFF',
                            fill_type='solid'
                            )
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    font_text = Font(name="Calibri", size=14, color='00FFFFFF', bold=True)
    font_text3 = Font(name="Calibri", size=11, color='00FFFFFF', bold=True)
    font_text2 = Font(name="Calibri", size=11, color='000000', bold=False, italic = True)
    alignment = Alignment(horizontal='center', vertical='center')

    sheet_report_review_Lenght_Byte.column_dimensions['F'].width = 20

    sheet_report_review_Lenght_Byte.column_dimensions['A'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_Lenght_Byte.column_dimensions['B'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_Lenght_Byte.column_dimensions['C'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_Lenght_Byte.column_dimensions['D'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_Lenght_Byte.column_dimensions['E'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_Lenght_Byte.column_dimensions['F'].number_format = numbers.FORMAT_TEXT

    sheet_report_review_Lenght_Byte.column_dimensions['A'].alignment = alignment
    sheet_report_review_Lenght_Byte.column_dimensions['B'].alignment = alignment
    sheet_report_review_Lenght_Byte.column_dimensions['C'].alignment = alignment
    sheet_report_review_Lenght_Byte.column_dimensions['D'].alignment = alignment
    sheet_report_review_Lenght_Byte.column_dimensions['E'].alignment = alignment
    sheet_report_review_Lenght_Byte.column_dimensions['F'].alignment = alignment


    k = 5
    o = 4
    j = 5
    count_length_byte = 0
    count_Value_DPT = 0
    count_DID_In_DPT = 0
    length_byte = []
    Value_DPT = []
    DID_In_DPT = []
    load_second = 0
    # print(row_read_DPT)
    
    while k < row_count_read_DPT:
        if str(Column_DID) == '':
            for row in range(o, j):
                for col in range(1, 2):
                    char = get_column_letter(col)
                    row_list_DID_In_DPT = ws_read_DPT[char + str(row)].value
                    # print(type(char))
                    # print(type(row))
                    # print(row_list_DID)
        else:
            row_list_DID_In_DPT = ws_read_DPT[Column_DID + str(line_DID)].value
            print(row_list_DID_In_DPT)

        if str(Column_Length_Byte) == '':      
            for row in range(o, j):
                for col in range(3, 4):
                    char = get_column_letter(col)
                    # print(char)
                    row_list_length_byte = ws_read_DPT[char + str(row)].value
                    # print(row_list_length_byte)
        else:
            row_list_length_byte = ws_read_DPT[Column_Length_Byte + str(line_Length_Byte)].value
            print(row_list_length_byte)
        
        if str(row_list_DID_In_DPT) != 'None':
            DID_In_DPT = DID_In_DPT + [row_list_DID_In_DPT]
            count_DID_In_DPT +=  1

        if str(row_list_length_byte) != 'None':
            length_byte = length_byte + [row_list_length_byte]
            count_length_byte +=  1
        
        
        o += 1
        j += 1
        k += 1
        line_DID = int(line_DID) + 1
        line_Length_Byte = int(line_Length_Byte) + 1



    
    id = 1
    r = 2
    with open(url, 'r') as f:

        ws_report_review_Lenght_Byte.append(['ID' , 'DID','RECEIVED','EXPECTED','DPT','RESULT'])
        
        for col in range(1, 7):
            cell_header = ws_report_review_Lenght_Byte.cell(1, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='000066CC', end_color='000066CC', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
            cell_header.alignment = alignment

        result_of_string = f.read()
        
        result_of_string = result_of_string.partition("EXPECTED: 62")[2]
        # print(len(result_of_string))
        while len(result_of_string) > 0:

            id += 1
            # print(len(result_of_string))
            percent.set(str((id//count_DID_In_DPT)*100)+"%")
            length_of_EXPECTED = result_of_string.find('<')
            # print(length_of_EXPECTED)
            

            result_of_EXPECTED = result_of_string
            result_of_EXPECTED = result_of_EXPECTED.split('<', 1)[0]

            value_of_EXPECTED = result_of_EXPECTED[4:length_of_EXPECTED]
            DID_of_EXPECTED = result_of_EXPECTED[0:4]

            # print(value_of_EXPECTED)

            try:
                bytes_object_of_EXPECTED = bytes.fromhex(value_of_EXPECTED)
                
            except:
                missing_lenght_byte_of_EXPECTED = value_of_EXPECTED.partition(".{")[2]
                missing_lenght_byte_of_EXPECTED = missing_lenght_byte_of_EXPECTED[:-1]

                length_of_EXPECTED = int(length_of_EXPECTED) + int(missing_lenght_byte_of_EXPECTED) - 4

                value_of_EXPECTED = value_of_EXPECTED.split('.', 1)[0]

                bytes_object_of_EXPECTED = bytes.fromhex(value_of_EXPECTED)

                ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")

                # Convert to bytes object
            try:
                ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")

            # Convert to ASCII representation
            except:
                print('khong')
            if str(value_of_EXPECTED) != '':
                length_of_EXPECTED = (length_of_EXPECTED - 4) // 2
            else:
                length_of_EXPECTED = (length_of_EXPECTED // 2) - 2

            result_of_string = result_of_string.partition("RECEIVED: 62")[2] 
            # print(result_of_string)

            ### This is the part that measures the length before '<'
            length_of_RECEIVED = result_of_string.find('<')
            # print(length)
            

            result_of_RECEIVED = result_of_string
            result_of_RECEIVED = result_of_RECEIVED.split('<', 1)[0]

            value_of_RECEIVED = result_of_RECEIVED[4:length_of_RECEIVED]
            DID_of_RECEIVED = result_of_RECEIVED[0:4]

            bytes_object_of_RECEIVED = bytes.fromhex(value_of_RECEIVED)
                # Convert to bytes object

            length_of_RECEIVED = (length_of_RECEIVED - 4) // 2

            try:
                ascii_string_of_RECEIVED = bytes_object_of_RECEIVED.decode("ASCII")
                # print(ascii_string_of_RECEIVED)
                ascii_string_of_RECEIVED = ILLEGAL_CHARACTERS_RE.sub(r'', ascii_string_of_RECEIVED)
                # print(ascii_string_of_RECEIVED)
                
            # Convert to ASCII representation
            except:
                ascii_string_of_RECEIVED = ILLEGAL_CHARACTERS_RE.sub(r'', ascii_string_of_RECEIVED)
                # print(ascii_string_of_RECEIVED)


            result_of_string = result_of_string.partition("EXPECTED: 62")[2]

            ws_report_review_Lenght_Byte.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(length_of_RECEIVED),str(length_of_EXPECTED),'' , '', '','',''])

            # try:
            #     ws_report_review_Lenght_Byte.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(length_of_RECEIVED),str(length_of_EXPECTED),'' , '', '','',''])
            # except:
            #     ws_report_review_Lenght_Byte.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(length_of_RECEIVED),str(length_of_EXPECTED),'' , '', '','',''])
            
            r += 1

    data_validation_data = '"PASS, FAIL"'
    
    for row in range(2, row_count_read_DPT):
        data_validation = DataValidation(type='list', formula1 = data_validation_data)
        ws_report_review_Lenght_Byte.add_data_validation(data_validation)
        data_validation.add(ws_report_review_Lenght_Byte['K'+str(row)])
    
    locate_save = Output_path_text.get()
    if locate_save == '':
            # print('Ok')
        wb_report_review_Length_Byte.save("Report_review.xlsx")
    else:
        wb_report_review_Length_Byte.save(locate_save + '/'+'Report_review.xlsx')
    
    wb_report_review_Length_Byte.close()

    l = 2
    h = 0
    
    wb_report_review_Length_Byte_add_dpt_value = load_workbook('Report_review.xlsx')
    ws_report_review_Lenght_Byte_add_dpt_value = wb_report_review_Length_Byte_add_dpt_value.active
    ws_report_review_Lenght_Byte_add_dpt_value = wb_report_review_Length_Byte_add_dpt_value['Length Byte']

    sheet_report_review_add_dpt_value = wb_report_review_Length_Byte_add_dpt_value.worksheets[1]
    row_count_report_review = sheet_report_review_add_dpt_value.max_row

    while l < row_count_report_review:

        ws_report_review_Lenght_Byte_add_dpt_value['E'+ str(l)] = str(length_byte[h])

        l += 1

        ws_report_review_Lenght_Byte_add_dpt_value['E'+ str(l)] = str(length_byte[h])
        
        l += 1
        h += 1

    if locate_save == '':
        wb_report_review_Length_Byte_add_dpt_value.save("Report_review.xlsx")
    else:
        wb_report_review_Length_Byte_add_dpt_value.save(locate_save + '/'+'Report_review.xlsx')

    
    wb_report_review_Length_Byte_add_dpt_value.close()


    wb_report_review_Length_Byte_add_result = load_workbook('Report_review.xlsx')
    ws_report_review_Lenght_Byte_add_result = wb_report_review_Length_Byte_add_result.active
    ws_report_review_Lenght_Byte_add_result = wb_report_review_Length_Byte_add_result['Length Byte']

    sheet_report_review_add_result = wb_report_review_Length_Byte_add_result.worksheets[1]
    row_count_report_review_add_result = sheet_report_review_add_result.max_row


    z = 2
    m = 2
    n = 3
    v = 2
    while z < row_count_report_review_add_result + 1:
        for row in range(m, n):
            for col in range(3, 4):
                char = get_column_letter(col)
                length_of_RECEIVED_report = ws_report_review_Lenght_Byte_add_result[char + str(row)].value
                # print(length_of_RECEIVED_report)
                

        for row in range(m, n):
            for col in range(4, 5):
                char = get_column_letter(col)
                length_of_EXPECTED_report = ws_report_review_Lenght_Byte_add_result[char + str(row)].value
                # print(length_of_EXPECTED)
            
        for row in range(m, n):
            for col in range(5, 6):
                char = get_column_letter(col)
                length_of_DID_DPT_report = ws_report_review_Lenght_Byte_add_result[char + str(row)].value
                # print(length_of_DID_DPT_report)
        
        if str(length_of_EXPECTED_report) == str(length_of_RECEIVED_report) and str(length_of_EXPECTED_report) == str(length_of_DID_DPT_report):
            # print('pass length ')
            result_length_value = 'PASS'
        else:
            result_length_value = 'FAIL'

        ws_report_review_Lenght_Byte_add_result['F'+ str(z)] = str(result_length_value)

        if str(result_length_value) == 'FAIL':
            for col in range(2, 7):
                cell_header_result_length_value = ws_report_review_Lenght_Byte_add_result.cell(v, col)
            # used hex code for red color
                cell_header_result_length_value.fill = PatternFill(
                    start_color='00FF0000', end_color='00FF0000', fill_type="solid")
                cell_header_result_length_value.border = border
                cell_header_result_length_value.font = font_text3
                cell_header_result_length_value.alignment = alignment

        else:
            for col in range(6, 7):
                cell_header_result_length_value = ws_report_review_Lenght_Byte_add_result.cell(v, col)
            # used hex code for red color
                cell_header_result_length_value.fill = PatternFill(
                    start_color='0099CC00', end_color='0099CC00', fill_type="solid")
                cell_header_result_length_value.border = border
                cell_header_result_length_value.font = font_text3
                cell_header_result_length_value.alignment = alignment

        n += 1
        m += 1
        v += 1
        z += 1

    if locate_save == '':
        wb_report_review_Length_Byte_add_result.save("Report_review.xlsx")
    else:
        wb_report_review_Length_Byte_add_result.save(locate_save + '/'+'Report_review.xlsx')

    
    wb_report_review_Length_Byte_add_result.close()

    # messagebox.showinfo("COMPLETE", "File Report_review.xlsx has been created in the same folder tool successfully")
    # print('tao thanh cong')

def Hex_value(ws_report_review_Hex_value, wb_report_review_Hex_value,url):

    Sheet_DPT_Name = Input_Sheet_Name_text.get()

    Column_DID_DPT = Input_Column_DID_text.get()
    Column_DID = Column_DID_DPT[:-1]
    # print(Column_DID)
    # print(type(Column_DID))
    line_DID = Column_DID_DPT[-1:]

    # print(Column_DID + line_DID)
    

    Column_Length_Byte_DPT = Input_Column_Length_Byte_text.get()
    Column_Length_Byte = Column_Length_Byte_DPT[:-1]
    # print(Column_Length_Byte)
    # print(type(Column_Length_Byte))
    line_Length_Byte = Column_Length_Byte_DPT[-1:]

    Column_Value_DPT = Input_Column_Value_text.get()
    Column_Value = Column_Value_DPT[:-1]
    # print(Column_Value)
    # print(type(Column_Value))
    line_Value = Column_Value_DPT[-1:]
    
    wb_report_review_Hex_value = load_workbook('Report_review.xlsx')
    wb_add_dpt_value_to_report_review = load_workbook('Report_review.xlsx')
    wb_read_DPT = load_workbook('DPT_can_load.xlsx')
    wb_report_review_Hex_value_add_result = load_workbook('Report_review.xlsx')
    wb_report_review_Hex_value_Clear = load_workbook('Report_review.xlsx')

    ws_report_review_Hex_value = wb_report_review_Hex_value.active
    ws_read_DPT = wb_read_DPT.active
    ws_report_review_Hex_value_Clear = wb_report_review_Hex_value.active
    ws_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review.active
    ws_report_review_Hex_value_add_result = wb_report_review_Hex_value.active

    ws_report_review_Hex_value = wb_report_review_Hex_value['Hex Value']
    ws_report_review_Hex_value_Clear = wb_report_review_Hex_value['Hex Value']
    ws_read_DPT = wb_read_DPT[str(Sheet_DPT_Name)]
    ws_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review['Hex Value']
    ws_report_review_Hex_value_add_result = wb_report_review_Hex_value_add_result['Hex Value']

    count_sheet_check = 0
    x = 0
    while x < len(wb_read_DPT.sheetnames):
        if wb_read_DPT.sheetnames[x] == str(Sheet_DPT_Name):
            count_sheet_check = x
            print(count_sheet_check)
            x = len(wb_read_DPT.sheetnames)
        x += 1

    sheet_report_review_Hex_value = wb_report_review_Hex_value.worksheets[2]
    sheet_read_DPT = wb_read_DPT.worksheets[count_sheet_check]
    sheet_report_review_Hex_value_Clear = wb_report_review_Hex_value.worksheets[2]
    sheet_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review.worksheets[2]
    sheet_report_review_Hex_value_add_result = wb_report_review_Hex_value_add_result.worksheets[2]


    row_count_report_review = sheet_report_review_Hex_value.max_row
    row_count_read_DPT = sheet_read_DPT.max_row
    row_count_report_review_Clear = sheet_report_review_Hex_value_Clear.max_row    
    row_count_add_dpt_value_to_report_review = sheet_add_dpt_value_to_report_review.max_row

    if row_count_report_review_Clear != '':
        b = 0
        while b < row_count_report_review_Clear:
            ws_report_review_Hex_value.delete_rows(1)
            b += 1
            print("dang xoa")
        print("done")
        b = 0

    noneFill = PatternFill(start_color='00FFFFFF',
                            end_color='00FFFFFF',
                            fill_type='solid'
                            )
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    font_text = Font(name="Calibri", size=14, color='00FFFFFF', bold=True)
    font_text3 = Font(name="Calibri", size=11, color='00FFFFFF', bold=True)
    font_text2 = Font(name="Calibri", size=11, color='000000', bold=False, italic = True)
    alignment = Alignment(horizontal='center', vertical='center')

    sheet_report_review_Hex_value.column_dimensions['B'].width = 40
    sheet_report_review_Hex_value.column_dimensions['C'].width = 40
    sheet_report_review_Hex_value.column_dimensions['D'].width = 40
    sheet_report_review_Hex_value.column_dimensions['F'].width = 20

    sheet_report_review_Hex_value.column_dimensions['A'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_Hex_value.column_dimensions['B'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_Hex_value.column_dimensions['C'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_Hex_value.column_dimensions['D'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_Hex_value.column_dimensions['E'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_Hex_value.column_dimensions['F'].number_format = numbers.FORMAT_TEXT

    sheet_report_review_Hex_value.column_dimensions['A'].alignment = alignment
    sheet_report_review_Hex_value.column_dimensions['B'].alignment = alignment
    sheet_report_review_Hex_value.column_dimensions['C'].alignment = alignment
    sheet_report_review_Hex_value.column_dimensions['D'].alignment = alignment
    sheet_report_review_Hex_value.column_dimensions['E'].alignment = alignment
    sheet_report_review_Hex_value.column_dimensions['F'].alignment = alignment

    k = 5
    o = 4
    j = 5
    count_length_byte = 0
    count_Value_DPT = 0
    count_DID_In_DPT = 0
    length_byte = []
    Value_DPT = []
    DID_In_DPT = []
    load_second = 0
    # print(row_read_DPT)
    row_list_hex_value = ''
    while k < row_count_read_DPT:
        row_list_hex_value = ''
        if str(Column_DID) == '':
            for row in range(o, j):
                for col in range(1, 2):
                    char = get_column_letter(col)
                    row_list_DID_In_DPT = ws_read_DPT[char + str(row)].value
                    # print(row_list_DID)
        else:
            row_list_DID_In_DPT = ws_read_DPT[Column_DID + str(line_DID)].value

        if str(Column_Value) == '':
            for row in range(o, j):
                for col in range(21, 22):
                    char = get_column_letter(col)
                    row_list_value = ws_read_DPT[char + str(row)].value
                    if str(row_list_value) != 'None':
                        for i in str(row_list_value):
                            row_list_hex_value += hex(ord(i))[2:]
                        # print(row_list_hex_value)
                        # print(row_list_value)
        else:
            row_list_value = ws_read_DPT[Column_Value + str(line_Value)].value
            if str(row_list_value) != 'None':
                for i in str(row_list_value):
                    row_list_hex_value += hex(ord(i))[2:]
                # print(row_list_hex_value)
                # print(row_list_value)


        if str(row_list_DID_In_DPT) != 'None':
            DID_In_DPT = DID_In_DPT + [row_list_DID_In_DPT]
            count_DID_In_DPT +=  1

        
        if str(row_list_value) != 'None':
            Value_DPT = Value_DPT + [row_list_hex_value]
            count_Value_DPT +=  1
        
        o += 1
        j += 1
        k += 1
        line_DID = int(line_DID) + 1
        line_Value = int(line_Value) + 1


    
    id = 1
    r = 2
    with open(url, 'r') as f:

        ws_report_review_Hex_value.append(['ID' , 'DID','RECEIVED','EXPECTED', 'DPT', 'RESULT'])
        
        for col in range(1, 7):
            cell_header = ws_report_review_Hex_value.cell(1, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='000066CC', end_color='000066CC', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
            cell_header.alignment = alignment

        result_of_string = f.read()
        
        result_of_string = result_of_string.partition("EXPECTED: 62")[2]
        # print(len(result_of_string))
        while len(result_of_string) > 0:

            id += 1
            # print(len(result_of_string))
            percent.set(str((id//count_DID_In_DPT)*100)+"%")
            length_of_EXPECTED = result_of_string.find('<')
            # print(length_of_EXPECTED)
            length_of_EXPECTED_real = int(length_of_EXPECTED) - 4
            # print(length_of_EXPECTED_real)
            

            result_of_EXPECTED = result_of_string
            result_of_EXPECTED = result_of_EXPECTED.split('<', 1)[0]

            value_of_EXPECTED = result_of_EXPECTED[4:length_of_EXPECTED]
            DID_of_EXPECTED = result_of_EXPECTED[0:4]

            # print(value_of_EXPECTED)
            missing_lenght_byte_of_EXPECTED = ''
            try:
                bytes_object_of_EXPECTED = bytes.fromhex(value_of_EXPECTED)
                
            except:
                
                missing_lenght_byte_of_EXPECTED = value_of_EXPECTED.partition(".{")[2]
                missing_lenght_byte_of_EXPECTED = missing_lenght_byte_of_EXPECTED[:-1]

                length_of_EXPECTED = int(length_of_EXPECTED) + int(missing_lenght_byte_of_EXPECTED) - 4

                value_of_EXPECTED = value_of_EXPECTED.split('.', 1)[0]

                bytes_object_of_EXPECTED = bytes.fromhex(value_of_EXPECTED)

                ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")

                # Convert to bytes object
            try:
                ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")

            # Convert to ASCII representation
            except:
                print('khong')
            if str(value_of_EXPECTED) != '':
                length_of_EXPECTED = (length_of_EXPECTED - 4) // 2
            else:
                length_of_EXPECTED = (length_of_EXPECTED // 2) - 2

            result_of_string = result_of_string.partition("RECEIVED: 62")[2] 
            # print(result_of_string)

            ### This is the part that measures the length before '<'
            length_of_RECEIVED = result_of_string.find('<')
            length_of_RECEIVED_real = int(length_of_RECEIVED) - 4
            # print(length_of_RECEIVED_real)
            

            result_of_RECEIVED = result_of_string
            result_of_RECEIVED = result_of_RECEIVED.split('<', 1)[0]

            value_of_RECEIVED = result_of_RECEIVED[4:length_of_RECEIVED]
            DID_of_RECEIVED = result_of_RECEIVED[0:4]

            if str(missing_lenght_byte_of_EXPECTED) != '': 
                # length_minus = int(length_of_RECEIVED_real) - int(length_of_EXPECTED_real)
                # print(str(missing_lenght_byte_of_EXPECTED))
                value_of_RECEIVED = value_of_RECEIVED[:-int(missing_lenght_byte_of_EXPECTED)]

            bytes_object_of_RECEIVED = bytes.fromhex(value_of_RECEIVED)
                # Convert to bytes object

            length_of_RECEIVED = (length_of_RECEIVED - 4) // 2

            try:
                ascii_string_of_RECEIVED = bytes_object_of_RECEIVED.decode("ASCII")
                # print(ascii_string_of_RECEIVED)
                ascii_string_of_RECEIVED = ILLEGAL_CHARACTERS_RE.sub(r'', ascii_string_of_RECEIVED)
                # print(ascii_string_of_RECEIVED)
                
            # Convert to ASCII representation
            except:
                ascii_string_of_RECEIVED = ILLEGAL_CHARACTERS_RE.sub(r'', ascii_string_of_RECEIVED)
                # print(ascii_string_of_RECEIVED)


            result_of_string = result_of_string.partition("EXPECTED: 62")[2]

            ws_report_review_Hex_value.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(value_of_RECEIVED),str(value_of_EXPECTED),'' , '', '','',''])

            # try:
            #     ws_report_review_Hex_value.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(value_of_RECEIVED),str(value_of_EXPECTED),'' , '', '','',''])
            # except:
            #     ws_report_review_Hex_value.append(['ID_'+str(id) , str(DID_of_EXPECTED), str(value_of_RECEIVED),str(value_of_EXPECTED),'' , '', '','',''])
            
            r += 1

    data_validation_data = '"PASS, FAIL"'
    
    for row in range(2, row_count_read_DPT):
        data_validation = DataValidation(type='list', formula1 = data_validation_data)
        ws_report_review_Hex_value.add_data_validation(data_validation)
        data_validation.add(ws_report_review_Hex_value['K'+str(row)])
    
    locate_save = Output_path_text.get()
    if locate_save == '':
            # print('Ok')
        wb_report_review_Hex_value.save("Report_review.xlsx")
    else:
        wb_report_review_Hex_value.save(locate_save + '/'+'Report_review.xlsx')
    
    wb_report_review_Hex_value.close()

    l = 2
    h = 0
    
    wb_report_review_Hex_value_add_dpt_value = load_workbook('Report_review.xlsx')
    ws_report_review_Hex_value_add_dpt_value = wb_report_review_Hex_value_add_dpt_value.active
    ws_report_review_Hex_value_add_dpt_value = wb_report_review_Hex_value_add_dpt_value['Hex Value']

    sheet_report_review_add_dpt_value = wb_report_review_Hex_value_add_dpt_value.worksheets[2]
    row_count_report_review = sheet_report_review_add_dpt_value.max_row

    while l < row_count_report_review:
        ws_report_review_Hex_value_add_dpt_value['E'+ str(l)] = str(Value_DPT[h])

        l += 1

        ws_report_review_Hex_value_add_dpt_value['E'+ str(l)] = str(Value_DPT[h])
        
        l += 1
        h += 1

    if locate_save == '':
        wb_report_review_Hex_value_add_dpt_value.save("Report_review.xlsx")
    else:
        wb_report_review_Hex_value_add_dpt_value.save(locate_save + '/'+'Report_review.xlsx')

    
    wb_report_review_Hex_value_add_dpt_value.close()


    wb_report_review_Hex_value_add_result = load_workbook('Report_review.xlsx')
    ws_report_review_Hex_value_add_result = wb_report_review_Hex_value_add_result.active
    ws_report_review_Hex_value_add_result = wb_report_review_Hex_value_add_result['Hex Value']

    sheet_report_review_add_result = wb_report_review_Hex_value_add_result.worksheets[2]
    row_count_report_review_add_result = sheet_report_review_add_result.max_row

    wb_report_review_DID_Compare_add_result = load_workbook('Report_review.xlsx')
    ws_report_review_DID_Compare_add_result = wb_report_review_DID_Compare_add_result.active
    ws_report_review_DID_Compare_add_result = wb_report_review_DID_Compare_add_result['DID']

    sheet_report_review_add_result = wb_report_review_DID_Compare_add_result.worksheets[0]
    row_count_report_review_add_result = sheet_report_review_add_result.max_row

    wb_report_review_Length_Byte_add_result = load_workbook('Report_review.xlsx')
    ws_report_review_Lenght_Byte_add_result = wb_report_review_Length_Byte_add_result.active
    ws_report_review_Lenght_Byte_add_result = wb_report_review_Length_Byte_add_result['Length Byte']

    sheet_report_review_add_result = wb_report_review_Length_Byte_add_result.worksheets[1]
    row_count_report_review_add_result = sheet_report_review_add_result.max_row
    

    z = 2
    m = 2
    n = 3
    v = 2
    while z < row_count_report_review_add_result + 1:

        for row in range(m, n):
            for col in range(3, 4):
                char = get_column_letter(col)
                value_of_RECEIVED_report = ws_report_review_Hex_value_add_result[char + str(row)].value
                # print(value_of_RECEIVED)
        
        for row in range(m, n):
            for col in range(4, 5):
                char = get_column_letter(col)
                value_of_EXPECTED_report = ws_report_review_Hex_value_add_result[char + str(row)].value
                # print(value_of_EXPECTED)

        for row in range(m, n):
            for col in range(5, 6):
                char = get_column_letter(col)
                value_of_DPT_report = ws_report_review_Hex_value_add_result[char + str(row)].value
                # print(value_of_EXPECTED)

        for row in range(m, n):
            for col in range(2, 3):
                char = get_column_letter(col)
                DID_of_RECEIVED_report = ws_report_review_DID_Compare_add_result[char + str(row)].value
                # print(DID_of_RECEIVED_report)
                

        for row in range(m, n):
            for col in range(3, 4):
                char = get_column_letter(col)
                DID_of_EXPECTED_report = ws_report_review_DID_Compare_add_result[char + str(row)].value
                # print(DID_of_EXPECTED)
            
        for row in range(m, n):
            for col in range(4, 5):
                char = get_column_letter(col)
                DID_of_DID_DPT_report = ws_report_review_DID_Compare_add_result[char + str(row)].value

        for row in range(m, n):
            for col in range(3, 4):
                char = get_column_letter(col)
                length_of_RECEIVED_report = ws_report_review_Lenght_Byte_add_result[char + str(row)].value
                # print(length_of_RECEIVED_report)
                

        for row in range(m, n):
            for col in range(4, 5):
                char = get_column_letter(col)
                length_of_EXPECTED_report = ws_report_review_Lenght_Byte_add_result[char + str(row)].value
                # print(length_of_EXPECTED)
            
        for row in range(m, n):
            for col in range(5, 6):
                char = get_column_letter(col)
                length_of_DID_DPT_report = ws_report_review_Lenght_Byte_add_result[char + str(row)].value
                # print(length_of_DID_DPT_report)

        
        
        if str(value_of_EXPECTED_report) == str(value_of_RECEIVED_report) and str(value_of_EXPECTED_report) == str(value_of_DPT_report):
            # print('pass hex value')
            result_hex_value = 'PASS'
        else:
            if str(value_of_EXPECTED_report) == str(value_of_RECEIVED_report):
                if str(length_of_EXPECTED_report) == str(length_of_RECEIVED_report) and str(length_of_EXPECTED_report) == str(length_of_DID_DPT_report):
                    if str(DID_of_EXPECTED_report) == str(DID_of_RECEIVED_report) and str(DID_of_EXPECTED_report) == str(DID_of_DID_DPT_report):
                        result_hex_value = 'PASS'
                    else:
                        result_hex_value = 'FAIL'
                else:
                    result_hex_value = 'FAIL'
            else:
                result_hex_value = 'FAIL'

        # if str(value_of_EXPECTED_report) == str(value_of_RECEIVED_report):
        #     # print('pass hex value')
        #     result_hex_value = 'PASS'
        # else:
        #     result_hex_value = 'FAIL'

        ws_report_review_Hex_value_add_result['F'+ str(z)] = str(result_hex_value)

        if  str(result_hex_value) == 'FAIL':
            for col in range(2, 7):
                cell_header_result_hex_value = ws_report_review_Hex_value_add_result.cell(v, col)
            # used hex code for red color
                cell_header_result_hex_value.fill = PatternFill(
                    start_color='00FF0000', end_color='00FF0000', fill_type="solid")
                cell_header_result_hex_value.border = border
                cell_header_result_hex_value.font = font_text3
                cell_header_result_hex_value.alignment = alignment

        else:
            for col in range(6, 7):
                cell_header_result_hex_value = ws_report_review_Hex_value_add_result.cell(v, col)
            # used hex code for red color
                cell_header_result_hex_value.fill = PatternFill(
                    start_color='0099CC00', end_color='0099CC00', fill_type="solid")
                cell_header_result_hex_value.border = border
                cell_header_result_hex_value.font = font_text3
                cell_header_result_hex_value.alignment = alignment

        n += 1
        m += 1
        v += 1
        z += 1

    if locate_save == '':
        wb_report_review_Hex_value_add_result.save("Report_review.xlsx")
    else:
        wb_report_review_Hex_value_add_result.save(locate_save + '/'+'Report_review.xlsx')

    
    wb_report_review_Hex_value_add_result.close()

    # messagebox.showinfo("COMPLETE", "File Report_review.xlsx has been created in the same folder tool successfully")
    # print('tao thanh cong')

def ASCII_Value(ws_report_review_ASCII_value, wb_report_review_ASCII_value,url):

    Sheet_DPT_Name = Input_Sheet_Name_text.get()

    Column_DID_DPT = Input_Column_DID_text.get()
    Column_DID = Column_DID_DPT[:-1]
    # print(Column_DID)
    # print(type(Column_DID))
    line_DID = Column_DID_DPT[-1:]

    # print(Column_DID + line_DID)
    

    Column_Length_Byte_DPT = Input_Column_Length_Byte_text.get()
    Column_Length_Byte = Column_Length_Byte_DPT[:-1]
    # print(Column_Length_Byte)
    # print(type(Column_Length_Byte))
    line_Length_Byte = Column_Length_Byte_DPT[-1:]

    Column_Value_DPT = Input_Column_Value_text.get()
    Column_Value = Column_Value_DPT[:-1]
    # print(Column_Value)
    # print(type(Column_Value))
    line_Value = Column_Value_DPT[-1:]
    
    wb_report_review_ASCII_value = load_workbook('Report_review.xlsx')
    wb_add_dpt_value_to_report_review = load_workbook('Report_review.xlsx')
    wb_read_DPT = load_workbook('DPT_can_load.xlsx')
    wb_report_review_ASCII_value_add_result = load_workbook('Report_review.xlsx')
    wb_report_review_ASCII_value_Clear = load_workbook('Report_review.xlsx')

    ws_report_review_ASCII_value = wb_report_review_ASCII_value.active
    ws_read_DPT = wb_read_DPT.active
    ws_report_review_ASCII_value_Clear = wb_report_review_ASCII_value.active
    ws_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review.active
    ws_report_review_ASCII_value_add_result = wb_report_review_ASCII_value.active

    ws_report_review_ASCII_value = wb_report_review_ASCII_value['ASCII Value']
    ws_report_review_ASCII_value_Clear = wb_report_review_ASCII_value['ASCII Value']
    ws_read_DPT = wb_read_DPT[str(Sheet_DPT_Name)]
    ws_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review['ASCII Value']
    ws_report_review_ASCII_value_add_result = wb_report_review_ASCII_value_add_result['ASCII Value']

    count_sheet_check = 0
    x = 0
    while x < len(wb_read_DPT.sheetnames):
        if wb_read_DPT.sheetnames[x] == str(Sheet_DPT_Name):
            count_sheet_check = x
            print(count_sheet_check)
            x = len(wb_read_DPT.sheetnames)
        x += 1

    sheet_report_review_ASCII_value = wb_report_review_ASCII_value.worksheets[3]
    sheet_read_DPT = wb_read_DPT.worksheets[count_sheet_check]
    sheet_report_review_ASCII_value_Clear = wb_report_review_ASCII_value.worksheets[3]
    sheet_add_dpt_value_to_report_review = wb_add_dpt_value_to_report_review.worksheets[3]
    sheet_report_review_ASCII_value_add_result = wb_report_review_ASCII_value_add_result.worksheets[3]


    row_count_report_review = sheet_report_review_ASCII_value.max_row
    row_count_read_DPT = sheet_read_DPT.max_row
    row_count_report_review_Clear = sheet_report_review_ASCII_value_Clear.max_row    
    row_count_add_dpt_value_to_report_review = sheet_add_dpt_value_to_report_review.max_row

    if row_count_report_review_Clear != '':
        b = 0
        while b < row_count_report_review_Clear:
            ws_report_review_ASCII_value.delete_rows(1)
            b += 1
            print("dang xoa")
        print("done")
        b = 0

    noneFill = PatternFill(start_color='00FFFFFF',
                            end_color='00FFFFFF',
                            fill_type='solid'
                            )
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    font_text = Font(name="Calibri", size=14, color='00FFFFFF', bold=True)
    font_text3 = Font(name="Calibri", size=11, color='00FFFFFF', bold=True)
    font_text2 = Font(name="Calibri", size=11, color='000000', bold=False, italic = True)
    alignment = Alignment(horizontal='center', vertical='center')

    
    sheet_report_review_ASCII_value.column_dimensions['C'].width = 30
    sheet_report_review_ASCII_value.column_dimensions['D'].width = 30
    sheet_report_review_ASCII_value.column_dimensions['E'].width = 30
    sheet_report_review_ASCII_value.column_dimensions['F'].width = 20
    

    sheet_report_review_ASCII_value.column_dimensions['A'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_ASCII_value.column_dimensions['B'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_ASCII_value.column_dimensions['C'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_ASCII_value.column_dimensions['D'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_ASCII_value.column_dimensions['E'].number_format = numbers.FORMAT_TEXT
    sheet_report_review_ASCII_value.column_dimensions['F'].number_format = numbers.FORMAT_TEXT

    sheet_report_review_ASCII_value.column_dimensions['A'].alignment = alignment
    sheet_report_review_ASCII_value.column_dimensions['B'].alignment = alignment
    sheet_report_review_ASCII_value.column_dimensions['C'].alignment = alignment
    sheet_report_review_ASCII_value.column_dimensions['D'].alignment = alignment
    sheet_report_review_ASCII_value.column_dimensions['E'].alignment = alignment
    sheet_report_review_ASCII_value.column_dimensions['F'].alignment = alignment

    k = 5
    o = 4
    j = 5
    count_length_byte = 0
    count_Value_DPT = 0
    count_DID_In_DPT = 0
    length_byte = []
    Value_DPT = []
    DID_In_DPT = []
    load_second = 0
    # print(row_read_DPT)
    
    while k < row_count_read_DPT:
        if str(Column_DID) == '':
            for row in range(o, j):
                for col in range(1, 2):
                    char = get_column_letter(col)
                    row_list_DID_In_DPT = ws_read_DPT[char + str(row)].value
                    # print(row_list_DID)
        else:
            row_list_DID_In_DPT = ws_read_DPT[Column_DID + str(line_DID)].value

        if str(Column_Value) == '':        
            for row in range(o, j):
                for col in range(21, 22):
                    char = get_column_letter(col)
                    row_list_value = ws_read_DPT[char + str(row)].value
                    # print(row_list_value)
        else:
            row_list_value = ws_read_DPT[Column_Value + str(line_Value)].value

        
        if str(row_list_DID_In_DPT) != 'None':
            DID_In_DPT = DID_In_DPT + [row_list_DID_In_DPT]
            count_DID_In_DPT +=  1

        if str(row_list_value) != 'None':
            Value_DPT = Value_DPT + [row_list_value]
            count_Value_DPT +=  1
        
        o += 1
        j += 1
        k += 1
        line_DID = int(line_DID) + 1
        line_Value = int(line_Value) + 1


    
    id = 1
    r = 2
    with open(url, 'r') as f:

        ws_report_review_ASCII_value.append(['ID' , 'DID','RECEIVED', 'EXPECTED','DPT', 'RESULT'])
        
        for col in range(1, 7):
            cell_header = ws_report_review_ASCII_value.cell(1, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='000066CC', end_color='000066CC', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
            cell_header.alignment = alignment

        result_of_string = f.read()
        
        result_of_string = result_of_string.partition("EXPECTED: 62")[2]
        # print(len(result_of_string))
        while len(result_of_string) > 0:

            id += 1
            # print(len(result_of_string))
            percent.set(str((id//count_DID_In_DPT)*100)+"%")
            length_of_EXPECTED = result_of_string.find('<')
            # print(length_of_EXPECTED)
            

            result_of_EXPECTED = result_of_string
            result_of_EXPECTED = result_of_EXPECTED.split('<', 1)[0]

            value_of_EXPECTED = result_of_EXPECTED[4:length_of_EXPECTED]
            DID_of_EXPECTED = result_of_EXPECTED[0:4]

            # print(value_of_EXPECTED)

            try:
                bytes_object_of_EXPECTED = bytes.fromhex(value_of_EXPECTED)
                
            except:
                missing_lenght_byte_of_EXPECTED = value_of_EXPECTED.partition(".{")[2]
                missing_lenght_byte_of_EXPECTED = missing_lenght_byte_of_EXPECTED[:-1]

                length_of_EXPECTED = int(length_of_EXPECTED) + int(missing_lenght_byte_of_EXPECTED) - 4

                value_of_EXPECTED = value_of_EXPECTED.split('.', 1)[0]

                bytes_object_of_EXPECTED = bytes.fromhex(value_of_EXPECTED)

                ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")

                # Convert to bytes object
            try:
                ascii_string_of_EXPECTED = bytes_object_of_EXPECTED.decode("ASCII")

            # Convert to ASCII representation
            except:
                print('khong')
            if str(value_of_EXPECTED) != '':
                length_of_EXPECTED = (length_of_EXPECTED - 4) // 2
            else:
                length_of_EXPECTED = (length_of_EXPECTED // 2) - 2

            result_of_string = result_of_string.partition("RECEIVED: 62")[2] 
            # print(result_of_string)

            ### This is the part that measures the length before '<'
            length_of_RECEIVED = result_of_string.find('<')
            # print(length)
            

            result_of_RECEIVED = result_of_string
            result_of_RECEIVED = result_of_RECEIVED.split('<', 1)[0]

            value_of_RECEIVED = result_of_RECEIVED[4:length_of_RECEIVED]
            DID_of_RECEIVED = result_of_RECEIVED[0:4]

            bytes_object_of_RECEIVED = bytes.fromhex(value_of_RECEIVED)
                # Convert to bytes object

            length_of_RECEIVED = (length_of_RECEIVED - 4) // 2

            try:
                ascii_string_of_RECEIVED = bytes_object_of_RECEIVED.decode("ASCII")
                # print(ascii_string_of_RECEIVED)
                ascii_string_of_RECEIVED = ILLEGAL_CHARACTERS_RE.sub(r'', ascii_string_of_RECEIVED)
                # print(ascii_string_of_RECEIVED)
                
            # Convert to ASCII representation
            except:
                ascii_string_of_RECEIVED = ILLEGAL_CHARACTERS_RE.sub(r'', ascii_string_of_RECEIVED)
                # print(ascii_string_of_RECEIVED)


            result_of_string = result_of_string.partition("EXPECTED: 62")[2]

            try:
                ws_report_review_ASCII_value.append(['ID_'+str(id) , str(DID_of_EXPECTED),str(ascii_string_of_RECEIVED), str(ascii_string_of_EXPECTED),'' , '', '','',''])
            except:
                ws_report_review_ASCII_value.append(['ID_'+str(id) , str(DID_of_EXPECTED),'', '' , '', '','',''])
            
            r += 1

    data_validation_data = '"PASS, FAIL"'
    
    for row in range(2, row_count_read_DPT):
        data_validation = DataValidation(type='list', formula1 = data_validation_data)
        ws_report_review_ASCII_value.add_data_validation(data_validation)
        data_validation.add(ws_report_review_ASCII_value['K'+str(row)])
    
    locate_save = Output_path_text.get()
    if locate_save == '':
            # print('Ok')
        wb_report_review_ASCII_value.save("Report_review.xlsx")
    else:
        wb_report_review_ASCII_value.save(locate_save + '/'+'Report_review.xlsx')
    
    wb_report_review_ASCII_value.close()

    l = 2
    h = 0
    
    wb_report_review_ASCII_value_add_dpt_value = load_workbook('Report_review.xlsx')
    ws_report_review_ASCII_value_add_dpt_value = wb_report_review_ASCII_value_add_dpt_value.active
    ws_report_review_ASCII_value_add_dpt_value = wb_report_review_ASCII_value_add_dpt_value['ASCII Value']

    sheet_report_review_add_dpt_value = wb_report_review_ASCII_value_add_dpt_value.worksheets[3]
    row_count_report_review = sheet_report_review_add_dpt_value.max_row

    while l < row_count_report_review:
        ws_report_review_ASCII_value_add_dpt_value['E'+ str(l)] = str(Value_DPT[h])

        l += 1

        ws_report_review_ASCII_value_add_dpt_value['E'+ str(l)] = str(Value_DPT[h])
        
        l += 1
        h += 1

    if locate_save == '':
        wb_report_review_ASCII_value_add_dpt_value.save("Report_review.xlsx")
    else:
        wb_report_review_ASCII_value_add_dpt_value.save(locate_save + '/'+'Report_review.xlsx')

    
    wb_report_review_ASCII_value_add_dpt_value.close()


    wb_report_review_ASCII_value_add_result = load_workbook('Report_review.xlsx')
    ws_report_review_ASCII_value_add_result = wb_report_review_ASCII_value_add_result.active
    ws_report_review_ASCII_value_add_result = wb_report_review_ASCII_value_add_result['ASCII Value']

    sheet_report_review_add_result = wb_report_review_ASCII_value_add_result.worksheets[3]
    row_count_report_review_add_result = sheet_report_review_add_result.max_row

    wb_report_review_Length_Byte_add_result = load_workbook('Report_review.xlsx')
    ws_report_review_Lenght_Byte_add_result = wb_report_review_Length_Byte_add_result.active
    ws_report_review_Lenght_Byte_add_result = wb_report_review_Length_Byte_add_result['Length Byte']

    sheet_report_review_add_result = wb_report_review_Length_Byte_add_result.worksheets[1]
    row_count_report_review_add_result = sheet_report_review_add_result.max_row

    wb_report_review_DID_Compare_add_result = load_workbook('Report_review.xlsx')
    ws_report_review_DID_Compare_add_result = wb_report_review_DID_Compare_add_result.active
    ws_report_review_DID_Compare_add_result = wb_report_review_DID_Compare_add_result['DID']

    sheet_report_review_add_result = wb_report_review_DID_Compare_add_result.worksheets[0]
    row_count_report_review_add_result = sheet_report_review_add_result.max_row


    z = 2
    m = 2
    n = 3
    v = 2
    while z < row_count_report_review_add_result + 1:

        for row in range(m, n):
            for col in range(3, 4):
                char = get_column_letter(col)
                ascii_string_of_RECEIVED_report = ws_report_review_ASCII_value_add_result[char + str(row)].value
                # print(ascii_string_of_RECEIVED)

        
        for row in range(m, n):
            for col in range(4, 5):
                char = get_column_letter(col)
                ascii_string_of_EXPECTED_report = ws_report_review_ASCII_value_add_result[char + str(row)].value
                # print(ascii_string_of_EXPECTED)

        for row in range(m, n):
            for col in range(5, 6):
                char = get_column_letter(col)
                ascii_string_of_DPT_report = ws_report_review_ASCII_value_add_result[char + str(row)].value
                # print(ascii_string_of_EXPECTED)

        for row in range(m, n):
            for col in range(2, 3):
                char = get_column_letter(col)
                DID_of_RECEIVED_report = ws_report_review_DID_Compare_add_result[char + str(row)].value
                # print(DID_of_RECEIVED_report)
                

        for row in range(m, n):
            for col in range(3, 4):
                char = get_column_letter(col)
                DID_of_EXPECTED_report = ws_report_review_DID_Compare_add_result[char + str(row)].value
                # print(DID_of_EXPECTED)
            
        for row in range(m, n):
            for col in range(4, 5):
                char = get_column_letter(col)
                DID_of_DID_DPT_report = ws_report_review_DID_Compare_add_result[char + str(row)].value

        for row in range(m, n):
            for col in range(3, 4):
                char = get_column_letter(col)
                length_of_RECEIVED_report = ws_report_review_Lenght_Byte_add_result[char + str(row)].value
                # print(length_of_RECEIVED_report)
                

        for row in range(m, n):
            for col in range(4, 5):
                char = get_column_letter(col)
                length_of_EXPECTED_report = ws_report_review_Lenght_Byte_add_result[char + str(row)].value
                # print(length_of_EXPECTED)
            
        for row in range(m, n):
            for col in range(5, 6):
                char = get_column_letter(col)
                length_of_DID_DPT_report = ws_report_review_Lenght_Byte_add_result[char + str(row)].value
                # print(length_of_DID_DPT_report)

        
        
        if str(ascii_string_of_EXPECTED_report) == str(ascii_string_of_RECEIVED_report) and str(ascii_string_of_EXPECTED_report) == str(ascii_string_of_DPT_report):
            # print('pass hex value')
            result_ascii_value = 'PASS'
        else:
            if str(ascii_string_of_EXPECTED_report) == str(ascii_string_of_RECEIVED_report):
                if str(length_of_EXPECTED_report) == str(length_of_RECEIVED_report) and str(length_of_EXPECTED_report) == str(length_of_DID_DPT_report):
                    if str(DID_of_EXPECTED_report) == str(DID_of_RECEIVED_report) and str(DID_of_EXPECTED_report) == str(DID_of_DID_DPT_report):
                        result_ascii_value = 'PASS'
                    else:
                        result_ascii_value = 'FAIL'
                else:
                    result_ascii_value = 'FAIL'
            else:
                result_ascii_value = 'FAIL'


        ws_report_review_ASCII_value_add_result['F'+ str(z)] = str(result_ascii_value)

        if  str(result_ascii_value) == 'FAIL':
            for col in range(2, 7):
                cell_header_result_ascii_value = ws_report_review_ASCII_value_add_result.cell(v, col)
            # used hex code for red color
                cell_header_result_ascii_value.fill = PatternFill(
                    start_color='00FF0000', end_color='00FF0000', fill_type="solid")
                cell_header_result_ascii_value.border = border
                cell_header_result_ascii_value.font = font_text3
                cell_header_result_ascii_value.alignment = alignment

        else:
            for col in range(6, 7):
                cell_header_result_ascii_value = ws_report_review_ASCII_value_add_result.cell(v, col)
            # used hex code for red color
                cell_header_result_ascii_value.fill = PatternFill(
                    start_color='0099CC00', end_color='0099CC00', fill_type="solid")
                cell_header_result_ascii_value.border = border
                cell_header_result_ascii_value.font = font_text3
                cell_header_result_ascii_value.alignment = alignment

        n += 1
        m += 1
        v += 1
        z += 1

    if locate_save == '':
        wb_report_review_ASCII_value_add_result.save("Report_review.xlsx")
    else:
        wb_report_review_ASCII_value_add_result.save(locate_save + '/'+'Report_review.xlsx')

    
    wb_report_review_ASCII_value_add_result.close()

    # messagebox.showinfo("COMPLETE", "File Report_review.xlsx has been created in the same folder tool successfully")
    # print('tao thanh cong')
# review_report(ws_report_review, wb_report_review)


def start_program():
    try:
        os.remove('save_data_input.txt')
    except:
        print('')
    direct_DPT = Input_DPT_path_text.get()
    direct_report = Input_report_path_text.get()
    direct_save = Output_path_text.get()
    Sheet_DPT_Name = Input_Sheet_Name_text.get()
    Column_DID_DPT = Input_Column_DID_text.get()
    Column_Length_Byte_DPT = Input_Column_Length_Byte_text.get()
    Column_Value_DPT = Input_Column_Value_text.get()

    data_input_save =',' + direct_DPT + ',' + direct_report + ',' + direct_save + ',' + Sheet_DPT_Name + ',' + Column_DID_DPT + ',' + Column_Length_Byte_DPT + ',' + Column_Value_DPT
    url = ''
    url = create_DPT_can_load_file(direct_DPT,url)

    with open('save_data_input.txt', 'w') as f:
        f.write(data_input_save)

    if direct_report == '':
        messagebox.showerror("ERROR", "Locate file report fail, Please add locate of report file")
        print('tao thanh cong')
    else:
        url = direct_report

    print(url)
    DID_Compare(ws_report_review, wb_report_review,url)
    Length_Byte(ws_report_review, wb_report_review,url)
    Hex_value(ws_report_review, wb_report_review,url)
    ASCII_Value(ws_report_review, wb_report_review,url)
    os.remove("DPT_can_load.xlsx")
    messagebox.showinfo("COMPLETE", "File Report_review.xlsx has been created in the same folder tool successfully")
    print('tao thanh cong')


# app front end
app = tk.Tk()

app.title('Reviewer Tool')
app.geometry('700x420')

def open_DPT_file():
    browse_input_DPT_path_text.set("loading...")
    file_path = askopenfile(parent=app, mode='rb', title="Choose location take file", filetype=[
                            ("excel file", ".xlsx")])
    print("Original string: " + str(file_path))

    result_str = ""
    final_str = ""
    for i in range(0, len(str(file_path))):
        if i >= 26:
            result_str = result_str + str(file_path)[i]
    reverse_str = result_str[::-1]
    for i in range(0, len(reverse_str)):
        if i >= 2:
            final_str = final_str + reverse_str[i]
    complete_str = final_str[::-1]
    print(type(complete_str))
    print(complete_str)
    if file_path:
        Input_DPT_path_text.set(str(complete_str))
        browse_input_DPT_path_text.set("Browse")
    return complete_str

def open_report_file():
    browse_input_report_path_text.set("loading...")
    file_path2 = askopenfile(parent=app, mode='rb', title="Choose location take file", filetype=[
                            ("excel file", ".html")])
    print("Original string: " + str(file_path2))

    result_str = ""
    final_str = ""
    for i in range(0, len(str(file_path2))):
        if i >= 26:
            result_str = result_str + str(file_path2)[i]
    reverse_str = result_str[::-1]
    for i in range(0, len(reverse_str)):
        if i >= 2:
            final_str = final_str + reverse_str[i]
    complete_str = final_str[::-1]
    print(type(complete_str))
    print(complete_str)
    if file_path2:
        Input_report_path_text.set(str(complete_str))
        browse_input_report_path_text.set("Browse")
    return complete_str

def save_file():
    # print("is this working??")
    browse_output_path_text.set("loading...")
    file_path3 = filedialog.askdirectory()
    print(file_path3)
    if file_path3:
        Output_path_text.set(str(file_path3))
        browse_output_path_text.set("Browse")
    return str(file_path3)


file_path = ""

# Part Base SW
frameall = tk.Frame(app)
frame1 = tk.Frame(frameall)
frame2 = tk.Frame(frameall)
frame3 = tk.Frame(frameall)
    
    

Input_DPT_path_text = tk.StringVar()
Input_DPT_path_label = tk.Label(frame1, text='Input DPT file path', font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=0, column=0, sticky='w')
Input_DPT_path_entry = tk.Entry(frame1, textvariable=Input_DPT_path_text,
                            font='large_font', width=55).grid(row=1, column=0, sticky='w')

Input_report_path_text = tk.StringVar()
Input_report_path_label = tk.Label(frame1, text='Input report file path', font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=2, column=0, sticky='w')
Input_report_path_entry = tk.Entry(frame1, textvariable=Input_report_path_text,
                            font='large_font', width=55).grid(row=3, column=0, sticky='w')

Output_path_text = tk.StringVar()
Output_path_label = tk.Label(frame1, text='Output path', font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=4, column=0, sticky='w')
Output_path_entry = tk.Entry(frame1, textvariable=Output_path_text,
                            font='large_font', width=55).grid(row=5, column=0, sticky='w')

# browse button open file
browse_input_DPT_path_text = tk.StringVar()
browse_btn_input_DPT_path = tk.Button(frame1, textvariable=browse_input_DPT_path_text, command=lambda: open_DPT_file(
), font="bold", width=7, height=1).grid(row=1, column=1, pady=5, padx=10)
browse_input_DPT_path_text.set("Browse")

browse_input_report_path_text = tk.StringVar()
browse_btn_input_report_path = tk.Button(frame1, textvariable=browse_input_report_path_text, command=lambda: open_report_file(
), font="bold", width=7, height=1).grid(row=3, column=1, pady=5, padx=10)
browse_input_report_path_text.set("Browse")

# browse button save file
browse_output_path_text = tk.StringVar()
browse_btn_output_path = tk.Button(frame1, textvariable=browse_output_path_text, command=lambda: save_file(
), font="bold", width=7, height=1).grid(row=5, column=1, pady=5, padx=10)
browse_output_path_text.set("Browse")

# Run program Buttons
run_btn_text = tk.StringVar()
run_btn = tk.Button(frame1, textvariable=run_btn_text, command=start_program,
                    font="bold", width=15).grid(row=6, column=0, columnspan=2)
run_btn_text.set("RUN")

percent = tk.StringVar()
percentLabel = tk.Label(frame1, textvariable=percent, font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=7, column=0,columnspan=2)

frame1.pack()

Input_Sheet_Name_text = tk.StringVar()
Input_Sheet_Name_label = tk.Label(frame2, text='Sheet Name', font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=0, column=0,padx = 2, sticky='w')
Input_Sheet_Name_entry = tk.Entry(frame2, textvariable=Input_Sheet_Name_text,
                            font='large_font', width=10).grid(row=1, column=0,padx = 10, sticky='w')

Input_Column_DID_text = tk.StringVar()
Input_Column_DID_label = tk.Label(frame2, text='DID', font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=0, column=1,padx = 38, sticky='w')
Input_Column_DID_entry = tk.Entry(frame2, textvariable=Input_Column_DID_text,
                            font='large_font', width=10).grid(row=1, column=1,padx = 10, sticky='w')

Input_Column_Length_Byte_text = tk.StringVar()
Input_Column_Length_Byte_label = tk.Label(frame2, text='Length_Byte', font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=0, column=2,padx = 2, sticky='w')
Input_Column_Length_Byte_entry = tk.Entry(frame2, textvariable=Input_Column_Length_Byte_text,
                            font='large_font', width=10).grid(row=1, column=2,padx = 10, sticky='w')

Input_Column_Value_text = tk.StringVar()
Input_Column_Value_label = tk.Label(frame2, text='Value', font=(
    'bold', 14), bg="#20bebe", fg="black").grid(row=0, column=3,padx = 30, sticky='w')
Input_Column_Value_entry = tk.Entry(frame2, textvariable=Input_Column_Value_text,
                            font='large_font', width=10).grid(row=1, column=3,padx = 10, sticky='w')

frame2.pack()

# frame1.pack()


frameall.place(relx=0.5, rely=0.5, anchor=CENTER)

# instruction

instruction = tk.Label(
    app, text="    Welcome to Reviewer tool create by dev Huynh Minh Dang", font=("helvetica", 14))
instruction_version = tk.Label(app, text="R1.1.1", font=("helvetica", 14))
instruction_version.pack(side="right", anchor='s')
instruction.pack(side="bottom", fill='both', anchor=CENTER)

frameall.configure(background="#20bebe")
frame1.configure(background="#20bebe")
frame2.configure(background="#20bebe")

app.configure(background="#20bebe")


try:
    with open('save_data_input.txt', 'r') as f:
        data_input = f.read()
        # print(data_input)
    
        # DPT
        data_input = data_input.partition(",")[2] 
        direct_DPT = data_input
        direct_DPT = direct_DPT.split(',', 1)[0]
        Input_DPT_path_text.set(str(direct_DPT))
    
        # report
        data_input = data_input.partition(",")[2] 
        direct_report = data_input
        direct_report = direct_report.split(',', 1)[0]
        Input_report_path_text.set(direct_report)

        # Save
        data_input = data_input.partition(",")[2] 
        direct_save = data_input
        direct_save = direct_save.split(',', 1)[0]
        Output_path_text.set(direct_save)

        # sheet name
        data_input = data_input.partition(",")[2] 
        Sheet_DPT_Name = data_input
        Sheet_DPT_Name = Sheet_DPT_Name.split(',', 1)[0]
        Input_Sheet_Name_text.set(Sheet_DPT_Name)

        # DID
        data_input = data_input.partition(",")[2] 
        Column_DID_DPT = data_input
        Column_DID_DPT = Column_DID_DPT.split(',', 1)[0]
        Input_Column_DID_text.set(Column_DID_DPT)

        # length bytes
        data_input = data_input.partition(",")[2]
        Column_Length_Byte_DPT = data_input
        Column_Length_Byte_DPT = Column_Length_Byte_DPT.split(',', 1)[0]
        Input_Column_Length_Byte_text.set(Column_Length_Byte_DPT)

        # Value
        data_input = data_input.partition(",")[2]
        Column_Value_DPT = data_input
        Column_Value_DPT = Column_Value_DPT.split(',', 1)[0]
        Input_Column_Value_text.set(Column_Value_DPT)

except:
    print('chua co file data input')

# Start program
app.mainloop()
