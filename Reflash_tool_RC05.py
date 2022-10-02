import xlrd
import matplotlib
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import time


def DID_baseSW(ws2, wb, id, number1, number2, number3, number4):
    baseSW = "CA_CD569ICA_BL03_V4"
    tiket_baseSW = "abc_323"

    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    column_count = sheet.max_column

    i = 0
    o = 2
    j = 3
    # id = 2
    k = 1
    number4 += 1

    count = 0

    # gap doi so dong  de xoa cac du lieu cu
    dbrow = row_count + row_count
    i = 0
    hexvalue = ""
    while k < row_count:
        for row in range(o, j):
            for col in range(1, 2):
                char = get_column_letter(col)
                row_list_DID = ws[char + str(row)].value
                # print(ws[char + str(row)].value)
        for row in range(o, j):
            for col in range(2, 3):
                char = get_column_letter(col)
                row_list_name = ws[char + str(row)].value
                # print(ws[char + str(row)].value)
        for row in range(o, j):
            for col in range(4, 5):
                char = get_column_letter(col)
                row_list_values = ws[char + str(row)].value

                # print(ws[char + str(row)].value)
        id += 1
        hexvalue = ""
        for i in str(row_list_values):
            hexvalue += hex(ord(i))[2:]
        # print(hexvalue)
        ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID) + ' ' + str(row_list_name), 'To check value of the DID ' + str(row_list_DID), '1) Send service 0x22 to the camera for the DID ' +
                    str(row_list_DID) + ' using physical addressing', '1) -', '1) RequestResponse(' + str(row_list_DID) + ','+str(hexvalue) + ', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
        number4 += 1
        o += 1
        j += 1
        k += 1
    return id


def DID_latestSW(ws3, wb, id, number1, number2, number3, number4):
    latestSW = "CA_CD569ICA_BL03_RC05"
    tiket_latestSW = "abc_779"
    
    sheet = wb.worksheets[1]
    row_count2 = sheet.max_row

    i = 0
    o = 2
    j = 3
    k = 1
    number4 += 1

    hexvalue = ""
    
    # gap doi so dong  de xoa cac du lieu cu
    # dbrow = row_count + row_count
    
    i = 0
    hexvalue = ""
    hexvalue = ""
    while k < row_count2:
        for row in range(o, j):
            for col in range(1, 2):
                char = get_column_letter(col)
                row_list_DID = ws3[char + str(row)].value
                # print(ws3[char + str(row)].value)
        for row in range(o, j):
            for col in range(2, 3):
                char = get_column_letter(col)
                row_list_name = ws3[char + str(row)].value
                # print(ws3[char + str(row)].value)
        for row in range(o, j):
            for col in range(4, 5):
                char = get_column_letter(col)
                row_list_values = ws3[char + str(row)].value

                # print(ws3[char + str(row)].value)
        id += 1
        hexvalue = ""
        for i in str(row_list_values):
            hexvalue += hex(ord(i))[2:]
        # print(hexvalue)
        ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID) + ' ' + str(row_list_name), 'To check value of the DID ' + str(row_list_DID), '1) Send service 0x22 to the camera for the DID ' +
                    str(row_list_DID) + ' using physical addressing', '1) -', '1) RequestResponse(' + str(row_list_DID) + ','+str(hexvalue) + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
        number4 += 1
        o += 1
        j += 1
        k += 1
    return id

# -------------------------------------------------------------------------------------------------------------
# base SW to latestSW M3

id = 2
number1 = 1
number2 = 1
number3 = 1
number4 = 1
baseSW = "CA_CD569ICA_BL03_V4"
latestSW = "CA_CD569ICA_BL03_RC05"
tiket_baseSW = "abc_323"
tiket_latestSW = "abc_779"

wb = load_workbook('RFvalue2.xlsx')
wb2 = load_workbook('TC_RF.xlsx')
ws2 = wb2 .active
ws2 = wb2['TC_RF']

ws = wb .active
ws = wb['RFvalue_baseSW']

# script begin

ws2.append(['ID', 'XXX Component',  'Test Description', 'Test Steps',  'Test Response',
            'Teststep keywords', 'ObjectType', 'TestStatus', 'Project', 'TestResult'])
ws2.append(['ID_'+str(id),  '1 REFFLASH', '',
            '', '', '', '', 'Test group', '', ''])

# test case 1 script
id += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + ' Base SW to Latest SW M3',
            '', '', '', '', '', 'Test group', '', ''])

# # step 1 script

id += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + 'Flash Base SW via UART',
            '', '', '', '', '', 'Test group', '', ''])
id += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Flash SW', '',
            '', '', '', '', 'Test group', '', ''])
id += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART flash'+baseSW, 'Detail information is mentioned in the ticket: '+tiket_baseSW,
            'Screen shot the successful flash procress', 'Screen shot the successful flash procress', 'Manual Testcase', 'Manual Testcase', 'implemented', baseSW, ''])
id += 1
number3 += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
            '', '', '', '', '', 'Test group', '', ''])
id += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Select variant\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
            '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(2e014001, 6e0140, Equal)\n7) RequestResponse(22F1F0, 62014001, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])
# id += 1
# number4 += 1
id = DID_baseSW(ws2, wb, id, number1, number2, number3, number4)
# id = DID_baseSW(ws2, wb, id, number1, number2, number3, number4)
print(id)

# programing couter
id += 1
number3 += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
            '', '', '', '', '', 'Test group', '', ''])
print(id)
id += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
            '1) -', '1) RequestResponse(220200, 620200.{3}0, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
id += 1
number4 += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
            '1) -', '1) RequestResponse(220201, 620201.{7}0, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])


# RBEOL read
id += 1
number3 += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
            '', '', '', '', '', 'Test group', '', ''])
id += 1
number3 += 1
number4 = 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
            '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.{5}3 ,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
id += 1
number4 += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
            '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
id += 1
number4 += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
            '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
id += 1
number4 += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
            '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

# ws2.title = "TC_RF"
# wb2.save('TC_RF.xlsx')

# Step2

# # Reflash latest SW M3 via xflash tool

ws3 = wb .active
ws3 = wb['RFvalue_latestSW']

# number = df.shape[0]
# print(number)
sheet = wb.worksheets[1]
row_count = sheet.max_row

id += 1
number2 += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Latest SW M3 via X-Flash 1st',
            '', '', '', '', '', 'Test group', '', ''])
id += 1
number3 = 1
number4 = 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash '+latestSW, 'Detail information is mentioned in the ticket: '+tiket_latestSW,
            'Screen shot the successful flash procress', 'Screen shot the successful flash procress', 'Manual Testcase', 'Manual Testcase', 'implemented', latestSW, ''])
id += 1
number3 += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
            '', '', '', '', '', 'Test group', '', ''])
id += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Select variant\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
            '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(2e014001, 6e0140, Equal)\n7) RequestResponse(22F1F0, 62014001, Equal)', 'Automated Testcase', 'implemented', latestSW, ''])

# DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
print(id)

# programing couter
id += 1
number3 += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
            '', '', '', '', '', 'Test group', '', ''])
print(id)
id += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
            '1) -', '1) RequestResponse(220200, 620200.{3}0, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
id += 1
number4 += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
            '1) -', '1) RequestResponse(220201, 620201.{7}0, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])


# RBEOL read
id += 1
number3 += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
            '', '', '', '', '', 'Test group', '', ''])
id += 1
number3 += 1
number4 = 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
            '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.{5}3 ,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
id += 1
number4 += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
            '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
id += 1
number4 += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
            '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
id += 1
number4 += 1
ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
            '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

ws2.title = "TC_RF"
wb2.save('TC_RF.xlsx')


# # ws2.title = "TC_RF"
# # wb2.save('TC_RF.xlsx')

# # --------------------------------------------------------------------------------------------------------
# # latest SW script
# # n = 1
# # wb = load_workbook('RFvalue2.xlsx')
# # wb2 = load_workbook('TC_RF.xlsx')
# # ws2 = wb2 .active
# # ws2 = wb2['TC_RF']
# # sheets = wb.sheetnames
# # ws = wb[sheets[n]]
# ws3 = wb .active
# ws3 = wb['RFvalue_latestSW']

# # number = df.shape[0]
# # print(number)
# sheet = wb.worksheets[1]
# row_count = sheet.max_row
# column_count = sheet.max_column

# # base SW script
# # chon sheet
# # n = 0
# i = 0
# o = 2
# j = 3
# id = 2
# number = 1
# count = 0
# baseSW = "CA_CD569ICA_BL03_V4"
# latestSW = "CA_CD569ICA_BL03_RC05"
# tiket_baseSW = "abc_323"
# tiket_latestSW = "abc_779"
# # ws3 = wb[sheets[n]]
# # wb = Workbook()
# # ws3 = wb .active
# # ws3.title = "RFvalue_baseSW"
# # print(wb.sheetnames[0])
# # print(wb.sheetnames[1])
# # ws3['A2'].value = "Test"


# ws2.append(['ID', 'XXX Component',  'Test Description', 'Test Steps',  'Test Response',
#             'Teststep keywords', 'ObjectType', 'TestStatus', 'Project', 'TestResult'])
# ws2.append(['ID_'+str(id),  '1 REFFLASH', '',
#             '', '', '', '', 'Test group', '', ''])
# id += 1
# # base SW script
# ws2.append(['ID_'+str(id),  '1.1 Latest SW to Latest SW M3',
#             '', '', '', '', '', 'Test group', '', ''])
# id += 1
# ws2.append(['ID_'+str(id),  '1.1.1 Flash Latest SW via UART',
#             '', '', '', '', '', 'Test group', '', ''])
# id += 1
# ws2.append(['ID_'+str(id),  '1.1.1.1 Flash SW', '',
#             '', '', '', '', 'Test group', '', ''])
# id += 1
# ws2.append(['ID_'+str(id),  '1.1.1.1.1 UART flash'+latestSW, 'Detail information is mentioned in the ticket: '+tiket_latestSW,
#             'Screen shot the successful flash procress', 'Screen shot the successful flash procress', 'Manual Testcase', 'Manual Testcase', 'implemented', latestSW, ''])
# id += 1
# ws2.append(['ID_'+str(id),  '1.1.1.2 Variant and Software  Identification',
#             '', '', '', '', '', 'Test group', '', ''])
# id += 1
# ws2.append(['ID_'+str(id),  '1.1.1.2.1 Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Select variant\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
#             '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(2e014001, 6e0140, Equal)\n7) RequestResponse(22F1F0, 62014001, Equal)', 'Automated Testcase', 'implemented', latestSW, ''])

# # book = openpyxl.open_workbook("RFvalue2.xlsx")
# # values = []
# # for sheet in row_count:
# #     print(wb.sheetnames)
# # header = file.readline()
# # row = file.readline()
# print(row_count)
# k = 1
# # gap doi so dong  de xoa cac du lieu cu
# dbrow = row_count + row_count
# i = 0
# hexvalue = ""
# while k < row_count:
#     for row in range(o, j):
#         for col in range(1, 2):
#             char = get_column_letter(col)
#             row_list_DID = ws3[char + str(row)].value
#             # print(ws3[char + str(row)].value)
#     for row in range(o, j):
#         for col in range(2, 3):
#             char = get_column_letter(col)
#             row_list_name = ws3[char + str(row)].value
#             # print(ws3[char + str(row)].value)
#     for row in range(o, j):
#         for col in range(4, 5):
#             char = get_column_letter(col)
#             row_list_values = ws3[char + str(row)].value

#             # print(ws3[char + str(row)].value)
#     id += 1
#     hexvalue = ""
#     for i in str(row_list_values):
#         hexvalue += hex(ord(i))[2:]
#     print(hexvalue)
#     ws2.append(['ID_'+str(id),  '1.1.1.2.' + str(number) + ' ' + str(row_list_DID) + ' ' + str(row_list_name), 'To check value of the DID ' + str(row_list_DID), '1) Send service 0x22 to the camera for the DID ' +
#                 str(row_list_DID) + ' using physical addressing', '1) -', '1) RequestResponse(' + str(row_list_DID) + ','+str(hexvalue) + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
#     number += 1
#     o += 1
#     j += 1
#     k += 1
# # programing couter
# id += 1
# ws2.append(['ID_'+str(id),  '1.1.1.3 Programming Counter and Programming Attempt Counter',
#             '', '', '', '', '', 'Test group', '', ''])
# id += 1
# ws2.append(['ID_'+str(id),  '1.1.1.3.1 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
#             '1) -', '1) RequestResponse(220200, 620200.{3}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
# id += 1
# ws2.append(['ID_'+str(id),  '1.1.1.3.1 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
#             '1) -', '1) RequestResponse(220201, 620201.{7}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
# # RBEOL read
# id += 1
# ws2.append(['ID_'+str(id),  '1.1.1.4 DID in RBEOL',
#             '', '', '', '', '', 'Test group', '', ''])
# id += 1
# ws2.append(['ID_'+str(id),  '1.1.1.4.1 F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
#             '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.{5}3 ,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
# id += 1
# ws2.append(['ID_'+str(id),  '1.1.1.4.2 F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
#             '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
# id += 1
# ws2.append(['ID_'+str(id),  '1.1.1.4.2 4255', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
#             '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
# id += 1
# ws2.append(['ID_'+str(id),  '1.1.1.4.2 4259', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
#             '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

# ws2.title = "TC_RF"
# wb2.save('TC_RF.xlsx')
