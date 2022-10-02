from email import message
import xlrd
import matplotlib
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import time
import os , sys
from struct import pack
import tkinter as tk
import PyPDF2
from PIL import Image, ImageTk
from tkinter.filedialog import askopenfile
from tkinter import HORIZONTAL, messagebox
from tkinter import ttk
import tkinter.ttk
import sqlite3
import time


try:
    # fd = os.open("TC_RF.xlsx", os.O_RDWR | os.O_CREAT)
    # fd = os.open("TC_RF.xlsx", os.O_RDWR)
    # os.close(fd)
    os.system('TASKKILL /F /IM EXCEL.exe')
    os.remove("TC_RF.xlsx")
    # time.sleep(1)
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "TC_RF"
    wb2.save("TC_RF.xlsx")
    os.close("TC_RF.xlsx")
except:
    try:
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "TC_RF"
        wb2.save("TC_RF.xlsx")
        # os.close("TC_RF.xlsx")
        os.system('TASKKILL /F /IM EXCEL.exe')
    except OSError:
        print('Failed creating the file')
    else:
        print('File created')


def DID_baseSW(ws2, wb, id, number1, number2, number3, number4):
    # baseSW = "CA_CD569ICA_BL03_V4"
    baseSW = Base_SW_entry.get()
    tiket_baseSW = Ticket_BaseSW_entry.get()
    latestSW = Latest_SW_entry.get()
    tiket_latestSW = Ticket_Latest_SW_entry.get()
    tiket_baseSW = "abc_323"
    ws = wb .active
    ws = wb['RFvalue_baseSW']
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    column_count = sheet.max_column

    i = 0
    o = 2
    j = 3
    # id = 2
    k = 1
    number4 += 1

    count = 0

    # gap doi so dong  de xoa cac du lieu cu
    dbrow = row_count + row_count
    i = 0
    hexvalue = ""
    while k < row_count:
        for row in range(o, j):
            for col in range(1, 2):
                char = get_column_letter(col)
                row_list_DID = ws[char + str(row)].value
                # print(ws[char + str(row)].value)
        for row in range(o, j):
            for col in range(2, 3):
                char = get_column_letter(col)
                row_list_name = ws[char + str(row)].value
                # print(ws[char + str(row)].value)
        for row in range(o, j):
            for col in range(2, 3):
                char = get_column_letter(col)
                row_list_length_byte = ws[char + str(row)].value
                # print(ws[char + str(row)].value)
        for row in range(o, j):
            for col in range(4, 5):
                char = get_column_letter(col)
                row_list_values = ws[char + str(row)].value

                # print(ws[char + str(row)].value)
        id += 1
        

        hexvalue = ""
        for i in str(row_list_values):
            hexvalue += hex(ord(i))[2:]
        # print(hexvalue)
        ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID) + ' ' + str(row_list_name), 'To check value of the DID ' + str(row_list_DID), '1) Send service 0x22 to the camera for the DID ' +
                    str(row_list_DID) + ' using physical addressing', '1) -', '1) RequestResponse(' +'22'+ str(row_list_DID) + ','+'62'+str(row_list_DID)+str(hexvalue) + ', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
        number4 += 1
        o += 1
        j += 1
        k += 1
    return id


def DID_latestSW(ws3, wb, id, number1, number2, number3, number4):
    # latestSW = "CA_CD569ICA_BL03_RC05"
    baseSW = Base_SW_entry.get()
    tiket_baseSW = Ticket_BaseSW_entry.get()
    latestSW = Latest_SW_entry.get()
    tiket_latestSW = Ticket_Latest_SW_entry.get()
    tiket_latestSW = "abc_779"
    ws3 = wb .active
    ws3 = wb['RFvalue_latestSW']
    sheet = wb.worksheets[1]
    row_count2 = sheet.max_row

    i = 0
    o = 2
    j = 3
    k = 1
    number4 += 1

    hexvalue = ""

    # gap doi so dong  de xoa cac du lieu cu
    # dbrow = row_count + row_count

    i = 0
    hexvalue = ""
    hexvalue = ""
    while k < row_count2:
        for row in range(o, j):
            for col in range(1, 2):
                char = get_column_letter(col)
                row_list_DID = ws3[char + str(row)].value
                # print(ws3[char + str(row)].value)
        for row in range(o, j):
            for col in range(2, 3):
                char = get_column_letter(col)
                row_list_name = ws3[char + str(row)].value
                # print(ws3[char + str(row)].value)
        for row in range(o, j):
            for col in range(4, 5):
                char = get_column_letter(col)
                row_list_values = ws3[char + str(row)].value

                # print(ws3[char + str(row)].value)
        id += 1
        

        hexvalue = ""
        for i in str(row_list_values):
            hexvalue += hex(ord(i))[2:]
        # print(hexvalue)
        ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID) + ' ' + str(row_list_name), 'To check value of the DID ' + str(row_list_DID), '1) Send service 0x22 to the camera for the DID ' +
                    str(row_list_DID) + ' using physical addressing', '1) -', '1) RequestResponse(' +'22' + str(row_list_DID) + ','+'62'+str(row_list_DID)+str(hexvalue) + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
        number4 += 1
        o += 1
        j += 1
        k += 1
    return id


# def procress(ws, ws3, wb, id):
#     ws3 = wb .active
#     ws3 = wb['RFvalue_latestSW']
#     sheet = wb.worksheets[1]
#     ws = wb .active
#     ws = wb['RFvalue_baseSW']
#     sheet = wb.worksheets[0]
#     row_count = sheet.max_row
#     column_count = sheet.max_column
#     row_count2 = sheet.max_row
#     x = 0
#     tasks = row_count2 + row_count
#     bar['value'] += 1
#     percent.set(str((id/tasks)*100)+"%")
#     app.update_idletasks()

# -------------------------------------------------------------------------------------------------------------

# SOURCE

def run_program():
    run_btn_text.set("Loading...")
    baseSW = Base_SW_entry.get()
    tiket_baseSW = Ticket_BaseSW_entry.get()
    latestSW = Latest_SW_entry.get()
    tiket_latestSW = Ticket_Latest_SW_entry.get()
    id = 2
    number1 = 1
    number2 = 1
    number3 = 1
    number4 = 1
    # baseSW = "CA_CD569ICA_BL03_V4"
    # latestSW = "CA_CD569ICA_BL03_RC05"
    # tiket_baseSW = "abc_323"
    # tiket_latestSW = "abc_779"

    wb = load_workbook('RFvalue2.xlsx')
    wb2 = load_workbook('TC_RF.xlsx')
    ws2 = wb2 .active
    ws2 = wb2['TC_RF']

    ws = wb .active
    ws = wb['RFvalue_baseSW']
    ws3 = wb['RFvalue_latestSW']
    sheet2 = wb.worksheets[1]
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    row_count2 = sheet2.max_row
    tasks = row_count2 + row_count + 192
    # print(row_count)
    # print(row_count2)
    
    # script begin
    ws2.append(['ID', 'XXX Component',  'Test Description', 'Test Steps',  'Test Response',
                'Teststep keywords', 'ObjectType', 'TestStatus', 'Project', 'TestResult'])
    ws2.append(['ID_'+str(id),  '1 REFFLASH', '',
                '', '', '', 'Test group', '', '', ''])

    # ------------------------------------------------------------------------------------------------------
    # BEGIN TEST CASE 1
    # TEST CASE 1 base SW to latestSW M3
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + ' Base SW to Latest SW M3',
                '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH BASE_SW VIA UART script
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Flash Base SW via UART',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Flash SW', '',
                '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART flash'+baseSW, 'Detail information is mentioned in the ticket: '+tiket_baseSW,
                'Screen shot the successful flash procress', 'Screen shot the successful flash procress', 'Manual Testcase', 'Manual Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Select variant\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(2e014001, 6e0140, Equal)\n7) RequestResponse(22F1F0, 62014001, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])
    # id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    # number4 += 1
    id = DID_baseSW(ws2, wb, id, number1, number2, number3, number4)
    # id = DID_baseSW(ws2, wb, id, number1, number2, number3, number4)
    print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])
    print(id)
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.{3}0, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.{7}0, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])


    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    # ws2.title = "TC_RF"
    # wb2.save('TC_RF.xlsx')

    # Step2 FLASH LATEST_SW M3 VIA Xflash TOOLS

    ws3 = wb .active
    ws3 = wb['RFvalue_latestSW']

    # number = df.shape[0]
    # print(number)
    sheet = wb.worksheets[1]
    row_count = sheet.max_row
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Latest SW M3 via X-Flash 1st',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash '+latestSW, 'Detail information is mentioned in the ticket: '+tiket_latestSW,
                'Screen shot the successful flash procress', 'Screen shot the successful flash procress', 'Manual Testcase', 'Manual Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Select variant\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(2e014001, 6e0140, Equal)\n7) RequestResponse(22F1F0, 62014001, Equal)', 'Automated Testcase', 'implemented', latestSW, ''])

    # DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
    print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])
    print(id)
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.{3}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.{7}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])


    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # ws2.title = "TC_RF"
    # wb2.save('TC_RF.xlsx')

    # END TEST CASE 1

    # ------------------------------------------------------------------------------------------------------
    # BEGIN TEST CASE 2
    # TEST CASE 2 base SW to latestSW M5
    number1 += 1
    number2 = 1
    number3 = 1
    number4 = 1
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + ' Base SW to Latest SW M5',
                '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH BASE_SW VIA UART script
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Flash Base SW via UART',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Flash SW', '',
                '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART flash'+baseSW, 'Detail information is mentioned in the ticket: '+tiket_baseSW,
                'Screen shot the successful flash procress', 'Screen shot the successful flash procress', 'Manual Testcase', 'Manual Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Select variant\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(2e014001, 6e0140, Equal)\n7) RequestResponse(22F1F0, 62014001, Equal)', 'Automated Testcase', 'implemented', baseSW, ''])
    # id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    # number4 += 1
    id = DID_baseSW(ws2, wb, id, number1, number2, number3, number4)
    # id = DID_baseSW(ws2, wb, id, number1, number2, number3, number4)
    print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])
    print(id)
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.{3}0, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.{7}0, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])


    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    # ws2.title = "TC_RF"
    # wb2.save('TC_RF.xlsx')

    # Step2 FLASH LATEST_SW M5 1st VIA Xflash TOOLS

    # # Reflash latest SW M5 via xflash tool

    ws3 = wb .active
    ws3 = wb['RFvalue_latestSW']

    # number = df.shape[0]
    # print(number)
    sheet = wb.worksheets[1]
    row_count = sheet.max_row
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Latest SW M5 via X-Flash 1st',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash '+latestSW, 'Detail information is mentioned in the ticket: '+tiket_latestSW,
                'Screen shot the successful flash procress', 'Screen shot the successful flash procress', 'Manual Testcase', 'Manual Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Select variant\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(2e014001, 6e0140, Equal)\n7) RequestResponse(22F1F0, 62014001, Equal)', 'Automated Testcase', 'implemented', latestSW, ''])

    # DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
    print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])
    print(id)
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.{3}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.{7}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])


    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*5,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # Step2 FLASH LATEST_SW M3 2nd VIA Xflash TOOLS

    # # Reflash latest SW M3 via xflash tool

    ws3 = wb .active
    ws3 = wb['RFvalue_latestSW']

    # number = df.shape[0]
    # print(number)
    sheet = wb.worksheets[1]
    row_count = sheet.max_row
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Latest SW M3 via X-Flash 2nd',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash '+latestSW, 'Detail information is mentioned in the ticket: '+tiket_latestSW,
                'Screen shot the successful flash procress', 'Screen shot the successful flash procress', 'Manual Testcase', 'Manual Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Select variant\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(2e014001, 6e0140, Equal)\n7) RequestResponse(22F1F0, 62014001, Equal)', 'Automated Testcase', 'implemented', latestSW, ''])

    # DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
    print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])
    print(id)
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.{3}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.{7}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])


    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # ws2.title = "TC_RF"
    # wb2.save('TC_RF.xlsx')

    # END TEST CASE 2

    # ------------------------------------------------------------------------------------------------------

    # BEGIN TEST CASE 3
    # TEST CASE 3 latest SW to DummySW M3
    ws3 = wb .active
    ws3 = wb['RFvalue_latestSW']
    sheet2 = wb.worksheets[1]
    row_count2 = sheet2.max_row
    number1 += 1
    number2 = 1
    number3 = 1
    number4 = 1
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + ' latest SW to Dummy SW M3',
                '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH latest_SW VIA UART script

    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Flash latest SW via UART',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Flash SW', '',
                '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART flash'+latestSW, 'Detail information is mentioned in the ticket: '+tiket_latestSW,
                'Screen shot the successful flash procress', 'Screen shot the successful flash procress', 'Manual Testcase', 'Manual Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Select variant\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(2e014001, 6e0140, Equal)\n7) RequestResponse(22F1F0, 62014001, Equal)', 'Automated Testcase', 'implemented', latestSW, ''])
    # id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    # number4 += 1
    id = DID_latestSW(ws2, wb, id, number1, number2, number3, number4)
    # id = DID_latestSW(ws2, wb, id, number1, number2, number3, number4)
    print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])
    print(id)
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.{3}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.{7}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])


    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # ws2.title = "TC_RF"
    # wb2.save('TC_RF.xlsx')

    # Step2 FLASH DUMMY_SW M3 VIA Xflash TOOLS

    ws3 = wb .active
    ws3 = wb['RFvalue_latestSW']

    # number = df.shape[0]
    # print(number)
    sheet = wb.worksheets[1]
    row_count = sheet.max_row

    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Dummy SW M3 via X-Flash 1st',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash '+latestSW, 'Detail information is mentioned in the ticket: '+tiket_latestSW,
                'Screen shot the successful flash procress', 'Screen shot the successful flash procress', 'Manual Testcase', 'Manual Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Select variant\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(2e014001, 6e0140, Equal)\n7) RequestResponse(22F1F0, 62014001, Equal)', 'Automated Testcase', 'implemented', latestSW, ''])

    # DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
    print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])
    print(id)
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.{3}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.{7}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])


    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # ws2.title = "TC_RF"
    # wb2.save('TC_RF.xlsx')

    # END TEST CASE 3

    # ------------------------------------------------------------------------------------------------------
    # BEGIN TEST CASE 4
    # TEST CASE 4 latest SW to DummySW M5
    number1 += 1
    number2 = 1
    number3 = 1
    number4 = 1
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + ' latest SW to Dummy SW M5',
                '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH latest_SW VIA UART script

    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Flash latest SW via UART',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Flash SW', '',
                '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART flash'+latestSW, 'Detail information is mentioned in the ticket: '+tiket_latestSW,
                'Screen shot the successful flash procress', 'Screen shot the successful flash procress', 'Manual Testcase', 'Manual Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Select variant\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(2e014001, 6e0140, Equal)\n7) RequestResponse(22F1F0, 62014001, Equal)', 'Automated Testcase', 'implemented', latestSW, ''])
    # id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    # number4 += 1
    id = DID_latestSW(ws2, wb, id, number1, number2, number3, number4)
    # id = DID_latestSW(ws2, wb, id, number1, number2, number3, number4)
    print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])
    print(id)
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.{3}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.{7}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])


    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # ws2.title = "TC_RF"
    # wb2.save('TC_RF.xlsx')

    # Step2 FLASH DUMMY_SW M5 1st VIA Xflash TOOLS

    # # Reflash Dummy SW M5 via xflash tool

    ws3 = wb .active
    ws3 = wb['RFvalue_latestSW']

    # number = df.shape[0]
    # print(number)
    sheet = wb.worksheets[1]
    row_count = sheet.max_row

    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Dummy SW M5 via X-Flash 1st',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash '+latestSW, 'Detail information is mentioned in the ticket: '+tiket_latestSW,
                'Screen shot the successful flash procress', 'Screen shot the successful flash procress', 'Manual Testcase', 'Manual Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Select variant\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(2e014001, 6e0140, Equal)\n7) RequestResponse(22F1F0, 62014001, Equal)', 'Automated Testcase', 'implemented', latestSW, ''])

    # DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
    print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])
    print(id)
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.{3}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.{7}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])


    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*5,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # Step2 FLASH DUMMY_SW M3 2nd VIA Xflash TOOLS

    # # Reflash Dummy SW M3 via xflash tool

    ws3 = wb .active
    ws3 = wb['RFvalue_latestSW']

    # number = df.shape[0]
    # print(number)
    sheet = wb.worksheets[1]
    row_count = sheet.max_row

    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Dummy SW M3 via X-Flash 2nd',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash '+latestSW, 'Detail information is mentioned in the ticket: '+tiket_latestSW,
                'Screen shot the successful flash procress', 'Screen shot the successful flash procress', 'Manual Testcase', 'Manual Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Select variant\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(2e014001, 6e0140, Equal)\n7) RequestResponse(22F1F0, 62014001, Equal)', 'Automated Testcase', 'implemented', latestSW, ''])

    # DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
    print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])
    print(id)
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.{3}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.{7}0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])


    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id/tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    ws2.title = "TC_RF"
    wb2.save('TC_RF.xlsx')
    # END TEST CASE 4
    
    run_btn_text.set("DONE")
    tkinter.messagebox.showinfo("GREAT!", "Test case RFlash tool created successfully")


app = tk.Tk()

app.title('Reflash Tool')
# app.iconbitmap('background_img.jpg')
app.geometry('700x420')


# logo
logo = Image.open('background_img.jpg')
logo = ImageTk.PhotoImage(logo)

# Creat canvas

canvas = tk.Canvas(app, width=700, height=420)

canvas.pack(fill='both', expand=True)
canvas.create_image(0, 0, image=logo, anchor='nw')

# instruction
instruction = tk.Label(
    app, text="THIS IS REFLASH TOOL CREATE BY HUYNH MINH DANG", font="Raleway")
Base_SW_entry_window = canvas.create_window(
    130, 390, anchor="nw", window=instruction)


# # Part Base SW
Base_SW_text = tk.StringVar()
Base_SW_label = canvas.create_text(80, 50, text='Base SW Name', font=('bold', 14), fill="Black")
Base_SW_entry = tk.Entry(app, textvariable=Base_SW_text, font='large_font')
Base_SW_entry_window = canvas.create_window(10, 70, anchor="nw", window=Base_SW_entry)


# # Ticket_BaseSW
Ticket_BaseSW_text = tk.StringVar()
Ticket_BaseSW_label = canvas.create_text(400, 50, text='Ticket BaseSW', font=('bold', 14), fill="Black")
Ticket_BaseSW_entry = tk.Entry(app, textvariable=Ticket_BaseSW_text, font='large_font')
Ticket_BaseSW_entry_window = canvas.create_window(325, 70, anchor="nw", window=Ticket_BaseSW_entry)


# # Part Latest SW
Latest_SW_text = tk.StringVar()
Latest_SW_label = canvas.create_text(80, 130, text='Latest SW Name', font=('bold', 14), fill="Black")
Latest_SW_entry = tk.Entry(app, textvariable=Latest_SW_text, font='large_font')
Latest_SW_entry_window = canvas.create_window(10, 150, anchor="nw", window=Latest_SW_entry)


# # Ticket_Latest_SW
Ticket_Latest_SW_text = tk.StringVar()
Ticket_Latest_SW_label = canvas.create_text(400, 130, text='Ticket Latest SW', font=('bold', 14), fill="Black")
Ticket_Latest_SW_entry = tk.Entry(app, textvariable=Ticket_Latest_SW_text, font='large_font')
Ticket_Latest_SW_entry_window = canvas.create_window(325, 150, anchor="nw", window=Ticket_Latest_SW_entry)


# # Buttons
run_btn_text = tk.StringVar()
run_btn = tk.Button(app, textvariable=run_btn_text,command = run_program, font="Raleway", width=15)
run_btn_text.set("RUN")

run_btn_window = canvas.create_window(390, 190, anchor="nw", window=run_btn)



# progress bar
bar = ttk.Progressbar(app, orient=HORIZONTAL, length=600, mode='determinate')
bar_window = canvas.create_window(40, 250, anchor="nw", window=bar)
# barpack(pady = 10)

percent = tk.StringVar()
# percentLabel = tk.Label(app, textvariable = percent).pack()
percentLabel = tk.Label(app, textvariable=percent)
percentLabel_window = canvas.create_window(
    330, 300, anchor="nw", window=percentLabel)


# Start program
app.mainloop()
